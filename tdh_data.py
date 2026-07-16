"""
TDH — data layer PURO (framework-agnostico).

Contiene SOLO calcolo/lettura dati: nessun import di Streamlit, nessun grafico
Plotly. È la fonte di verità dei dati, importabile sia dal cruscotto Streamlit
(`tdhlib`, che espone questi stessi risultati nei grafici) sia dalla futura app
Reflex. Multi-regione by construction (la regione è un codice NUTS2).

Note di portabilità:
- I path dei dati sono ASSOLUTI, radicati sulla cartella di questo file
  (`_ROOT/.cache/...`): funziona a prescindere dalla working directory, quindi
  anche lanciando l'app da un'altra cartella (es. il progetto Reflex).
- La cache di Streamlit (`@st.cache_data`) è sostituita da `functools.lru_cache`,
  che memorizza in-memory dentro il processo (nessuna dipendenza da Streamlit).

Fase 1 della migrazione a Reflex (decisa l'11-07-2026): scorporo del data layer.
"""
from __future__ import annotations

import functools
import importlib.util
import json
import os
import re
import sys

import pandas as pd

# ── radici dei dati (assolute) ──────────────────────────────────────────────
_ROOT = os.path.dirname(os.path.abspath(__file__))
# In locale la cache è `.cache`; nel bundle di deploy è rinominata `datacache`
# (il deploy Reflex pota le cartelle che iniziano con il punto → i dati andrebbero persi).
_CACHE = next((os.path.join(_ROOT, d) for d in (".cache", "datacache")
               if os.path.isdir(os.path.join(_ROOT, d))), os.path.join(_ROOT, ".cache"))

# `regions` è autonomo (solo stdlib): import normale, assicurando _ROOT su sys.path
# (garantisce l'import anche quando l'app gira da un'altra cartella, es. Reflex).
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)
import regions as RG  # noqa: E402


# `real_sources.py` è AUTONOMO (stdlib + numpy/pandas, nessun import relativo): lo
# carichiamo DIRETTAMENTE dal file, così NON scatta `tourism_wedge/__init__.py`, che
# importa `engine` → statsmodels & c. (inutili al data-layer e assenti nel venv Reflex).
def _load_by_path(name: str, path: str):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


RS = _load_by_path("tdh_real_sources", os.path.join(_ROOT, "tourism_wedge", "real_sources.py"))


def _cpath(*parts: str) -> str:
    """Path assoluto dentro la cartella .cache del progetto."""
    return os.path.join(_CACHE, *parts)


# ════════════════════════════════════════════════════════════════════════════
# BANCA D'ITALIA — spesa/notti/viaggiatori dei turisti stranieri
# ════════════════════════════════════════════════════════════════════════════
# colonna del foglio TS2 -> codice/i NUTS2 (Trentino = una colonna -> Bolzano+Trento).
_BDI_REG_COL = {4: ["ITC1"], 5: ["ITC2"], 6: ["ITC4"], 7: ["ITC3"], 9: ["ITD1", "ITD2"],
                10: ["ITD3"], 11: ["ITD4"], 12: ["ITD5"], 14: ["ITE1"], 15: ["ITE2"],
                16: ["ITE3"], 17: ["ITE4"], 19: ["ITF1"], 20: ["ITF2"], 21: ["ITF3"],
                22: ["ITF4"], 23: ["ITF5"], 24: ["ITF6"], 25: ["ITG1"], 26: ["ITG2"]}
_BDI_REG_SHEETS = {"spesa": "TS2-S-S", "notti": "TS2-N-S", "viaggiatori": "TS2-V-S"}

# fogli TS1 (nazionale per PAESE di origine)
_BDI_COLMAP = {4: "DE", 5: "FR", 6: "AT", 7: "ES", 9: "GB", 10: "CH", 11: "RU",
               13: "US", 14: "CA", 17: "JP"}
_BDI_SHEETS = {"notti": "TS1-N-S", "spesa": "TS1-S-S", "viaggiatori": "TS1-V-S"}
BDI_MARKETS = ["DE", "AT", "GB", "CH", "US", "FR", "ES"]

_BDI_XLSX = _cpath("bdi_turismo_ts.xlsx")


@functools.lru_cache(maxsize=None)
def bdi_region_long():
    """Spesa/notti/viaggiatori dei turisti stranieri per REGIONE visitata (trimestrale,
    1997-2025). DataFrame: date·code·spesa·notti·viaggiatori. None se l'xlsx manca."""
    if not os.path.exists(_BDI_XLSX):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(_BDI_XLSX, read_only=True, data_only=True)
    recs = []
    for metric, sheet in _BDI_REG_SHEETS.items():
        year = None
        for r in wb[sheet].iter_rows(values_only=True):
            if r and r[0] is not None:
                try:
                    year = int(r[0])
                except (TypeError, ValueError):
                    pass
            m = re.match(r"\s*(\d)", str((r[1] or r[2]) if r else ""))
            if not (year and m):
                continue
            q = int(m.group(1))
            date = pd.Timestamp(year, q * 3 - 2, 1)
            for col, codes in _BDI_REG_COL.items():
                if col < len(r) and r[col] is not None:
                    try:
                        val = float(r[col])
                    except (TypeError, ValueError):
                        continue
                    for code in codes:
                        recs.append((date, code, metric, val))
    if not recs:
        return None
    df = pd.DataFrame(recs, columns=["date", "code", "metric", "value"])
    return df.pivot_table(index=["date", "code"], columns="metric", values="value").reset_index()


def bdi_national_annual():
    """Spesa straniera ANNUALE per TUTTA Italia = somma delle regioni visitate (BdI).
    Deduplica il Trentino, che nei dati è duplicato su Bolzano (ITD1) e Trento (ITD2)."""
    df = bdi_region_long()
    if df is None or df.empty:
        return None
    reps = {codes[0] for codes in _BDI_REG_COL.values()}  # un codice per colonna-regione (no doppio Trentino)
    d = df[df["code"].isin(reps)].copy()
    if d.empty:
        return None
    qn = d.groupby("date").agg(spesa=("spesa", "sum"), notti=("notti", "sum"),
                               viaggiatori=("viaggiatori", "sum")).reset_index()
    qn["anno"] = qn["date"].dt.year
    return qn.groupby("anno").agg(spesa=("spesa", "sum"), notti=("notti", "sum"),
                                  viaggiatori=("viaggiatori", "sum"),
                                  trimestri=("date", "count")).reset_index()


def bdi_region_annual(code: str):
    """Aggregato ANNUALE (spesa M€, notti/viaggiatori in migliaia) per la regione,
    oppure il totale Italia se è selezionata la vista nazionale."""
    if RG.is_national(code):
        return bdi_national_annual()
    df = bdi_region_long()
    if df is None or df.empty:
        return None
    d = df[df["code"] == code].copy()
    if d.empty:
        return None
    d["anno"] = d["date"].dt.year
    return d.groupby("anno").agg(spesa=("spesa", "sum"), notti=("notti", "sum"),
                                 viaggiatori=("viaggiatori", "sum"),
                                 trimestri=("date", "count")).reset_index()


@functools.lru_cache(maxsize=None)
def bdi_country_long():
    """Flussi nazionali BdI per paese di origine: DataFrame date·code·notti·spesa·viaggiatori
    (trimestrale). Sorgente: fogli TS1 dell'xlsx BdI. None se il file manca."""
    if not os.path.exists(_BDI_XLSX):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(_BDI_XLSX, read_only=True, data_only=True)
    recs = []
    for metric, sheet in _BDI_SHEETS.items():
        rows = list(wb[sheet].iter_rows(values_only=True))
        h = next((i for i, r in enumerate(rows) if r and any(c == "Germania" for c in r if c)), None)
        if h is None:
            continue
        year = None
        for r in rows[h + 1:]:
            if r[0] is not None:
                year = int(r[0])
            m = re.match(r"\s*(\d)", str(r[1] or r[2] or ""))
            if not (year and m):
                continue
            q = int(m.group(1))
            date = pd.Timestamp(year, q * 3 - 2, 1)
            for col, code in _BDI_COLMAP.items():
                if col < len(r) and r[col] is not None:
                    recs.append((date, code, metric, float(r[col])))
    if not recs:
        return None
    df = pd.DataFrame(recs, columns=["date", "code", "metric", "value"])
    return df.pivot_table(index=["date", "code"], columns="metric", values="value").reset_index()


@functools.lru_cache(maxsize=None)
def bdi_extended():
    """Aggregati BdI pre-calcolati (spesa regioni 2024, struttura/motivo nazionale...)."""
    p = _cpath("bdi_extended.json")
    return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else None


# ════════════════════════════════════════════════════════════════════════════
# ISTAT — esteri per PAESE × territorio, ANNUALE (cube DCSC_TUR_9)
# ════════════════════════════════════════════════════════════════════════════
_TUR9_PATH = _cpath("istat_estero_regione_prov_annuale.csv")

# Nomi leggibili dei codici COUNTRY_RES_GUESTS (ISO2 + aggregati). Pass-through se ignoto.
_ISTAT_COUNTRY_NAME = {
    "DE": "Germania", "FR": "Francia", "AT": "Austria", "ES": "Spagna",
    "GB": "Regno Unito", "UK": "Regno Unito", "CH_LI": "Svizzera e Liechtenstein",
    "CH": "Svizzera", "US": "Stati Uniti", "CA": "Canada", "NL": "Paesi Bassi",
    "BE": "Belgio", "RU": "Russia", "CN": "Cina", "JP": "Giappone", "BR": "Brasile",
    "AR": "Argentina", "AU": "Australia", "PL": "Polonia", "SE": "Svezia",
    "DK": "Danimarca", "FI": "Finlandia", "NO": "Norvegia", "IE": "Irlanda",
    "PT": "Portogallo", "GR": "Grecia", "CZ": "Rep. Ceca", "HU": "Ungheria",
    "HR": "Croazia", "SI": "Slovenia", "SK": "Slovacchia", "RO": "Romania",
    "BG": "Bulgaria", "EE": "Estonia", "LT": "Lituania", "LV": "Lettonia",
    "CY": "Cipro", "MT": "Malta", "LU": "Lussemburgo", "IN": "India",
    "KR": "Corea del Sud", "MX": "Messico", "TR": "Turchia", "IL": "Israele",
    "EG": "Egitto", "ZA": "Sudafrica", "NZ": "Nuova Zelanda",
    "WORLD": "Mondo (totale)", "WRL_X_ITA": "Estero (tutti i paesi)",
    "IT": "Italia (residenti)", "EU": "Unione Europea",
    "EUR_NEU": "Altri Europa non-UE", "EUR_OTH": "Altri Europa",
    "AFRMED": "Africa mediterranea", "AFR_OTH": "Altri Africa",
    "AME_N_OTH": "Altri America del Nord", "AME_C_S_OTH": "America centro-sud (altri)",
    "ASI_OTH": "Altri Asia", "ASI_W_OTH": "Asia occidentale (altri)",
    "OCE_OTH": "Altri Oceania",
}


def estero_country_name(code: str) -> str:
    """Nome leggibile del codice paese ISTAT (pass-through se non mappato)."""
    return _ISTAT_COUNTRY_NAME.get(code, code)


def _tur9_is_aggregate(c: str) -> bool:
    """True se il codice è un aggregato (mondo/ripartizioni continentali/UE) o l'Italia,
    non un singolo paese estero. CH_LI (Svizzera+Liechtenstein) è tenuto come mercato."""
    if c in {"WORLD", "WRL_X_ITA", "IT", "EU"}:
        return True
    if c.endswith("_OTH") or c.endswith("_NEU"):
        return True
    return c in {"AFRMED", "EUR", "AFR", "ASI", "AME", "OCE"}


@functools.lru_cache(maxsize=None)
def estero_regione_long():
    """Presenze/arrivi turisti ESTERI per PAESE × territorio (naz/regione/provincia),
    ANNUALE (ISTAT DCSC_TUR_9). DataFrame: area·datatype·country·anno·valore.
    None se la cache manca."""
    if not os.path.exists(_TUR9_PATH):
        return None
    df = pd.read_csv(_TUR9_PATH, dtype={"REF_AREA": str, "DATA_TYPE": str,
                                        "COUNTRY_RES_GUESTS": str})
    if df.empty:
        return None
    df = df.rename(columns={"REF_AREA": "area", "DATA_TYPE": "datatype",
                            "COUNTRY_RES_GUESTS": "country", "TIME_PERIOD": "anno",
                            "OBS_VALUE": "valore"})
    df["anno"] = pd.to_numeric(df["anno"], errors="coerce").astype("Int64")
    df["valore"] = pd.to_numeric(df["valore"], errors="coerce")
    return df.dropna(subset=["anno", "valore"])


def estero_markets(code: str, datatype: str = "NI", year: int | None = None,
                   only_countries: bool = True, top: int | None = None):
    """Classifica dei mercati esteri per la regione (o Italia): DataFrame
    country·nome·valore per l'anno scelto (o l'ultimo disponibile), ordinata desc.
    datatype: 'NI' presenze, 'AR' arrivi. only_countries esclude gli aggregati."""
    df = estero_regione_long()
    if df is None:
        return None
    area = RG.istat_area(code)
    d = df[(df["area"] == area) & (df["datatype"] == datatype)]
    if only_countries:
        d = d[~d["country"].map(_tur9_is_aggregate)]
    if d.empty:
        return None
    if year is None:
        year = int(d["anno"].max())
    d = d[d["anno"] == year].copy()
    if d.empty:
        return None
    d["nome"] = d["country"].map(estero_country_name)
    out = d.sort_values("valore", ascending=False)[["country", "nome", "valore"]].reset_index(drop=True)
    return out.head(top).reset_index(drop=True) if top else out


def estero_years(code: str, datatype: str = "NI"):
    """Anni disponibili per la regione/territorio (lista ordinata di int)."""
    df = estero_regione_long()
    if df is None:
        return []
    area = RG.istat_area(code)
    d = df[(df["area"] == area) & (df["datatype"] == datatype)]
    return sorted(int(a) for a in d["anno"].dropna().unique())


def estero_country_series(code: str, country: str, datatype: str = "NI"):
    """Serie storica annuale di un paese in una regione: DataFrame anno·valore."""
    df = estero_regione_long()
    if df is None:
        return None
    area = RG.istat_area(code)
    d = df[(df["area"] == area) & (df["datatype"] == datatype) & (df["country"] == country)]
    if d.empty:
        return None
    return d.sort_values("anno")[["anno", "valore"]].reset_index(drop=True)


# ════════════════════════════════════════════════════════════════════════════
# ISTAT — presenze mensili e capacità (per regione, tutte le 20)
# ════════════════════════════════════════════════════════════════════════════
def region_overview(code: str) -> dict:
    """Presenze mensili (totale + stranieri) e capacità della regione richiesta.
    Funziona per tutte le 20 regioni (ISTAT per NUTS2). Legge dai CSV di cache."""
    info = RG.region(code)
    area = RG.istat_area(code)
    stran = RS.fetch_istat_presences(area=area, country="WRL_X_ITA", start="2019-01", cache_dir=_CACHE)
    tot = RS.fetch_istat_presences(area=area, country="WORLD", start="2019-01", cache_dir=_CACHE)
    df = (stran.rename(columns={"presences": "stranieri"})
          .merge(tot.rename(columns={"presences": "totale"}), on="date", how="outer")
          .sort_values("date").reset_index(drop=True))
    letti = anno = None
    try:
        cap = RS.fetch_istat_capacity(area=area, cache_dir=_CACHE)
        letti, anno = int(cap["letti"].iloc[-1]), int(cap["anno"].iloc[-1])
    except Exception:  # noqa: BLE001
        pass
    return {"info": info, "presenze": df, "letti": letti, "anno_letti": anno}


# Registro variabili annuali della scheda Regione: chiave · etichetta · unità · decimali.
REGION_VARS = [
    ("presenze_tot", "Presenze totali", "n", 0),
    ("presenze_str", "Presenze straniere", "n", 0),
    ("quota_str", "Quota stranieri", "%", 1),
    ("spesa", "Spesa straniera", "M€", 0),
    ("spesa_per_viagg", "Spesa per viaggiatore", "€", 0),
    ("spesa_per_notte", "Spesa per notte", "€", 0),
    ("viaggiatori", "Viaggiatori stranieri", "migliaia", 0),
    ("notti", "Pernottamenti stranieri", "migliaia", 0),
    ("letti", "Posti letto", "n", 0),
    ("occ", "Occupazione", "%", 1),
]
REGION_VAR_LABEL = {k: lab for k, lab, _u, _d in REGION_VARS}
REGION_VAR_UNIT = {k: u for k, _l, u, _d in REGION_VARS}
REGION_VAR_DEC = {k: d for k, _l, _u, d in REGION_VARS}


@functools.lru_cache(maxsize=None)
def region_annual_panel(code: str) -> pd.DataFrame:
    """Pannello ANNUALE (anno × variabili) per regione o Italia. Unisce ISTAT (presenze
    mensili→somma annua su anni completi, capacità) e Banca d'Italia (spesa/notti/viaggiatori,
    anni completi). Aggiunge le derivate. Le celle mancanti restano NaN."""
    out: dict[int, dict] = {}
    try:
        p = region_overview(code)["presenze"].copy()
        p["anno"] = p["date"].dt.year
        for col, key in (("totale", "presenze_tot"), ("stranieri", "presenze_str")):
            if col in p:
                g = p.dropna(subset=[col]).groupby("anno")[col].agg(["sum", "count"])
                for y, r in g.iterrows():
                    if r["count"] >= 12:
                        out.setdefault(int(y), {})[key] = float(r["sum"])
    except Exception:  # noqa: BLE001
        pass
    g = bdi_region_annual(code)
    if g is not None and not g.empty:
        for _, r in g.iterrows():
            if r["trimestri"] >= 4:
                d = out.setdefault(int(r["anno"]), {})
                d["spesa"], d["notti"], d["viaggiatori"] = float(r["spesa"]), float(r["notti"]), float(r["viaggiatori"])
    try:
        cap = RS.fetch_istat_capacity(area=RG.istat_area(code), cache_dir=_CACHE)
        for _, r in cap.iterrows():
            out.setdefault(int(r["anno"]), {})["letti"] = float(r["letti"])
    except Exception:  # noqa: BLE001
        pass
    if not out:
        return pd.DataFrame()
    df = pd.DataFrame(out).T.sort_index()
    df.index.name = "anno"
    if {"presenze_str", "presenze_tot"} <= set(df.columns):
        df["quota_str"] = df["presenze_str"] / df["presenze_tot"] * 100
    if {"spesa", "viaggiatori"} <= set(df.columns):
        df["spesa_per_viagg"] = df["spesa"] * 1e6 / (df["viaggiatori"] * 1e3)
    if {"spesa", "notti"} <= set(df.columns):
        df["spesa_per_notte"] = df["spesa"] * 1e6 / (df["notti"] * 1e3)
    if {"presenze_tot", "letti"} <= set(df.columns):
        df["occ"] = df["presenze_tot"] / (df["letti"] * 365) * 100
    cols = [k for k, *_ in REGION_VARS if k in df.columns]
    return df[cols]


# ════════════════════════════════════════════════════════════════════════════
# RANKING regioni per spesa straniera (multi-regione by construction)
# ════════════════════════════════════════════════════════════════════════════
@functools.lru_cache(maxsize=None)
def regions_spend_ranking() -> tuple:
    """Classifica delle regioni per spesa turistica straniera 2024 (Banca d'Italia).
    Tupla di dict: code (NUTS2) · regione · spesa_M · rank. (tupla per l'lru_cache)."""
    ext = bdi_extended()
    spese = (ext or {}).get("regioni_2024", {})
    if not spese:
        return ()
    rows = []
    for code, info in RG.REGIONS.items():
        sp = spese.get(info["bdi"])
        if sp is not None:
            rows.append({"code": code, "regione": info["nome"], "spesa_M": float(sp)})
    rows.sort(key=lambda r: r["spesa_M"], reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return tuple(rows)


@functools.lru_cache(maxsize=None)
def bdi_region_years() -> tuple:
    """Anni con i 4 trimestri completi nella serie regionale BdI (per i selettori)."""
    df = bdi_region_long()
    if df is None or df.empty:
        return ()
    g = df.assign(anno=df["date"].dt.year).groupby("anno")["date"].nunique()
    return tuple(sorted(int(y) for y, n in g.items() if n >= 4))


@functools.lru_cache(maxsize=None)
def regions_spend_ranking_year(year: int) -> tuple:
    """Come regions_spend_ranking() ma per l'ANNO indicato (spesa straniera M€, BdI)."""
    df = bdi_region_long()
    if df is None or df.empty:
        return ()
    d = df[df["date"].dt.year == year]
    if d.empty:
        return ()
    by_name: dict[str, float] = {}
    for code, val in d.groupby("code")["spesa"].sum().items():
        info = RG.REGIONS.get(code)
        if info:
            by_name[info["bdi"]] = float(val)
    rows = []
    for code, info in RG.REGIONS.items():
        sp = by_name.get(info["bdi"])
        if sp is not None:
            rows.append({"code": code, "regione": info["nome"], "spesa_M": sp})
    rows.sort(key=lambda r: r["spesa_M"], reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return tuple(rows)


def region_spend(code: str):
    """(spesa_M, rank, totale_regioni) per la regione richiesta, o None.
    Per la vista nazionale: (totale_Italia, None, n_regioni) — rank=None = «Italia»."""
    rk = regions_spend_ranking()
    if RG.is_national(code):
        return (sum(r["spesa_M"] for r in rk), None, len(rk)) if rk else None
    bdi = RG.region(code)["bdi"]
    hit = next((r for r in rk if RG.REGIONS[r["code"]]["bdi"] == bdi), None)
    return (hit["spesa_M"], hit["rank"], len(rk)) if hit else None


# ════════════════════════════════════════════════════════════════════════════
# API DI ALTO LIVELLO per il layer UI (Reflex/Streamlit): tutto in tipi Python
# semplici (str/num/list), niente pandas/plotly. Multi-regione.
# ════════════════════════════════════════════════════════════════════════════
MESI_IT = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]


def _it_int(n) -> str:
    """Intero con separatore delle migliaia all'italiana (1.234.567). '—' se None/NaN."""
    if n is None:
        return "—"
    try:
        return f"{int(round(float(n))):,}".replace(",", ".")
    except (TypeError, ValueError):
        return "—"


def _fmt_val(v, unit: str, dec: int) -> str:
    """Formatta un valore secondo l'unità del registro REGION_VARS (all'italiana)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    if unit == "%":
        return f"{v:.{dec}f}".replace(".", ",") + "%"
    if unit == "M€":
        return f"{_it_int(v)} M€"
    if unit == "€":
        return f"{_it_int(v)} €"
    return _it_int(v)  # n, migliaia


def regione_snapshot(code: str, top_mercati: int = 5) -> dict:
    """Quadro turistico sintetico di una regione (o Italia), in tipi Python puri.
    Pensato per alimentare direttamente lo stato di una UI (Reflex/Streamlit) senza
    che questa conosca pandas. Tutti i campi sono robusti ai dati mancanti."""
    info = RG.region(code)
    ov = region_overview(code)
    pres = ov["presenze"]

    # serie mensile presenze straniere (ultimi 24 mesi disponibili)
    serie_x: list[str] = []
    serie_y: list[float] = []
    ultimo_mese = None
    if pres is not None and not pres.empty and "stranieri" in pres:
        s = pres.dropna(subset=["stranieri"]).tail(24)
        for _, r in s.iterrows():
            d = r["date"]
            serie_x.append(f"{MESI_IT[d.month - 1]} {str(d.year)[2:]}")
            serie_y.append(float(r["stranieri"]))
        if serie_y:
            ultimo_mese = serie_y[-1]

    # spesa e posizione in classifica
    sp = region_spend(code)
    if sp is None:
        spesa_M, rank, n_reg = None, None, None
    else:
        spesa_M, rank, n_reg = sp

    # top mercati esteri (presenze, ultimo anno disponibile)
    mkt = estero_markets(code, top=top_mercati)
    mercati_top = ([] if mkt is None
                   else [{"nome": r["nome"], "valore": float(r["valore"])} for _, r in mkt.iterrows()])
    anni = estero_years(code)
    anno_mkt = anni[-1] if anni else None
    n_mercati_tot = estero_markets(code)
    n_mercati = 0 if n_mercati_tot is None else len(n_mercati_tot)

    if RG.is_national(code):
        rank_label = "Totale Italia"
    elif rank and n_reg:
        rank_label = f"{rank}ª su {n_reg} regioni"
    else:
        rank_label = "—"

    return {
        "code": code,
        "nome": info["nome"],
        "titolo": f"{info['nome']} — quadro turistico",
        # KPI (stringhe formattate, pronte per la UI)
        "kpi_presenze": _it_int(ultimo_mese),
        "kpi_presenze_periodo": (serie_x[-1] if serie_x else "—"),
        "kpi_letti": _it_int(ov["letti"]),
        "kpi_letti_anno": (f"capacità {ov['anno_letti']}" if ov["anno_letti"] else "capacità"),
        "kpi_spesa": (f"{_it_int(spesa_M)} M€" if spesa_M is not None else "—"),
        "kpi_spesa_rank": rank_label,
        "kpi_mercati": str(n_mercati),
        "kpi_mercati_anno": (f"paesi di origine · {anno_mkt}" if anno_mkt else "paesi di origine"),
        # serie e classifiche (dati grezzi per i grafici)
        "serie_mesi": serie_x,
        "serie_presenze": serie_y,
        "mercati_top": mercati_top,
        "mercati_anno": anno_mkt,
    }


def base_dati_snapshot(code: str) -> dict:
    """Pannello ANNUALE della regione in tipi Python puri (per la pagina 'Base dati
    regionale'): serie per i grafici, tabella formattata (anno più recente in alto),
    KPI = ultimo valore disponibile di alcune derivate. Robusto ai dati mancanti."""
    empty = {"years": [], "pres_tot": [], "pres_str": [], "spesa": [], "rows": [], "kpi": {}}
    panel = region_annual_panel(code)
    if panel is None or getattr(panel, "empty", True):
        return empty
    years = [int(y) for y in panel.index.tolist()]

    def col(k):
        if k not in panel.columns:
            return [None] * len(years)
        return [None if pd.isna(v) else float(v) for v in panel[k].tolist()]

    pres_tot, pres_str = col("presenze_tot"), col("presenze_str")
    spesa, quota, occ = col("spesa"), col("quota_str"), col("occ")

    rows = []
    for i, y in enumerate(years):
        rows.append({
            "anno": str(y),
            "pres_tot": _fmt_val(pres_tot[i], "n", 0),
            "pres_str": _fmt_val(pres_str[i], "n", 0),
            "quota": _fmt_val(quota[i], "%", 1),
            "spesa": _fmt_val(spesa[i], "M€", 0),
            "occ": _fmt_val(occ[i], "%", 1),
        })
    rows = rows[::-1]  # anno più recente in alto

    def last_valid(k):
        if k not in panel.columns:
            return (None, None)
        s = panel[k].dropna()
        return (None, None) if s.empty else (int(s.index[-1]), float(s.iloc[-1]))

    kpi = {}
    for k in ("spesa_per_viagg", "spesa_per_notte", "occ", "quota_str"):
        yr, val = last_valid(k)
        kpi[k] = {"val": _fmt_val(val, REGION_VAR_UNIT.get(k, ""), REGION_VAR_DEC.get(k, 0)),
                  "anno": (str(yr) if yr else "—")}

    return {"years": years, "pres_tot": pres_tot, "pres_str": pres_str,
            "spesa": spesa, "rows": rows, "kpi": kpi}


def spesa_snapshot(code: str) -> dict:
    """Spesa turistica STRANIERA (Banca d'Italia) della regione: serie annuali di spesa,
    pernottamenti, viaggiatori + economia PER-VISITATORE (spesa/viaggiatore, spesa/notte,
    permanenza media). Solo anni completi (4 trimestri). Tipi Python puri."""
    empty = {"years": [], "spesa": [], "sp_viagg": [], "sp_notte": [], "permanenza": [],
             "rows": [], "kpi": {}}
    g = bdi_region_annual(code)
    if g is None or getattr(g, "empty", True):
        return empty
    g = g[g["trimestri"] >= 4].sort_values("anno")
    if g.empty:
        return empty
    years = [int(a) for a in g["anno"]]
    spesa = [float(v) for v in g["spesa"]]           # M€
    notti = [float(v) for v in g["notti"]]           # migliaia
    viagg = [float(v) for v in g["viaggiatori"]]     # migliaia
    sp_viagg = [(s * 1e6) / (v * 1e3) if v else None for s, v in zip(spesa, viagg)]   # €/viaggiatore
    sp_notte = [(s * 1e6) / (n * 1e3) if n else None for s, n in zip(spesa, notti)]   # €/notte
    permanenza = [(n / v) if v else None for n, v in zip(notti, viagg)]              # notti/viaggiatore

    rows = []
    for i in range(len(years) - 1, -1, -1):  # anno più recente in alto
        rows.append({
            "anno": str(years[i]),
            "spesa": _fmt_val(spesa[i], "M€", 0),
            "viaggiatori": _it_int(viagg[i] * 1e3),
            "notti": _it_int(notti[i] * 1e3),
            "sp_viagg": _fmt_val(sp_viagg[i], "€", 0),
            "permanenza": (f"{permanenza[i]:.1f}".replace(".", ",") if permanenza[i] is not None else "—"),
        })
    li = len(years) - 1
    kpi = {
        "anno": str(years[li]),
        "spesa": _fmt_val(spesa[li], "M€", 0),
        "sp_viagg": _fmt_val(sp_viagg[li], "€", 0),
        "sp_notte": _fmt_val(sp_notte[li], "€", 0),
        "permanenza": (f"{permanenza[li]:.1f}".replace(".", ",") if permanenza[li] is not None else "—"),
    }
    return {"years": years, "spesa": spesa, "sp_viagg": sp_viagg, "sp_notte": sp_notte,
            "permanenza": permanenza, "rows": rows, "kpi": kpi}


@functools.lru_cache(maxsize=1)
def _prov_map() -> dict:
    """Mappa NUTS2 → {NUTS3: nome} letta a PATH ASSOLUTO (regions._prov_map usa un path
    relativo → vuoto se cwd ≠ TDH_Engine, come nell'app Reflex)."""
    p = os.path.join(_ROOT, "assets", "nuts3_provinces.json")
    return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else {}


def province_snapshot(code: str) -> dict:
    """Presenze per PROVINCIA (NUTS3) della regione, SOLO da cache (nessun fetch live).
    Le province senza CSV in cache vengono conteggiate in `n_mancanti` e saltate.
    Nazionale (ITALIA): has_region=False (le province sono un concetto regionale)."""
    empty = {"has_region": True, "n": 0, "n_mancanti": 0, "rows": [], "bar_nomi": [],
             "bar_val": [], "top_nome": "—", "top_presenze": "—", "top_peso": "—", "share_top3": "—"}
    if RG.is_national(code):
        return {**empty, "has_region": False}
    provs = _prov_map().get(code, {})
    recs, mancanti = [], 0
    for area, nome in provs.items():
        totp = _cpath(f"istat_{area}_NI_ALL_WORLD.csv")
        if not os.path.exists(totp):
            mancanti += 1
            continue
        tot = pd.read_csv(totp, parse_dates=["date"]).sort_values("date")
        t12 = float(tot.tail(12)["presences"].sum())
        e12 = None
        estp = _cpath(f"istat_{area}_NI_ALL_WRL_X_ITA.csv")
        if os.path.exists(estp):
            est = pd.read_csv(estp, parse_dates=["date"]).sort_values("date")
            e12 = float(est.tail(12)["presences"].sum())
        recs.append({"nome": nome, "presenze": t12, "stranieri": e12,
                     "quota": (e12 / t12 * 100 if (t12 and e12 is not None) else None)})
    if not recs:
        return {**empty, "n_mancanti": mancanti}
    recs.sort(key=lambda r: -r["presenze"])
    tot_reg = sum(r["presenze"] for r in recs) or 1.0
    rows = [{"provincia": r["nome"], "presenze": _it_int(r["presenze"]),
             "peso": f"{r['presenze'] / tot_reg * 100:.1f}".replace(".", ",") + "%",
             "stranieri": (_it_int(r["stranieri"]) if r["stranieri"] is not None else "—"),
             "quota": (f"{r['quota']:.0f}".replace(".", ",") + "%" if r["quota"] is not None else "—")}
            for r in recs]
    top = recs[0]
    share3 = sum(r["presenze"] for r in recs[:3]) / tot_reg * 100
    return {"has_region": True, "n": len(recs), "n_mancanti": mancanti, "rows": rows,
            "bar_nomi": [r["nome"] for r in recs], "bar_val": [r["presenze"] for r in recs],
            "top_nome": top["nome"], "top_presenze": _it_int(top["presenze"]),
            "top_peso": f"{top['presenze'] / tot_reg * 100:.0f}%", "share_top3": f"{share3:.0f}%"}


# ════════════════════════════════════════════════════════════════════════════
# ALLOCATORE — "Azioni & budget": ranking mercati esteri (il cuore-motore)
# Usa il MOTORE VERO (tourism_wedge.rank_markets, fonte unica) via import LAZY +
# chdir temporaneo su TDH_Engine (assemble_real usa .cache relativo alla cwd).
# ════════════════════════════════════════════════════════════════════════════
_MK_ISTAT = {"DE": "DE", "AT": "AT", "FR": "FR", "ES": "ES", "US": "US",
             "NL": "NL", "CH": "CH_LI", "GB": "GB"}
_MK_BDI = {"DE": "DE", "AT": "AT", "FR": "FR", "ES": "ES", "US": "US",
           "NL": None, "CH": "CH", "GB": "GB"}


@functools.lru_cache(maxsize=None)
def _bdi_spend_per_market() -> dict:
    p = _cpath("bdi_spend_per_market.csv")
    if not os.path.exists(p):
        return {}
    df = pd.read_csv(p)
    return {str(r["code"]): float(r["eur_per_viaggiatore"])
            for _, r in df.iterrows() if pd.notna(r["eur_per_viaggiatore"])}


def _bdi_spend_per_night(year: int) -> dict:
    long = bdi_country_long()
    if long is None or long.empty:
        return {}
    d = long.copy()
    d["anno"] = d["date"].dt.year
    d = d[d["anno"] == year]
    if d.empty:
        return {}
    g = d.groupby("code").agg(spesa=("spesa", "sum"), notti=("notti", "sum"), q=("date", "count"))
    return {c: r["spesa"] * 1e6 / (r["notti"] * 1e3)
            for c, r in g.iterrows() if r["q"] >= 4 and r["notti"] > 0}


def _region_market_value(code: str, markets) -> dict:
    """Peso economico REALE di ogni mercato nella regione = presenze (ISTAT _9, ultimo anno)
    × spesa/notte del mercato (BdI). {code: M€}. {} se dati assenti."""
    df = estero_regione_long()
    if df is None:
        return {}
    area = RG.istat_area(code)
    ni = df[(df["area"] == area) & (df["datatype"] == "NI")]
    if ni.empty:
        return {}
    yr = int(ni["anno"].max())
    ni = ni[ni["anno"] == yr]
    spn = _bdi_spend_per_night(yr)
    if not spn:
        return {}
    med = float(pd.Series(list(spn.values())).median())
    pres = dict(zip(ni["country"], ni["valore"]))
    out = {}
    for mk in markets:
        p = pres.get(_MK_ISTAT.get(mk.code, mk.code))
        if p is None and mk.code == "GB":
            p = pres.get("UK")
        eurn = spn.get(_MK_BDI.get(mk.code) or "", med)
        out[mk.code] = round(float(p or 0) * eurn / 1e6, 3)
    return out if any(out.values()) else {}


def _reco_cat(reco: str) -> str:
    for k in ("Aumentare", "Ridurre", "Monitorare"):
        if reco.startswith(k):
            return k
    return "Mantenere"


@functools.lru_cache(maxsize=None)
def azioni_snapshot(code: str) -> dict:
    """Azioni raccomandate per la regione: ranking mercati esteri con priorità.
    score = max(forza,0) × momentum × peso_economico × fattibilità (dal motore).
    Vista nazionale (ITALIA): usa la regione di default come esempio (has_national=True)."""
    national = RG.is_national(code)
    eng = RG.DEFAULT_REGION if national else code
    base = {"actions": [], "n_alta": 0, "n_aumentare": 0, "picco_top": "—",
            "region": RG.region(eng)["nome"], "has_national": national, "error": ""}
    cwd = os.getcwd()
    try:
        os.chdir(_ROOT)
        from tourism_wedge import engine_aggregate as _EA
        from tourism_wedge import DEFAULT_MARKETS as _MKTS
        info = RG.region(eng)
        df = _EA.assemble_real(start="2019-01", region_code=eng, trends_kw=info["trends_kw"])
        ranked = _EA.rank_markets(df, value_override=(_bdi_spend_per_market() or None),
                                  weight_override=(_region_market_value(eng, _MKTS) or None))
    except Exception as e:  # noqa: BLE001
        return {**base, "error": f"{type(e).__name__}: {e}"}
    finally:
        os.chdir(cwd)

    def peak(mk_code):
        col = f"search_{mk_code}"
        if col not in df.columns:
            return None
        d = df.dropna(subset=[col])
        if d.empty:  # mercato senza serie Trends utile → nessun picco (niente idxmax su tutti-NA)
            return None
        prof = d.assign(m=d["date"].dt.month).groupby("m")[col].mean().dropna()
        return int(prof.idxmax()) if not prof.empty else None

    order = {"Alta": 0, "Media": 1, "Bassa": 2}
    acts = []
    for c in ranked:
        cat = _reco_cat(c["raccomandazione"])
        forza = c.get("forza_anticipatrice") or 0
        mom = c.get("momentum_search_pct") or 0
        if cat == "Aumentare" and forza >= 0.4 and mom > 20:
            prio = "Alta"
        elif cat in ("Aumentare", "Ridurre"):
            prio = "Media"
        else:
            prio = "Bassa"
        pm = peak(c["code"])
        acts.append({
            "market": c["market"], "code": c["code"], "reco": c["raccomandazione"], "cat": cat,
            "priorita": prio, "forza": f"{c['forza_anticipatrice']:+.2f}",
            "momentum": f"{c['momentum_search_pct']:+.0f}%",
            "valore": f"€ {int(c['valore_eur_per_visitatore'])}",
            "picco": (MESI_IT[pm - 1] if pm else "—"),
            "score": int(c["score"]), "rank": c["rank"],
        })
    acts.sort(key=lambda a: (order[a["priorita"]], -a["score"]))
    return {**base, "actions": acts,
            "n_alta": sum(1 for a in acts if a["priorita"] == "Alta"),
            "n_aumentare": sum(1 for a in acts if a["cat"] == "Aumentare"),
            "picco_top": (acts[0]["picco"] if acts else "—")}


# ── INTERESSE ONLINE (Google Trends) — segnale leading per-mercato ──
# Trends normalizza 0-100 SEPARATAMENTE per query/geo: i livelli NON sono confrontabili
# tra paesi in assoluto, ma il MOMENTUM (variazione nel tempo dentro lo stesso paese) sì.
_TRENDS_MARKETS = [("DE", "Germania"), ("AT", "Austria"), ("GB", "Regno Unito"),
                   ("NL", "Paesi Bassi"), ("CH", "Svizzera"), ("US", "Stati Uniti")]


def _safe_kw(kw: str) -> str:
    return "".join(c if c.isalnum() else "_" for c in kw.lower())


def interesse_snapshot(code: str) -> dict:
    """Interesse di ricerca (Google Trends) per la regione, per i 6 mercati del motore.
    Serie mensili + momentum (media ultimi 3 mesi vs stessi 3 mesi un anno prima)."""
    empty = {"keyword": "—", "n": 0, "top_nome": "—", "top_mom": "—", "series": [], "rows": []}
    info = RG.region(code)
    kw = info.get("trends_kw") or info["nome"]
    sk = _safe_kw(kw)
    series, rows = [], []
    for mk, nome in _TRENDS_MARKETS:
        p = _cpath(f"trends_{sk}_{mk}.csv")
        if not os.path.exists(p):
            continue
        df = pd.read_csv(p, parse_dates=["date"]).sort_values("date")
        if df.empty or "search" not in df:
            continue
        s = df["search"].astype(float)
        recent = float(s.tail(3).mean()) if len(s) >= 3 else (float(s.mean()) if len(s) else 0.0)
        yoy = float(s.tail(15).head(3).mean()) if len(s) >= 15 else float("nan")
        mom = ((recent - yoy) / yoy * 100) if (pd.notna(yoy) and yoy) else None
        series.append({"code": mk, "nome": nome,
                       "dates": [d.strftime("%Y-%m") for d in df["date"]],
                       "values": [float(v) for v in s]})
        rows.append({"nome": nome, "recente": f"{recent:.0f}",
                     "momentum": (f"{mom:+.0f}%" if mom is not None else "—"),
                     "_m": (mom if mom is not None else -1e9)})
    if not series:
        return empty
    rows.sort(key=lambda r: -r["_m"])
    top = rows[0]
    return {"keyword": kw, "n": len(series), "top_nome": top["nome"], "top_mom": top["momentum"],
            "series": series,
            "rows": [{"nome": r["nome"], "recente": r["recente"], "momentum": r["momentum"]} for r in rows]}


@functools.lru_cache(maxsize=None)
def forecast_snapshot(code: str) -> dict:
    """Previsione delle presenze straniere (motore statistico): storico + previsione con
    intervallo, confronto anno prima, e qualità del modello (MAPE, batte il naïve?, lag).
    Vista nazionale (ITALIA): usa la regione di default come esempio."""
    national = RG.is_national(code)
    eng = RG.DEFAULT_REGION if national else code
    base = {"region": RG.region(eng)["nome"], "has_national": national, "error": "",
            "hist_x": [], "hist_y": [], "fc_x": [], "fc_mean": [], "fc_lo": [], "fc_hi": [],
            "fc_lastyear": [], "mape": "—", "beats": "—", "lag": "—", "reading": ""}
    cwd = os.getcwd()
    try:
        os.chdir(_ROOT)
        from tourism_wedge import engine_aggregate as _EA
        info = RG.region(eng)
        df = _EA.assemble_real(start="2019-01", region_code=eng, trends_kw=info["trends_kw"])
        fit = _EA.fit_aggregate(df)
        fc = _EA.forecast_aggregate(fit)
        obs = df.dropna(subset=["presences"])[["date", "presences"]]
    except Exception as e:  # noqa: BLE001
        return {**base, "error": f"{type(e).__name__}: {e}"}
    finally:
        os.chdir(cwd)

    mape = float(getattr(fit, "mape_model", float("nan")))
    return {**base,
            "hist_x": [pd.Timestamp(d).strftime("%Y-%m") for d in obs["date"]],
            "hist_y": [float(v) for v in obs["presences"]],
            "fc_x": [pd.Timestamp(d).strftime("%Y-%m") for d in fc.dates],
            "fc_mean": [float(x) for x in fc.mean],
            "fc_lo": [float(x) for x in fc.lo],
            "fc_hi": [float(x) for x in fc.hi],
            "fc_lastyear": [None if pd.isna(x) else float(x) for x in fc.lastyear],
            "mape": (f"{mape * 100:.1f}%" if mape == mape and mape < 1 else (f"{mape:.1f}%" if mape == mape else "—")),
            "beats": ("sì" if getattr(fit, "beats_naive", False) else "no"),
            "lag": str(int(getattr(fit, "lag", 0))),
            "reading": (fit.coefficient_reading() if hasattr(fit, "coefficient_reading") else "")}


def confronto_snapshot() -> dict:
    """Confronto di TUTTE le regioni su più metriche (ultimo valore disponibile da
    region_annual_panel): spesa, presenze straniere, posti letto, occupazione,
    spesa/viaggiatore, quota stranieri. + arrays per scatter spesa×occupazione."""
    rows = []
    sc_names, sc_codes, sc_spesa, sc_occ, sc_letti = [], [], [], [], []
    for code, info in RG.REGIONS.items():
        panel = region_annual_panel(code)

        def lastv(k, _p=None):
            p = panel
            if p is None or getattr(p, "empty", True) or k not in p.columns:
                return None
            s = p[k].dropna()
            return float(s.iloc[-1]) if not s.empty else None

        spesa, presenze, letti = lastv("spesa"), lastv("presenze_str"), lastv("letti")
        occ, spv, quota = lastv("occ"), lastv("spesa_per_viagg"), lastv("quota_str")
        rows.append({"code": code, "regione": info["nome"],
                     "spesa": _fmt_val(spesa, "M€", 0), "presenze": _it_int(presenze),
                     "letti": _it_int(letti), "occ": _fmt_val(occ, "%", 1),
                     "sp_viagg": _fmt_val(spv, "€", 0), "quota": _fmt_val(quota, "%", 1),
                     "_spesa": (spesa if spesa is not None else -1.0),
                     "_occ": (occ if occ is not None else -1.0)})
        if spesa is not None and occ is not None:
            sc_names.append(info["nome"]); sc_codes.append(code)
            sc_spesa.append(spesa); sc_occ.append(occ); sc_letti.append(letti or 0.0)
    rows.sort(key=lambda r: -r["_spesa"])
    top_spesa = rows[0]["regione"] if rows else "—"
    top_occ = max(rows, key=lambda r: r["_occ"])["regione"] if rows else "—"
    clean = [{k: v for k, v in r.items() if not k.startswith("_")} for r in rows]
    return {"n": len(rows), "top_spesa": top_spesa, "top_occ": top_occ, "rows": clean,
            "sc_names": sc_names, "sc_codes": sc_codes,
            "sc_spesa": sc_spesa, "sc_occ": sc_occ, "sc_letti": sc_letti}


def gestione_snapshot() -> dict:
    """Inventario in sola lettura delle sorgenti dati, con stato letto DALLA CACHE reale
    (numero file, ultimo dato). Nessuna duplicazione di cataloghi statici → niente drift."""
    import glob as _glob

    def status(pat, date_col=None, anno_col=None):
        files = _glob.glob(_cpath(pat))
        if not files:
            return 0, None
        f = max(files, key=os.path.getmtime)
        ultimo = None
        try:
            df = pd.read_csv(f)
            if date_col and date_col in df.columns:
                t = pd.to_datetime(df[date_col], errors="coerce").max()
                ultimo = t.strftime("%m/%Y") if pd.notna(t) else None
            elif anno_col and anno_col in df.columns:
                y = pd.to_numeric(df[anno_col], errors="coerce").max()
                ultimo = str(int(y)) if pd.notna(y) else None
        except Exception:  # noqa: BLE001
            pass
        return len(files), ultimo

    rows = []

    def add(nome, ente, freq, livello, agg, n, ultimo, dett=""):
        rows.append({"nome": nome, "ente": ente, "freq": freq, "livello": livello, "agg": agg,
                     "stato": ("✅" if n else "—"), "ultimo": (ultimo or "—"), "dettaglio": dett})

    n, u = status("istat_????_NI_ALL_WRL_X_ITA.csv", date_col="date")
    add("Presenze straniere (mensile)", "ISTAT", "mensile", "Regionale", "live", n, u, f"{n} regioni")
    n, u = status("istat_?????_NI_ALL_WORLD.csv", date_col="date")
    add("Presenze per provincia", "ISTAT", "mensile", "Provinciale", "live", n, u, f"{n} province")
    n, u = status("istat_capacity_letti_*.csv", anno_col="anno")
    add("Capacità ricettiva (posti letto)", "ISTAT", "annuale", "Regionale", "live", n, u, f"{n} file")
    n, u = status("istat_estero_regione_prov_annuale.csv", anno_col="TIME_PERIOD")
    add("Esteri per paese × territorio", "ISTAT (cube _9)", "annuale", "Naz./Regionale", "live", n, u)
    n, u = status("trends_*.csv", date_col="date")
    add("Interesse di ricerca (Google Trends)", "Google Trends", "mensile", "mercato × regione", "live", n, u, f"{n} serie")
    add("Turismo internazionale (spesa · notti · viaggiatori)", "Banca d'Italia", "trimestrale",
        "Naz./Regionale", "manuale (xlsx)", 1 if os.path.exists(_cpath("bdi_turismo_ts.xlsx")) else 0,
        "2025", "1997–2025")
    add("Spesa straniera per regione", "Banca d'Italia", "annuale", "Regionale", "manuale",
        1 if os.path.exists(_cpath("bdi_extended.json")) else 0, "2024", "")
    add("Spesa €/viaggiatore per mercato", "Banca d'Italia", "annuale", "Nazionale", "manuale",
        1 if os.path.exists(_cpath("bdi_spend_per_market.csv")) else 0, "—", "")
    ns, us = 0, None
    try:
        pt = _cpath("str_airbnb_territorio.csv")
        if os.path.exists(pt):
            d = pd.read_csv(pt)
            ns = len(d)
            if "snapshot" in d.columns and ns:
                us = str(d["snapshot"].iloc[0])
    except Exception:  # noqa: BLE001
        pass
    add("Affitti brevi (Inside Airbnb)", "Inside Airbnb", "snapshot", "Città + 3 regioni",
        "manuale", ns, us, f"{ns} territori")

    return {"rows": rows, "n_fonti": len(rows), "n_ok": sum(1 for r in rows if r["stato"] == "✅"),
            "n_live": sum(1 for r in rows if r["agg"].startswith("live"))}


def mercati_snapshot(code: str, top_bar: int = 12, top_lines: int = 5) -> dict:
    """Mercati esteri della regione in tipi Python puri (per la pagina 'Mercati
    d'origine'): classifica ultimo anno (barre + tabella con quota) e serie storiche
    annuali dei primi `top_lines` paesi. Robusto ai dati mancanti."""
    empty = {"anno": None, "n": 0, "primo": "—", "primo_val": "—",
             "bar_nomi": [], "bar_val": [], "rows": [], "line_years": [], "line_series": []}
    mk = estero_markets(code)  # tutti i paesi, ultimo anno, ordinati desc
    if mk is None or mk.empty:
        return empty
    anni = estero_years(code)
    anno = anni[-1] if anni else None
    top = mk.head(top_bar)
    bar_nomi = [r["nome"] for _, r in top.iterrows()]
    bar_val = [float(r["valore"]) for _, r in top.iterrows()]
    tot = float(mk["valore"].sum()) or 1.0
    rows = [{"nome": r["nome"],
             "valore": _it_int(r["valore"]),
             "quota": f"{r['valore'] / tot * 100:.1f}".replace(".", ",") + "%"}
            for _, r in top.iterrows()]
    # serie storiche dei primi `top_lines` paesi, allineate agli anni della regione
    line_series = []
    for _, r in mk.head(top_lines).iterrows():
        s = estero_country_series(code, r["country"])
        m = {int(a): float(v) for a, v in zip(s["anno"], s["valore"])} if s is not None else {}
        line_series.append({"nome": r["nome"], "values": [m.get(y) for y in anni]})
    return {"anno": anno, "n": len(mk),
            "primo": str(mk.iloc[0]["nome"]), "primo_val": _it_int(mk.iloc[0]["valore"]),
            "bar_nomi": bar_nomi, "bar_val": bar_val, "rows": rows,
            "line_years": [int(y) for y in anni], "line_series": line_series}


# ════════════════════════════════════════════════════════════════════════════
# STR / AFFITTI BREVI (Inside Airbnb) — indipendente dalla regione
# (7 città + 3 regioni: Sicilia, Puglia, Trentino; no Abruzzo)
# ════════════════════════════════════════════════════════════════════════════
_STR_TERR_PATH = _cpath("str_airbnb_territorio.csv")
_STR_ZONA_PATH = _cpath("str_airbnb_zona.csv")
_STR_ROOM_PATH = _cpath("str_airbnb_roomtype.csv")
_STR_REV_PATH = _cpath("str_airbnb_recensioni_mese.csv")


@functools.lru_cache(maxsize=None)
def str_territori():
    return pd.read_csv(_STR_TERR_PATH) if os.path.exists(_STR_TERR_PATH) else pd.DataFrame()


@functools.lru_cache(maxsize=None)
def _str_zone_all():
    return pd.read_csv(_STR_ZONA_PATH) if os.path.exists(_STR_ZONA_PATH) else pd.DataFrame()


@functools.lru_cache(maxsize=None)
def _str_room_all():
    return pd.read_csv(_STR_ROOM_PATH) if os.path.exists(_STR_ROOM_PATH) else pd.DataFrame()


@functools.lru_cache(maxsize=None)
def str_reviews_monthly():
    if not os.path.exists(_STR_REV_PATH):
        return pd.DataFrame()
    df = pd.read_csv(_STR_REV_PATH)
    df["date"] = pd.to_datetime(df["mese"].astype(str) + "-01", errors="coerce")
    return df


def str_territori_list() -> list:
    """Lista (territorio, slug) dei territori STR coperti, ordinata per n. annunci desc."""
    df = str_territori()
    if df.empty:
        return []
    d = df.sort_values("n_annunci", ascending=False)
    return [(str(r["territorio"]), str(r["slug"])) for _, r in d.iterrows()]


def _pct(v, dec: int = 1) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"{float(v):.{dec}f}".replace(".", ",") + "%"


def str_snapshot(slug: str) -> dict:
    """Quadro affitti brevi (Inside Airbnb) del territorio `slug`, in tipi Python puri.
    Indipendente dalla regione. Robusto ai dati mancanti."""
    empty = {"slug": slug, "territorio": "—", "tipo": "—", "k_annunci": "—", "k_adr": "—",
             "k_intero": "—", "k_multihost": "—", "k_licenza": "—", "k_rating": "—",
             "adr_nomi": [], "adr_val": [], "adr_slugs": [], "zone_nomi": [], "zone_val": [],
             "room_nomi": [], "room_val": [], "rev_x": [], "rev_y": []}
    terr = str_territori()
    if terr.empty:
        return empty
    row = terr[terr["slug"] == slug]
    r = row.iloc[0].to_dict() if not row.empty else {}

    def num(k):
        v = r.get(k)
        return None if v is None or (isinstance(v, float) and pd.isna(v)) else float(v)

    d = terr.dropna(subset=["adr_mediano"]).sort_values("adr_mediano")
    adr_nomi = [str(x) for x in d["territorio"]]
    adr_val = [float(x) for x in d["adr_mediano"]]
    adr_slugs = [str(x) for x in d["slug"]]

    zdf = _str_zone_all()
    zone_nomi, zone_val = [], []
    if not zdf.empty:
        z = zdf[zdf["slug"] == slug].sort_values("n_annunci", ascending=False).head(12)
        zone_nomi = [str(x) for x in z["zona"]]
        zone_val = [float(x) for x in z["n_annunci"]]

    rdf = _str_room_all()
    room_nomi, room_val = [], []
    if not rdf.empty:
        rt = rdf[rdf["slug"] == slug].sort_values("n_annunci")
        room_nomi = [str(x) for x in rt["room_type"]]
        room_val = [float(x) for x in rt["n_annunci"]]

    rev = str_reviews_monthly()
    rev_x, rev_y = [], []
    if not rev.empty:
        rv = rev[rev["slug"] == slug].dropna(subset=["date"]).sort_values("date")
        rev_x = [t.strftime("%Y-%m") for t in rv["date"]]
        rev_y = [float(x) for x in rv["n_recensioni"]]

    adr = num("adr_mediano")
    rating = num("rating_medio")
    return {
        "slug": slug,
        "territorio": str(r.get("territorio", "—")),
        "tipo": str(r.get("tipo", "—")),
        "k_annunci": _it_int(num("n_annunci")),
        "k_adr": (f"€ {_it_int(adr)}" if adr is not None else "—"),
        "k_intero": _pct(num("pct_intero")),
        "k_multihost": _pct(num("pct_multihost")),
        "k_licenza": _pct(num("pct_licenza")),
        "k_rating": (f"{rating:.2f}".replace(".", ",") if rating is not None else "—"),
        "adr_nomi": adr_nomi, "adr_val": adr_val, "adr_slugs": adr_slugs,
        "zone_nomi": zone_nomi, "zone_val": zone_val,
        "room_nomi": room_nomi, "room_val": room_val,
        "rev_x": rev_x, "rev_y": rev_y,
    }


_GEOJSON_PATH = os.path.join(_ROOT, "assets", "italy_regions.geojson")


@functools.lru_cache(maxsize=None)
def _italy_geojson():
    """GeoJSON delle regioni italiane (feature.properties.reg_name). None se manca."""
    if not os.path.exists(_GEOJSON_PATH):
        return None
    return json.load(open(_GEOJSON_PATH, encoding="utf-8"))


def mappa_snapshot(year: int | None = None) -> dict:
    """Dati per la mappa coropletica d'Italia (spesa straniera per regione): geojson +
    liste allineate names/codes/z (in ordine di feature). `codes[i]` = NUTS2 della
    regione i-esima (per il click → selezione); z = spesa M€ (Banca d'Italia)."""
    gj = _italy_geojson()
    if gj is None:
        return {"geojson": None, "names": [], "codes": [], "z": [], "anno": year or 2024}
    rk = regions_spend_ranking() if year is None else regions_spend_ranking_year(year)
    spend = {RG.REGIONS[r["code"]]["bdi"]: r["spesa_M"] for r in rk}
    names, codes, z = [], [], []
    for f in gj["features"]:
        nm = f["properties"]["reg_name"]
        code = RG.code_for_geo(nm)
        names.append(nm)
        codes.append(code or "")
        z.append(spend.get(RG.region(code)["bdi"]) if code else None)
    return {"geojson": gj, "names": names, "codes": codes, "z": z, "anno": year or 2024}


REGIONI_SELECT = [(info["nome"], code) for code, info in RG.REGIONS.items()]


# ── self-check: `python tdh_data.py` stampa numeri REALI per alcune regioni ──
if __name__ == "__main__":
    for _c in ("ITE1", "ITC4", "ITF1", "ITALIA"):
        try:
            s = regione_snapshot(_c)
        except Exception as e:  # noqa: BLE001
            print(f"{_c}: ERRORE {type(e).__name__}: {e}")
            continue
        print(f"\n== {s['nome']} ({_c}) ==")
        print(f"  presenze straniere ultimo mese ({s['kpi_presenze_periodo']}): {s['kpi_presenze']}")
        print(f"  posti letto: {s['kpi_letti']} ({s['kpi_letti_anno']})")
        print(f"  spesa straniera: {s['kpi_spesa']}  [{s['kpi_spesa_rank']}]")
        print(f"  mercati esteri: {s['kpi_mercati']} ({s['kpi_mercati_anno']})")
        print(f"  serie: {len(s['serie_presenze'])} mesi; top mercati: "
              + ", ".join(f"{m['nome']} {m['valore']:,.0f}" for m in s["mercati_top"]))
