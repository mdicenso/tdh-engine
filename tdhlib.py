"""
Libreria condivisa del cruscotto multi-pagina TDH Engine.

Contiene: calcolo dati (sintetico/reale, in cache), costruttori di grafici Plotly,
allocatore di budget, KPI, e il componente assistente Claude (chat con tool).
Le pagine (app.py) usano queste funzioni; la logica del motore resta in tourism_wedge/.
"""
from __future__ import annotations

import datetime
import glob
import io
import json
import math
import os
import time
from dataclasses import asdict

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from tourism_wedge import (DEFAULT_MARKETS, SyntheticProvider,
                           fit_market, forecast_within_lead, build_card, portfolio)
from tourism_wedge.engine_aggregate import (assemble_real, fit_aggregate,
                                            forecast_aggregate, rank_markets)
from tourism_wedge import candidate_sources as CS
from tourism_wedge import real_sources as RS
import regions as RG
import econ_sources as ECS

MODEL = "claude-sonnet-4-6"
MODE_SYN = "🧪 Sintetico (collaudo)"
MODE_REAL = "🌍 Reale (ECB · ISTAT · Google Trends)"

# geografia per la mappa
CENTROID = {"DE": (51.2, 10.4), "AT": (47.6, 14.1), "GB": (54.0, -2.0),
            "NL": (52.1, 5.3), "CH": (46.8, 8.2), "US": (39.5, -98.5)}
ABRUZZO = (42.35, 13.4)
MONTHS_IT = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]


def reco_color(reco: str) -> str:
    if reco.startswith("Aumentare"):
        return "#16a34a"
    if reco.startswith("Ridurre"):
        return "#dc2626"
    if reco.startswith("Monitorare"):
        return "#f59e0b"
    return "#2563eb"


def reco_cat(reco: str) -> str:
    for k in ("Aumentare", "Ridurre", "Monitorare"):
        if reco.startswith(k):
            return k
    return "Mantenere"


# ════════════════════════════════════════════════════════════════════════════
# STILE / UI (direzione A — moderno/pulito; CSS puro, a prova di proxy)
# ════════════════════════════════════════════════════════════════════════════
FONT_STACK = "-apple-system, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif"
# ── Design system "Istituzionale" (Swiss/enterprise): petrolio + neutri freddi, filetti,
#    sidebar CHIARA, card a filetto (niente barra colorata), sobrio. Token in :root. ──
CSS = """<style>
  :root{
    --bg:#f5f7f7; --panel:#ffffff; --panel2:#fbfcfc;
    --ink:#10262a; --mut:#5c7176; --faint:#8aa0a4;
    --line:#e4ebec; --line2:#eef3f3;
    --accent:#0e6b70; --accent-ink:#0a4f53; --accent-soft:rgba(14,107,112,.10);
  }
  html, body, [class*="css"] { font-family:-apple-system,'Segoe UI',Roboto,Helvetica,Arial,sans-serif; }
  /* spazio in alto sufficiente a non finire sotto l'header fisso di Streamlit (topbar tagliata) */
  .block-container { padding-top: 3.9rem; max-width: 1320px; }

  /* Tipografia */
  h1 { color: var(--ink); font-weight: 700; font-size: 1.55rem; letter-spacing: -0.02em; }
  h2 { color: var(--ink); font-weight: 650; font-size: 1.12rem; letter-spacing: -0.01em;
       border-left: 3px solid var(--accent); padding-left: 10px; margin-top: 1.4rem; }
  h3 { color: #2c3f44; font-weight: 600; font-size: 1rem; }

  .stApp { background-color: var(--bg); }

  /* ---- Sidebar CHIARA ---- */
  [data-testid="stSidebar"] { background-color: var(--panel); border-right: 1px solid var(--line); }
  [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3,
  [data-testid="stSidebar"] p, [data-testid="stSidebar"] label,
  [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] { color: var(--ink) !important; }
  [data-testid="stSidebar"] [data-testid="stCaptionContainer"] { color: var(--mut) !important; }
  [data-testid="stSidebarNav"] a span, [data-testid="stSidebarNav"] a p { color: var(--mut) !important; opacity: 1 !important; }
  [data-testid="stSidebarNav"] span[data-testid="stIconMaterial"] { color: var(--mut) !important; }
  [data-testid="stSidebarNav"] a:hover { background: var(--line2); border-radius: 8px; }
  [data-testid="stSidebarNav"] a:hover span, [data-testid="stSidebarNav"] a:hover p { color: var(--ink) !important; }
  [data-testid="stSidebarNav"] a[aria-current="page"] {
    background: var(--accent-soft); border-radius: 8px; border-left: 3px solid var(--accent); }
  [data-testid="stSidebarNav"] a[aria-current="page"] span,
  [data-testid="stSidebarNav"] a[aria-current="page"] p { color: var(--accent-ink) !important; font-weight: 700; }
  [data-testid="stSidebarNavViewButton"],
  [data-testid="stSidebarNavViewButton"] span,
  [data-testid="stSidebarNavViewButton"] p { color: var(--mut) !important; opacity: 1 !important; }
  [data-testid="stSidebarNavViewButton"]:hover span,
  [data-testid="stSidebarNavViewButton"]:hover p { color: var(--accent-ink) !important; }
  [data-testid="stSidebar"] .stButton > button {
    background: var(--panel2); border: 1px solid var(--line); color: var(--ink);
    border-radius: 9px; font-weight: 500; }
  [data-testid="stSidebar"] .stButton > button:hover { border-color: var(--accent); color: var(--accent-ink); }

  /* ---- Metric card: filetto, niente barra colorata in cima ---- */
  [data-testid="stMetric"] {
    background: var(--panel); border: 1px solid var(--line);
    border-radius: 14px; padding: 16px 18px; box-shadow: none; }
  [data-testid="stMetricLabel"] { color: var(--mut); font-weight: 600; font-size: .78rem;
    text-transform: uppercase; letter-spacing: .04em; }
  [data-testid="stMetricValue"] { color: var(--ink); font-weight: 700;
    font-variant-numeric: tabular-nums; }

  /* ---- Grafici: contenitore a filetto ---- */
  [data-testid="stPlotlyChart"] > div { border: 1px solid var(--line); border-radius: 14px;
    overflow: hidden; box-shadow: none; }

  /* ---- Tabelle ---- */
  [data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; border: 1px solid var(--line); }

  /* ---- Bottoni ---- */
  .stButton > button { border-radius: 9px; border: 1px solid var(--line); color: var(--ink);
    font-weight: 500; transition: all .15s; }
  .stButton > button:hover { border-color: var(--accent); color: var(--accent-ink); }
  .stButton > button[kind="primary"] { background: var(--accent); border: 1px solid var(--accent); color: #fff; }
  .stButton > button[kind="primary"]:hover { background: var(--accent-ink); border-color: var(--accent-ink); }

  /* ---- Expander / Alert / Select / Slider ---- */
  [data-testid="stExpander"] { border-radius: 12px; border: 1px solid var(--line); box-shadow: none; }
  [data-testid="stAlert"] { border-radius: 10px; border-left-width: 4px; }
  [data-testid="stSelectbox"] > div > div { border-radius: 8px; border-color: var(--line); }
  [data-testid="stSlider"] [data-testid="stThumbValue"] { color: var(--accent-ink); font-weight: 700; }

  /* ---- Scrollbar ---- */
  * { scrollbar-width: auto; scrollbar-color: #b7c4c6 transparent; }
  ::-webkit-scrollbar { width: 14px; height: 14px; }
  ::-webkit-scrollbar-track { background: transparent; }
  ::-webkit-scrollbar-thumb { background: #b7c4c6; border-radius: 8px;
    border: 3px solid transparent; background-clip: padding-box; }
  ::-webkit-scrollbar-thumb:hover { background: #93a9ad; background-clip: padding-box; }
</style>"""


def inject_css():
    st.markdown(CSS, unsafe_allow_html=True)


def check_access() -> bool:
    """Gate d'accesso leggero. Attivo solo se esiste il secret APP_PASSWORD
    (Streamlit Cloud → Settings → Secrets). Senza secret non blocca nulla.
    Ritorna True se l'utente può procedere."""
    try:
        pwd = st.secrets.get("APP_PASSWORD")
    except Exception:  # noqa: BLE001 — nessun file secrets in locale
        pwd = None
    if not pwd:
        return True
    if st.session_state.get("_auth_ok"):
        return True

    _, c, _ = st.columns([1, 1.4, 1])
    with c:
        st.markdown(LOGO_SVG, unsafe_allow_html=True)
        st.markdown("#### Turism Data Hub — accesso riservato")
        st.caption("Prototipo Indra · accesso riservato. Inserisci la password ricevuta.")
        entered = st.text_input("Password", type="password", label_visibility="collapsed",
                                placeholder="Password")
        if st.button("Entra", type="primary", use_container_width=True):
            if entered == pwd:
                st.session_state["_auth_ok"] = True
                st.rerun()
            else:
                st.error("Password non corretta.")
    return False


def hero(subtitle: str = "", mode_label: str = ""):
    chip = (f"<span style='margin-left:auto;background:rgba(14,107,112,.10);color:#0a4f53;"
            f"border:1px solid rgba(14,107,112,.28);padding:4px 13px;border-radius:999px;"
            f"font-size:.8rem;font-weight:600'>{mode_label}</span>") if mode_label else ""
    st.markdown(f"""
    <div style="margin:.1rem 0 1rem;padding-bottom:.7rem;border-bottom:1px solid #e4ebec;
                display:flex;align-items:center;gap:10px;flex-wrap:wrap">
      <div>
        <div style="font-size:.68rem;letter-spacing:.12em;text-transform:uppercase;
                    color:#8aa0a4;font-weight:600">Strumento ad uso interno · scala nazionale</div>
        <div style="font-size:1.5rem;font-weight:700;letter-spacing:-.02em;margin-top:4px;
                    color:#10262a">Turism Data Hub</div>
        <div style="color:#5c7176;font-size:.86rem;margin-top:3px">{subtitle}</div>
      </div>{chip}
    </div>""", unsafe_allow_html=True)


def section_header(title: str, subtitle: str = "", icon: str = ""):
    """Intestazione di sezione con linea accent — sostituisce st.subheader() nelle pagine."""
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:10px;margin:1.4rem 0 .6rem;
                padding-bottom:.6rem;border-bottom:1px solid #e2e8f0">
      {"<span style='font-size:1.2rem'>"+icon+"</span>" if icon else ""}
      <div>
        <div style="font-size:1.05rem;font-weight:700;color:#0f172a">{title}</div>
        {"<div style='font-size:.8rem;color:#64748b;margin-top:1px'>"+subtitle+"</div>" if subtitle else ""}
      </div>
    </div>""", unsafe_allow_html=True)


def page_header(title: str, subtitle: str = "", group: str = "", emoji: str = "",
                region_code: str | None = None):
    """Header di pagina PIATTO (design Istituzionale): breadcrumb in maiuscoletto, titolo
    grande su fondo pagina, badge regione a pill petrolio, filetto di chiusura. `emoji` è
    accettato per compatibilità ma NON renderizzato (look pulito, senza icone-emoji)."""
    crumb = (f"{group} &nbsp;›&nbsp; {title}" if group else title)
    # la regione attiva è mostrata nella TOPBAR (in alto a destra): niente badge qui.
    sub = (f"<div style='font-size:.86rem;color:#5c7176;margin-top:5px;max-width:920px'>{subtitle}</div>"
           if subtitle else "")
    st.markdown(f"""
    <div style="margin:.1rem 0 1rem;padding-bottom:.7rem;border-bottom:1px solid #e4ebec">
      <div style="font-size:.68rem;letter-spacing:.12em;text-transform:uppercase;
                  color:#8aa0a4;font-weight:600">{crumb}</div>
      <div style="font-size:1.5rem;font-weight:700;letter-spacing:-.02em;line-height:1.15;
                  color:#10262a;margin-top:6px">{title}</div>
      {sub}
    </div>""", unsafe_allow_html=True)


def kpi_row(items: list[dict]):
    """Riga di KPI 'executive' (numeri grandi, label maiuscola, delta opzionale).
    items: [{label, value, delta?, delta_dir?('up'|'down'|'flat'), hint?}]."""
    for col, it in zip(st.columns(len(items)), items):
        dh = ""
        if it.get("delta"):
            d = it.get("delta_dir", "flat")
            color = {"up": "#16a34a", "down": "#dc2626", "flat": "#64748b"}.get(d, "#64748b")
            arrow = {"up": "▲", "down": "▼", "flat": "→"}.get(d, "")
            dh = (f"<div style='font-size:.8rem;font-weight:700;color:{color};margin-top:6px'>"
                  f"{arrow} {it['delta']}</div>")
        hint = (f"<div style='font-size:.72rem;color:#8aa0a4;margin-top:4px'>{it['hint']}</div>"
                if it.get("hint") else "")
        col.markdown(f"""
        <div style="background:#fff;border:1px solid #e4ebec;
                    border-radius:14px;padding:16px 18px;height:100%">
          <div style="font-size:.72rem;font-weight:600;color:#5c7176;text-transform:uppercase;
                      letter-spacing:.05em">{it['label']}</div>
          <div style="font-size:1.9rem;font-weight:700;color:#10262a;line-height:1.1;margin-top:7px;
                      font-variant-numeric:tabular-nums">{it['value']}</div>
          {dh}{hint}
        </div>""", unsafe_allow_html=True)


def badge_html(reco: str) -> str:
    c = reco_color(reco)
    return (f"<span style='background:{c}1f;color:{c};padding:3px 11px;border-radius:999px;"
            f"font-size:.8rem;font-weight:600;white-space:nowrap'>{reco_cat(reco)}</span>")


def _reco_css(val: str) -> str:
    return f"color:{reco_color(str(val))}; font-weight:600;"


def style_reco(df, col: str = "Raccomandazione"):
    """Styler che colora la colonna Raccomandazione come un semaforo."""
    return df.style.map(_reco_css, subset=[col]) if col in df.columns else df


def aggrid_table(df, height: int = 320, reco_col: str | None = None, key: str | None = None):
    """Tabella interattiva (ordinabile/filtrabile) con st_aggrid; fallback a st.dataframe.
    Se reco_col è indicato, colora quella colonna come semaforo della raccomandazione."""
    try:
        from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
    except Exception:
        st.dataframe(df, hide_index=True, use_container_width=True)
        return
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(resizable=True, sortable=True, filter=True)
    if reco_col and reco_col in df.columns:
        js = JsCode("function(p){var v=String(p.value||'');var c='#2563eb';"
                    "if(v.indexOf('Aumentare')===0)c='#16a34a';"
                    "else if(v.indexOf('Ridurre')===0)c='#dc2626';"
                    "else if(v.indexOf('Monitorare')===0)c='#f59e0b';"
                    "return {color:c,fontWeight:'600'};}")
        gb.configure_column(reco_col, cellStyle=js)
    try:
        AgGrid(df, gridOptions=gb.build(), height=height, theme="streamlit",
               allow_unsafe_jscode=True, fit_columns_on_grid_load=True, key=key)
    except Exception:
        st.dataframe(df, hide_index=True, use_container_width=True)


# Logo "inventato" del Turism Data Hub: sole + montagne + lago con riflesso.
LOGO_SVG = """<svg width="54" height="54" viewBox="0 0 64 64" fill="none" xmlns="http://www.w3.org/2000/svg">
  <circle cx="44" cy="20" r="9" fill="#fbbf24"/>
  <path d="M4 44 L22 16 L34 34 L44 22 L60 44 Z" fill="#0e6b70"/>
  <path d="M4 44 L22 16 L30 28 L20 44 Z" fill="#0c5566"/>
  <rect x="4" y="44" width="56" height="2.6" rx="1" fill="#7dd3fc"/>
  <path d="M11 49 H53 M17 53 H47" stroke="#bae6fd" stroke-width="1.6" stroke-linecap="round" opacity="0.75"/>
</svg>"""


@st.cache_data(show_spinner=False)
def home_hero_datauri() -> str | None:
    """Foto della homepage come data URI base64 (offline, niente CDN al render)."""
    import base64
    p = "assets/home_hero.jpg"
    if not os.path.exists(p):
        return None
    return "data:image/jpeg;base64," + base64.b64encode(open(p, "rb").read()).decode()


# ════════════════════════════════════════════════════════════════════════════
# DATI (in cache)
# ════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner="Calcolo il motore (sintetico)…")
def compute_synthetic() -> tuple[list[dict], dict]:
    provider = SyntheticProvider()
    cards, rows = [], {}
    for mk in DEFAULT_MARKETS:
        panel = provider.monthly_panel(mk)
        fit = fit_market(panel, mk)
        fc = forecast_within_lead(fit, mk)
        card = build_card(fit, fc, mk)
        cards.append(card)
        rows[mk.code] = {
            "code": mk.code, "name": mk.name, "currency": mk.currency,
            "spend_per_visitor": mk.spend_per_visitor, "capacity_ok": mk.capacity_ok,
            "card": asdict(card), "lag": int(fit.lag), "beats_naive": bool(fit.beats_naive),
            "mape_model": float(fit.mape_model), "coeff_reading": fit.coefficient_reading(),
            "history": fit.df[["date", "presences", "search", "fx"]].copy(),
            "forecast": {"dates": [pd.Timestamp(d) for d in fc.dates],
                         "mean": [float(x) for x in fc.mean], "lo": [float(x) for x in fc.lo],
                         "hi": [float(x) for x in fc.hi],
                         "lastyear": [None if pd.isna(x) else float(x) for x in fc.lastyear]},
        }
    return portfolio(cards), rows


@st.cache_data(show_spinner="Carico il motore per la regione (ECB · ISTAT · Trends)…")
def compute_real(region_code: str | None = None) -> dict:
    region_code = region_code or RG.DEFAULT_REGION
    info = RG.region(region_code)
    df = assemble_real(start="2019-01", region_code=region_code, trends_kw=info["trends_kw"])
    fit = fit_aggregate(df)
    fc = forecast_aggregate(fit)
    ranked = rank_markets(df, value_override=bdi_spend_per_market() or None,
                          feas_override=region_feas_weights(region_code) or None,
                          weight_override=region_market_value(region_code) or None)
    obs = df.dropna(subset=["presences"])[["date", "presences"]].reset_index(drop=True)
    return {
        "ranked": ranked,
        "agg": {"lag": int(fit.lag), "beats_naive": bool(fit.beats_naive),
                "mape_model": float(fit.mape_model), "coeff_reading": fit.coefficient_reading()},
        "forecast": {"dates": [pd.Timestamp(d) for d in fc.dates],
                     "mean": [float(x) for x in fc.mean], "lo": [float(x) for x in fc.lo],
                     "hi": [float(x) for x in fc.hi],
                     "lastyear": [None if pd.isna(x) else float(x) for x in fc.lastyear]},
        "history": obs, "panel": df,
    }


def region_feas_weights(code: str) -> dict:
    """Pesi di fattibilità (voli) per regione: connettività reale Eurostat sugli AEROPORTI
    della regione (registro). 0.4 se nessun volo diretto da quel mercato, fino a 1.0 per il
    mercato più connesso. Regioni senza aeroporti commerciali → {} (fattibilità di default)."""
    airports = RG.region(code)["airports"]
    if not airports:
        return {}
    try:
        conn = ECS.fetch_region_connectivity(airports, cache_name=f"conn_{code}.csv")
    except Exception:  # noqa: BLE001 — Eurostat non raggiungibile: fattibilità di default
        return {}
    if not conn:
        return {}
    mx = max(conn.values()) or 1
    return {mk.code: round(0.4 + 0.6 * min(conn.get(mk.code, 0) / mx, 1.0), 3) for mk in DEFAULT_MARKETS}


# --- Banca d'Italia: valore economico reale (spesa €/viaggiatore per mercato) ---
@st.cache_data(show_spinner=False)
def bdi_spend_per_market() -> dict:
    p = ".cache/bdi_spend_per_market.csv"
    if not os.path.exists(p):
        return {}
    df = pd.read_csv(p)
    return {str(r["code"]): float(r["eur_per_viaggiatore"])
            for _, r in df.iterrows() if pd.notna(r["eur_per_viaggiatore"])}


# --- Peso economico REALE per mercato NELLA REGIONE (ISTAT _9 × BdI €/notte) ---
_MK_ISTAT = {"DE": "DE", "AT": "AT", "FR": "FR", "ES": "ES", "US": "US",
             "NL": "NL", "CH": "CH_LI", "GB": "GB"}
_MK_BDI = {"DE": "DE", "AT": "AT", "FR": "FR", "ES": "ES", "US": "US",
           "NL": None, "CH": "CH", "GB": "GB"}


def region_market_value(code: str, markets=None) -> dict:
    """Peso economico REALE di ogni mercato NELLA regione = presenze reali (ISTAT _9, ultimo
    anno) × spesa media a notte del mercato (BdI naz; mediana come fallback, es. NL). {code: M€}.
    Un valore per OGNI mercato del motore (0 se il mercato non è presente nella regione).
    {} se manca la cache _9 o BdI → il motore ricade sul €/visitatore nazionale."""
    markets = markets or DEFAULT_MARKETS
    df = estero_regione_long()
    if df is None:
        return {}
    area = RG.istat_area(code)
    ni = df[(df["area"] == area) & (df["datatype"] == "NI")]
    if ni.empty:
        return {}
    yr = int(ni["anno"].max())
    ni = ni[ni["anno"] == yr]
    spn = bdi_spend_per_night(yr)
    if not spn:
        return {}
    med = float(pd.Series(list(spn.values())).median())
    pres = dict(zip(ni["country"], ni["valore"]))
    out = {}
    for mk in markets:
        p = pres.get(_MK_ISTAT.get(mk.code, mk.code))
        if p is None and mk.code == "GB":
            p = pres.get("UK")
        eurn = spn.get(_MK_BDI.get(mk.code) or "", med)
        out[mk.code] = round(float(p or 0) * eurn / 1e6, 3)   # M€ (0 se mercato assente)
    return out if any(out.values()) else {}


@st.cache_data(show_spinner=False)
def bdi_abruzzo_spend():
    p = ".cache/bdi_abruzzo.json"
    if not os.path.exists(p):
        return None
    return json.load(open(p, encoding="utf-8")).get("abruzzo_spesa_M_2024")


# --- Banca d'Italia per REGIONE visitata (TS2): spesa/notti/viaggiatori, trimestrale ---
# colonna del foglio -> codice/i NUTS2 (Trentino = una colonna -> Bolzano+Trento).
_BDI_REG_COL = {4: ["ITC1"], 5: ["ITC2"], 6: ["ITC4"], 7: ["ITC3"], 9: ["ITD1", "ITD2"],
                10: ["ITD3"], 11: ["ITD4"], 12: ["ITD5"], 14: ["ITE1"], 15: ["ITE2"],
                16: ["ITE3"], 17: ["ITE4"], 19: ["ITF1"], 20: ["ITF2"], 21: ["ITF3"],
                22: ["ITF4"], 23: ["ITF5"], 24: ["ITF6"], 25: ["ITG1"], 26: ["ITG2"]}
_BDI_REG_SHEETS = {"spesa": "TS2-S-S", "notti": "TS2-N-S", "viaggiatori": "TS2-V-S"}


@st.cache_data(show_spinner=False)
def bdi_region_long():
    """Spesa/notti/viaggiatori dei turisti stranieri per REGIONE visitata (trimestrale,
    1997-2025). DataFrame: date·code·spesa·notti·viaggiatori. None se l'xlsx manca."""
    import re
    p = ".cache/bdi_turismo_ts.xlsx"
    if not os.path.exists(p):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    recs = []
    for metric, sheet in _BDI_REG_SHEETS.items():
        year = None
        for r in wb[sheet].iter_rows(values_only=True):
            if r and r[0] is not None:
                try:
                    year = int(r[0])
                except (TypeError, ValueError):
                    pass
            m = re.match(r"\s*(\d)", str((r[1] or r[2]) if r else ""))
            if not (year and m):
                continue
            q = int(m.group(1))
            date = pd.Timestamp(year, q * 3 - 2, 1)
            for col, codes in _BDI_REG_COL.items():
                if col < len(r) and r[col] is not None:
                    try:
                        val = float(r[col])
                    except (TypeError, ValueError):
                        continue
                    for code in codes:
                        recs.append((date, code, metric, val))
    if not recs:
        return None
    df = pd.DataFrame(recs, columns=["date", "code", "metric", "value"])
    return df.pivot_table(index=["date", "code"], columns="metric", values="value").reset_index()


def bdi_national_annual():
    """Spesa straniera ANNUALE per TUTTA Italia = somma delle regioni visitate (BdI).
    Deduplica il Trentino, che nei dati è duplicato su Bolzano (ITD1) e Trento (ITD2)."""
    df = bdi_region_long()
    if df is None or df.empty:
        return None
    reps = {codes[0] for codes in _BDI_REG_COL.values()}  # un codice per colonna-regione (no doppio Trentino)
    d = df[df["code"].isin(reps)].copy()
    if d.empty:
        return None
    # nazionale per trimestre = somma sulle regioni; poi aggregato per anno
    qn = d.groupby("date").agg(spesa=("spesa", "sum"), notti=("notti", "sum"),
                               viaggiatori=("viaggiatori", "sum")).reset_index()
    qn["anno"] = qn["date"].dt.year
    return qn.groupby("anno").agg(spesa=("spesa", "sum"), notti=("notti", "sum"),
                                  viaggiatori=("viaggiatori", "sum"),
                                  trimestri=("date", "count")).reset_index()


def bdi_region_annual(code: str):
    """Aggregato ANNUALE (spesa M€, notti/viaggiatori in migliaia) per la regione,
    oppure il totale Italia se è selezionata la vista nazionale."""
    if RG.is_national(code):
        return bdi_national_annual()
    df = bdi_region_long()
    if df is None or df.empty:
        return None
    d = df[df["code"] == code].copy()
    if d.empty:
        return None
    d["anno"] = d["date"].dt.year
    return d.groupby("anno").agg(spesa=("spesa", "sum"), notti=("notti", "sum"),
                                 viaggiatori=("viaggiatori", "sum"),
                                 trimestri=("date", "count")).reset_index()


def chart_region_spend(code: str, yr_range=None) -> go.Figure | None:
    g = bdi_region_annual(code)
    if g is None or g.empty:
        return None
    if yr_range:
        g = g[(g["anno"] >= yr_range[0]) & (g["anno"] <= yr_range[1])]
    fig = go.Figure(go.Bar(x=g["anno"], y=g["spesa"], marker_color="#0e6b70",
                           hovertemplate="<b>%{x}</b><br>%{y:,.0f} M€<extra></extra>"))
    fig.update_yaxes(title="spesa turisti stranieri (M€)")
    return _layout(fig, h=340)


# --- ISTAT esteri per PAESE × territorio (regione/provincia), ANNUALE (cube DCSC_TUR_9) ---
_TUR9_PATH = ".cache/istat_estero_regione_prov_annuale.csv"

# Nomi leggibili dei codici COUNTRY_RES_GUESTS (ISO2 + aggregati). Pass-through se ignoto.
_ISTAT_COUNTRY_NAME = {
    "DE": "Germania", "FR": "Francia", "AT": "Austria", "ES": "Spagna",
    "GB": "Regno Unito", "UK": "Regno Unito", "CH_LI": "Svizzera e Liechtenstein",
    "CH": "Svizzera", "US": "Stati Uniti", "CA": "Canada", "NL": "Paesi Bassi",
    "BE": "Belgio", "RU": "Russia", "CN": "Cina", "JP": "Giappone", "BR": "Brasile",
    "AR": "Argentina", "AU": "Australia", "PL": "Polonia", "SE": "Svezia",
    "DK": "Danimarca", "FI": "Finlandia", "NO": "Norvegia", "IE": "Irlanda",
    "PT": "Portogallo", "GR": "Grecia", "CZ": "Rep. Ceca", "HU": "Ungheria",
    "HR": "Croazia", "SI": "Slovenia", "SK": "Slovacchia", "RO": "Romania",
    "BG": "Bulgaria", "EE": "Estonia", "LT": "Lituania", "LV": "Lettonia",
    "CY": "Cipro", "MT": "Malta", "LU": "Lussemburgo", "IN": "India",
    "KR": "Corea del Sud", "MX": "Messico", "TR": "Turchia", "IL": "Israele",
    "EG": "Egitto", "ZA": "Sudafrica", "NZ": "Nuova Zelanda",
    "WORLD": "Mondo (totale)", "WRL_X_ITA": "Estero (tutti i paesi)",
    "IT": "Italia (residenti)", "EU": "Unione Europea",
    "EUR_NEU": "Altri Europa non-UE", "EUR_OTH": "Altri Europa",
    "AFRMED": "Africa mediterranea", "AFR_OTH": "Altri Africa",
    "AME_N_OTH": "Altri America del Nord", "AME_C_S_OTH": "America centro-sud (altri)",
    "ASI_OTH": "Altri Asia", "ASI_W_OTH": "Asia occidentale (altri)",
    "OCE_OTH": "Altri Oceania",
}


def estero_country_name(code: str) -> str:
    """Nome leggibile del codice paese ISTAT (pass-through se non mappato)."""
    return _ISTAT_COUNTRY_NAME.get(code, code)


def _tur9_is_aggregate(c: str) -> bool:
    """True se il codice è un aggregato (mondo/ripartizioni continentali/UE) o l'Italia,
    non un singolo paese estero. CH_LI (Svizzera+Liechtenstein) è tenuto come mercato."""
    if c in {"WORLD", "WRL_X_ITA", "IT", "EU"}:
        return True
    if c.endswith("_OTH") or c.endswith("_NEU"):
        return True
    return c in {"AFRMED", "EUR", "AFR", "ASI", "AME", "OCE"}


@st.cache_data(show_spinner=False)
def estero_regione_long():
    """Presenze/arrivi turisti ESTERI per PAESE × territorio (naz/regione/provincia),
    ANNUALE (ISTAT DCSC_TUR_9). DataFrame: area·datatype·country·anno·valore.
    None se la cache manca."""
    if not os.path.exists(_TUR9_PATH):
        return None
    df = pd.read_csv(_TUR9_PATH, dtype={"REF_AREA": str, "DATA_TYPE": str,
                                        "COUNTRY_RES_GUESTS": str})
    if df.empty:
        return None
    df = df.rename(columns={"REF_AREA": "area", "DATA_TYPE": "datatype",
                            "COUNTRY_RES_GUESTS": "country", "TIME_PERIOD": "anno",
                            "OBS_VALUE": "valore"})
    df["anno"] = pd.to_numeric(df["anno"], errors="coerce").astype("Int64")
    df["valore"] = pd.to_numeric(df["valore"], errors="coerce")
    return df.dropna(subset=["anno", "valore"])


def estero_markets(code: str, datatype: str = "NI", year: int | None = None,
                   only_countries: bool = True, top: int | None = None):
    """Classifica dei mercati esteri per la regione (o Italia): DataFrame
    country·nome·valore per l'anno scelto (o l'ultimo disponibile), ordinata desc.
    datatype: 'NI' presenze, 'AR' arrivi. only_countries esclude gli aggregati."""
    df = estero_regione_long()
    if df is None:
        return None
    area = RG.istat_area(code)
    d = df[(df["area"] == area) & (df["datatype"] == datatype)]
    if only_countries:
        d = d[~d["country"].map(_tur9_is_aggregate)]
    if d.empty:
        return None
    if year is None:
        year = int(d["anno"].max())
    d = d[d["anno"] == year].copy()
    if d.empty:
        return None
    d["nome"] = d["country"].map(estero_country_name)
    return d.sort_values("valore", ascending=False)[["country", "nome", "valore"]] \
            .head(top).reset_index(drop=True) if top else \
            d.sort_values("valore", ascending=False)[["country", "nome", "valore"]] \
            .reset_index(drop=True)


def estero_country_series(code: str, country: str, datatype: str = "NI"):
    """Serie storica annuale di un paese in una regione: DataFrame anno·valore."""
    df = estero_regione_long()
    if df is None:
        return None
    area = RG.istat_area(code)
    d = df[(df["area"] == area) & (df["datatype"] == datatype) &
           (df["country"] == country)]
    if d.empty:
        return None
    return d.sort_values("anno")[["anno", "valore"]].reset_index(drop=True)


def estero_years(code: str, datatype: str = "NI"):
    """Anni disponibili per la regione/territorio (lista ordinata di int)."""
    df = estero_regione_long()
    if df is None:
        return []
    area = RG.istat_area(code)
    d = df[(df["area"] == area) & (df["datatype"] == datatype)]
    return sorted(int(a) for a in d["anno"].dropna().unique())


# --- Banca d'Italia per PAESE di origine (nazionale, trimestrale 1997-2025) ---
_BDI_COLMAP = {4: "DE", 5: "FR", 6: "AT", 7: "ES", 9: "GB", 10: "CH", 11: "RU",
               13: "US", 14: "CA", 17: "JP"}
_BDI_SHEETS = {"notti": "TS1-N-S", "spesa": "TS1-S-S", "viaggiatori": "TS1-V-S"}
BDI_MARKETS = ["DE", "AT", "GB", "CH", "US", "FR", "ES"]


@st.cache_data(show_spinner=False)
def bdi_country_long():
    """Flussi nazionali BdI per paese di origine: DataFrame date·code·notti·spesa·viaggiatori
    (trimestrale). Sorgente: fogli TS1 dell'xlsx BdI. None se il file manca."""
    import re
    p = ".cache/bdi_turismo_ts.xlsx"
    if not os.path.exists(p):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    recs = []
    for metric, sheet in _BDI_SHEETS.items():
        rows = list(wb[sheet].iter_rows(values_only=True))
        h = next((i for i, r in enumerate(rows) if r and any(c == "Germania" for c in r if c)), None)
        if h is None:
            continue
        year = None
        for r in rows[h + 1:]:
            if r[0] is not None:
                year = int(r[0])
            m = re.match(r"\s*(\d)", str(r[1] or r[2] or ""))
            if not (year and m):
                continue
            q = int(m.group(1))
            date = pd.Timestamp(year, q * 3 - 2, 1)
            for col, code in _BDI_COLMAP.items():
                if col < len(r) and r[col] is not None:
                    recs.append((date, code, metric, float(r[col])))
    if not recs:
        return None
    df = pd.DataFrame(recs, columns=["date", "code", "metric", "value"])
    return df.pivot_table(index=["date", "code"], columns="metric", values="value").reset_index()


def bdi_country_annual():
    """Aggregato ANNUALE per paese (notti/viaggiatori = somma, spesa = somma). Per le viste descrittive."""
    df = bdi_country_long()
    if df is None or df.empty:
        return None
    df = df.copy(); df["anno"] = df["date"].dt.year
    g = df.groupby(["anno", "code"]).agg(
        notti=("notti", "sum"), spesa=("spesa", "sum"), viaggiatori=("viaggiatori", "sum")).reset_index()
    return g[g["code"].isin(BDI_MARKETS)]


# --- Spesa STIMATA dei turisti stranieri per mercato × regione (ISTAT _9 × BdI €/notte) ---
# ISTAT country code -> BdI country code (per la spesa media a notte nazionale)
_ISTAT_TO_BDI = {"DE": "DE", "FR": "FR", "AT": "AT", "ES": "ES", "US": "US",
                 "CA": "CA", "JP": "JP", "RU": "RU", "GB": "GB", "UK": "GB",
                 "CH": "CH", "CH_LI": "CH"}


def bdi_spend_per_night(year: int) -> dict:
    """€/notte per paese (BdI) nell'anno indicato (solo anni completi, 4 trimestri):
    spesa in M€ / notti in migliaia → €/notte. {} se dati assenti."""
    long = bdi_country_long()
    if long is None or long.empty:
        return {}
    d = long.copy(); d["anno"] = d["date"].dt.year
    d = d[d["anno"] == year]
    if d.empty:
        return {}
    g = d.groupby("code").agg(spesa=("spesa", "sum"), notti=("notti", "sum"), q=("date", "count"))
    return {code: r["spesa"] * 1e6 / (r["notti"] * 1e3)
            for code, r in g.iterrows() if r["q"] >= 4 and r["notti"] > 0}


def estero_spesa_stimata(code: str, year: int | None = None, top: int | None = None):
    """STIMA della spesa dei turisti stranieri per mercato in una regione:
    notti reali (ISTAT cube _9) × spesa media a notte del mercato (BdI, nazionale).
    DataFrame country·nome·notti·eur_notte·spesa_m (M€), ordinata desc; solo i mercati
    presenti in ENTRAMBE le fonti. None se manca una delle due. Anno scelto in .attrs['anno']."""
    pres = estero_regione_long()
    if pres is None:
        return None
    area = RG.istat_area(code)
    p = pres[(pres["area"] == area) & (pres["datatype"] == "NI")]
    if p.empty:
        return None
    # anno: il più recente con dati ISTAT E spesa/notte BdI (anno completo)
    years = [year] if year else sorted((int(a) for a in p["anno"].dropna().unique()), reverse=True)
    chosen, spn = None, {}
    for y in years:
        spn = bdi_spend_per_night(y)
        if spn:
            chosen = y; break
    if chosen is None:
        return None
    rows = []
    for _, r in p[p["anno"] == chosen].iterrows():
        bdi = _ISTAT_TO_BDI.get(r["country"])
        if bdi and bdi in spn:
            notti = float(r["valore"]); eurn = spn[bdi]
            rows.append({"country": r["country"], "nome": estero_country_name(r["country"]),
                         "notti": notti, "eur_notte": eurn, "spesa_m": notti * eurn / 1e6})
    if not rows:
        return None
    out = pd.DataFrame(rows).sort_values("spesa_m", ascending=False).reset_index(drop=True)
    out = out.head(top) if top else out
    out.attrs["anno"] = chosen
    return out


def chart_estero_spesa_stimata(code: str, year: int | None = None) -> go.Figure | None:
    d = estero_spesa_stimata(code, year=year, top=12)
    if d is None or d.empty:
        return None
    d = d.sort_values("spesa_m")
    fig = go.Figure(go.Bar(x=d["spesa_m"], y=d["nome"], orientation="h", marker_color="#0e6b70",
                           hovertemplate="<b>%{y}</b><br>%{x:,.1f} M€ stimati<extra></extra>"))
    fig.update_xaxes(title="spesa stimata turisti stranieri (M€/anno)")
    return _layout(fig, h=380)


def estero_markets_table(code: str, datatype: str = "NI", year: int | None = None,
                         top: int | None = 12):
    """Tabella mercati esteri REALI per regione (ISTAT _9): country·nome·valore·quota·yoy.
    quota = % sul totale esteri (WRL_X_ITA) dell'anno; yoy = var% vs anno precedente.
    Anno usato in .attrs['anno']. None se dati assenti."""
    df = estero_regione_long()
    if df is None:
        return None
    area = RG.istat_area(code)
    base = df[(df["area"] == area) & (df["datatype"] == datatype)]
    if base.empty:
        return None
    yrs = sorted(int(a) for a in base["anno"].dropna().unique())
    if not yrs:
        return None
    y = year or yrs[-1]
    cur = base[base["anno"] == y]
    prev = base[base["anno"] == y - 1].set_index("country")["valore"].to_dict()
    tot = float(cur[cur["country"] == "WRL_X_ITA"]["valore"].sum()) or None
    rows = []
    for _, r in cur.iterrows():
        c = r["country"]
        if _tur9_is_aggregate(c):
            continue
        v = float(r["valore"]); pv = prev.get(c)
        rows.append({"country": c, "nome": estero_country_name(c), "valore": v,
                     "quota": (v / tot * 100 if tot else None),
                     "yoy": ((v / float(pv) - 1) * 100 if pv else None)})
    if not rows:
        return None
    out = pd.DataFrame(rows).sort_values("valore", ascending=False).reset_index(drop=True)
    out.attrs["anno"] = y
    return out.head(top) if top else out


def chart_estero_markets(code: str, datatype: str = "NI", year: int | None = None,
                         top: int = 12) -> go.Figure | None:
    d = estero_markets_table(code, datatype=datatype, year=year, top=top)
    if d is None or d.empty:
        return None
    d = d.sort_values("valore")
    unit = "presenze" if datatype == "NI" else "arrivi"
    fig = go.Figure(go.Bar(x=d["valore"], y=d["nome"], orientation="h", marker_color="#0e6b70",
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} " + unit + "<extra></extra>"))
    fig.update_xaxes(title=f"{unit} di turisti esteri")
    return _layout(fig, h=420)


def chart_estero_trend(code: str, countries: list[str], datatype: str = "NI") -> go.Figure | None:
    """Trend storico (2008→) dei mercati scelti per la regione. None se nessuna serie."""
    if not countries:
        return None
    fig = go.Figure()
    for c in countries:
        s = estero_country_series(code, c, datatype=datatype)
        if s is not None and not s.empty:
            fig.add_trace(go.Scatter(x=s["anno"], y=s["valore"], mode="lines+markers",
                                     name=estero_country_name(c)))
    if not fig.data:
        return None
    unit = "presenze" if datatype == "NI" else "arrivi"
    fig.update_yaxes(title=f"{unit} di turisti esteri")
    fig.update_xaxes(dtick=1, title=None)
    return _layout(fig, h=380)


# ════════════════════════════════════════════════════════════════════════════
# ALTO ADIGE / BOLZANO (ASTAT) — fonte COMPLEMENTARE di una sola provincia.
#   flussi MENSILI (arrivi + presenze, 1990→) e capacità ANNUALE per categoria
#   (esercizi + posti letto). Niente paese di origine: non alimenta il motore,
#   serve per stagionalità tempestiva + lato offerta. [[tdh-engine-project]]
# ════════════════════════════════════════════════════════════════════════════
# categorie di DETTAGLIO (escludo gli aggregati per non contare due volte)
_ASTAT_DETAIL_CATS = ["4-5 stelle", "3 stelle", "1-2 stelle", "Residence",
                      "Campeggi", "Alloggi privati", "Agriturismi", "Altri esercizi"]


@st.cache_data(show_spinner=False)
def astat_flussi() -> pd.DataFrame:
    """Flussi mensili Alto Adige: date·arrivi·presenze. Vuoto se cache assente."""
    try:
        return RS.fetch_astat_flussi_mensili(refresh=False)
    except Exception:  # noqa: BLE001
        return pd.DataFrame(columns=["date", "arrivi", "presenze"])


@st.cache_data(show_spinner=False)
def astat_capacita() -> pd.DataFrame:
    """Capacità ricettiva annuale Alto Adige: anno·categoria·esercizi·posti_letto."""
    try:
        return RS.fetch_astat_capacita_categoria(refresh=False)
    except Exception:  # noqa: BLE001
        return pd.DataFrame(columns=["anno", "cat_code", "categoria", "esercizi", "posti_letto"])


def astat_kpi() -> dict:
    """KPI sintetici per l'header: ultimo anno solare COMPLETO, YoY presenze, offerta."""
    f = astat_flussi()
    out = {"mese_ultimo": None, "presenze_anno": None, "yoy": None,
           "arrivi_anno": None, "anno": None, "esercizi": None, "posti_letto": None}
    if not f.empty:
        f = f.copy()
        f["anno"] = f["date"].dt.year
        out["mese_ultimo"] = f["date"].max().strftime("%m/%Y")
        # ultimo anno con 12 mesi presenti (anno solare completo)
        full = f.groupby("anno").size()
        completi = [y for y, n in full.items() if n >= 12]
        if completi:
            y = max(completi)
            out["anno"] = int(y)
            out["presenze_anno"] = float(f[f["anno"] == y]["presenze"].sum())
            out["arrivi_anno"] = float(f[f["anno"] == y]["arrivi"].sum())
            prev = f[f["anno"] == y - 1]["presenze"].sum()
            if prev:
                out["yoy"] = (out["presenze_anno"] - prev) / prev * 100.0
    c = astat_capacita()
    if not c.empty:
        last = c[c["anno"] == c["anno"].max()]
        tot = last[last["categoria"] == "Totale"]
        if not tot.empty:
            out["esercizi"] = float(tot["esercizi"].iloc[0])
            out["posti_letto"] = float(tot["posti_letto"].iloc[0])
    return out


def chart_astat_flussi(indicator: str = "presenze", start_year: int = 2015) -> go.Figure | None:
    """Serie mensile (presenze o arrivi) dell'Alto Adige da start_year."""
    f = astat_flussi()
    if f.empty:
        return None
    f = f[f["date"].dt.year >= start_year]
    if f.empty or indicator not in f.columns:
        return None
    label = "Presenze" if indicator == "presenze" else "Arrivi"
    fig = go.Figure(go.Scatter(x=f["date"], y=f[indicator], mode="lines",
                               line=dict(color="#0e6b70", width=2),
                               fill="tozeroy", fillcolor="rgba(14,116,144,0.08)",
                               hovertemplate="%{x|%m/%Y}<br>%{y:,.0f} " + label.lower() + "<extra></extra>"))
    fig.update_yaxes(title=f"{label} · Alto Adige")
    fig.update_xaxes(title=None)
    return _layout(fig, h=360)


def chart_astat_stagionalita(indicator: str = "presenze", years: int = 3) -> go.Figure | None:
    """Profilo stagionale: media per mese dell'anno sugli ultimi `years` anni completi."""
    f = astat_flussi()
    if f.empty or indicator not in f.columns:
        return None
    f = f.copy()
    f["anno"] = f["date"].dt.year
    f["mese"] = f["date"].dt.month
    yrs = sorted(f["anno"].unique())[-years:]
    f = f[f["anno"].isin(yrs)]
    prof = f.groupby("mese")[indicator].mean().reindex(range(1, 13))
    mesi = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]
    label = "presenze" if indicator == "presenze" else "arrivi"
    fig = go.Figure(go.Bar(x=mesi, y=prof.values, marker_color="#0e6b70",
                           hovertemplate="%{x}<br>%{y:,.0f} " + label + " (media)<extra></extra>"))
    fig.update_yaxes(title=f"{label.capitalize()} · media mensile")
    fig.update_xaxes(title=None)
    return _layout(fig, h=340)


def astat_capacita_table(year: int | None = None) -> pd.DataFrame:
    """Capacità per categoria di dettaglio (un anno). Colonne pronte per la vista."""
    c = astat_capacita()
    if c.empty:
        return c
    y = year or int(c["anno"].max())
    d = c[(c["anno"] == y) & (c["categoria"].isin(_ASTAT_DETAIL_CATS))].copy()
    d["categoria"] = pd.Categorical(d["categoria"], _ASTAT_DETAIL_CATS, ordered=True)
    d = d.sort_values("categoria")
    return d[["categoria", "esercizi", "posti_letto"]].reset_index(drop=True)


def chart_astat_capacita(year: int | None = None) -> go.Figure | None:
    """Posti letto per categoria ricettiva (un anno) — lato offerta."""
    d = astat_capacita_table(year)
    if d is None or d.empty:
        return None
    d = d.sort_values("posti_letto")
    fig = go.Figure(go.Bar(x=d["posti_letto"], y=d["categoria"], orientation="h",
                           marker_color="#f59e0b",
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} posti letto<extra></extra>"))
    fig.update_xaxes(title="Posti letto")
    return _layout(fig, h=380)


def _last_date_of(path: str):
    """Ultimo timestamp della colonna 'date' di un CSV in cache. None se assente/vuoto."""
    try:
        d = pd.read_csv(path, usecols=["date"])
        t = pd.to_datetime(d["date"], errors="coerce").dropna()
        return t.max() if len(t) else None
    except Exception:  # noqa: BLE001
        return None


def region_data_catalog(code: str, ov: dict | None = None) -> list[dict]:
    """Catalogo delle fonti che coprono la regione selezionata (OFFLINE dove possibile).
    Righe: Fonte·Dato·Granularità·Ultimo·Copertura·Stato (🟢/🟡/🔴). `ov` = region_overview
    già calcolato per lo snapshot (evita di rifare le chiamate ISTAT live)."""
    area = RG.istat_area(code)
    rows: list[dict] = []

    def add(fonte, dato, gran, ultimo, cov, stato):
        rows.append({"Fonte": fonte, "Dato": dato, "Granularità": gran,
                     "Ultimo": ultimo, "Copertura": cov, "Stato": stato})

    # 1) ISTAT presenze totali/straniere (mensile, region-aware, via live) — riuso `ov`
    last_pres, nrp = "—", "tutte le regioni"
    if ov is not None and isinstance(ov.get("presenze"), pd.DataFrame) and not ov["presenze"].empty:
        df = ov["presenze"]
        obs = df.dropna(subset=["stranieri"]) if "stranieri" in df else df
        if not obs.empty:
            last_pres = obs["date"].max().strftime("%Y-%m")
    add("ISTAT · Movimento clienti (_7)", "Presenze totali e straniere", "mensile",
        last_pres, nrp, "🟢")

    # 2) ISTAT esteri per singolo PAESE × territorio (_9, annuale, OFFLINE)
    d9 = estero_regione_long()
    if d9 is not None and (s := d9[d9["area"] == area]) is not None and not s.empty:
        paesi = s[~s["country"].map(_tur9_is_aggregate)]["country"].nunique()
        add("ISTAT · _9 (paese × territorio)", "Esteri per singolo paese", "annuale",
            str(int(s["anno"].max())), f"{paesi} paesi", "🟢")
    else:
        add("ISTAT · _9 (paese × territorio)", "Esteri per singolo paese", "annuale",
            "—", "non per questo territorio", "🔴")

    # 3) ISTAT capacità (posti letto, annuale, via live) — riuso `ov`
    cap_last = str(ov["anno_letti"]) if (ov and ov.get("anno_letti")) else "—"
    add("ISTAT · Capacità esercizi", "Posti letto", "annuale", cap_last,
        "regionale", "🟢" if cap_last != "—" else "🟡")

    # 4) Banca d'Italia — spesa turisti stranieri (trimestrale, OFFLINE)
    g = bdi_region_annual(code)
    if g is not None and not g.empty:
        add("Banca d'Italia · turismo int.", "Spesa turisti stranieri", "trimestrale",
            str(int(g["anno"].max())), "totale regione", "🟢")
    else:
        add("Banca d'Italia · turismo int.", "Spesa turisti stranieri", "trimestrale",
            "—", "—", "🔴")

    # 5) Google Trends (mensile, OFFLINE — conta i file presenti per la regione)
    kw = _safe_kw(RG.region(code)["trends_kw"])
    geos = [mk.code for mk in DEFAULT_MARKETS]
    present = [p for p in (f".cache/trends_{kw}_{gg}.csv" for gg in geos) if os.path.exists(p)]
    tlast = "—"
    if present:
        r = _last_date_of(present[0])
        tlast = r.strftime("%Y-%m") if r is not None else "—"
    add("Google Trends", "Interesse di ricerca per mercato", "mensile", tlast,
        f"{len(present)}/{len(geos)} mercati",
        "🟢" if len(present) == len(geos) else ("🟡" if present else "🔴"))

    # 6) Eurostat — connettività aerea (annuale, OFFLINE file-check)
    airports = RG.region(code)["airports"]
    if airports and os.path.exists(f".cache/conn_{code}.csv"):
        add("Eurostat · avia_par", "Passeggeri voli diretti per mercato", "annuale",
            "2024", ", ".join(airports), "🟢")
    elif airports:
        add("Eurostat · avia_par", "Passeggeri voli diretti per mercato", "annuale",
            "—", ", ".join(airports), "🟡")
    else:
        add("Eurostat · avia_par", "Passeggeri voli diretti per mercato", "annuale",
            "—", "nessun aeroporto mappato", "🔴")

    # 6b) ISTAT Viaggi&Vacanze — domanda dei residenti per scopo (annuale, OFFLINE)
    tp = turnot_purpose()
    if not tp.empty:
        s = tp[tp["area"] == area]
        if not s.empty:
            add("ISTAT · Viaggi e Vacanze", "Notti residenti per scopo del viaggio", "annuale",
                str(int(s["anno"].max())), "domanda · tutte le regioni", "🟢")

    # 7) Fonte LOCALE complementare (solo dove esiste)
    if code == "ITD1" and os.path.exists(".cache/astat_bolzano_flussi_mensili.csv"):
        r = _last_date_of(".cache/astat_bolzano_flussi_mensili.csv")
        add("ASTAT · Bolzano (LOCALE)", "Flussi mensili + capacità per categoria", "mensile",
            r.strftime("%Y-%m") if r is not None else "—", "provincia di Bolzano", "🟢")
    elif code == "ITC4" and os.path.exists(_LOMB_PATH):
        try:
            yy = int(pd.read_csv(_LOMB_PATH, usecols=["anno"])["anno"].max())
        except Exception:  # noqa: BLE001
            yy = None
        add("Open Data Lombardia (LOCALE)", "Flussi mensili per provincia × mercato estero", "mensile",
            str(yy) if yy else "—", "12 province lombarde", "🟢")
    elif code == "ITE1" and os.path.exists(_TOSC_PATH):
        try:
            yy = int(pd.read_csv(_TOSC_PATH, usecols=["anno"])["anno"].max())
        except Exception:  # noqa: BLE001
            yy = None
        add("Open Data Regione Toscana (LOCALE)", "Movimento per comune (italiani/stranieri)", "annuale",
            str(yy) if yy else "—", "~272 comuni, 10 province", "🟢")
    elif code == "ITG2" and os.path.exists(_SARD_CM_PATH):
        try:
            yy = int(pd.read_csv(_SARD_CM_PATH, usecols=["anno"])["anno"].max())
        except Exception:  # noqa: BLE001
            yy = None
        add("Osservatorio Sardegna (LOCALE)", "Movimento mensile per comune × mercato estero", "mensile",
            str(yy) if yy else "—", "~330 comuni; export manuale", "🟢")
    return rows


def chart_region_letti(code: str) -> go.Figure | None:
    """Posti letto per anno della regione (ISTAT capacità). None se non disponibile."""
    panel = region_annual_panel(code)
    if panel is None or panel.empty or "letti" not in panel.columns:
        return None
    d = panel["letti"].dropna()
    if d.empty:
        return None
    fig = go.Figure(go.Bar(x=d.index.astype(int), y=d.values, marker_color="#16a34a",
                           hovertemplate="<b>%{x}</b><br>%{y:,.0f} posti letto<extra></extra>"))
    fig.update_yaxes(title="posti letto")
    fig.update_xaxes(dtick=1, title=None)
    return _layout(fig, h=340)


# ════════════════════════════════════════════════════════════════════════════
# LOMBARDIA (Open Data Lombardia, Socrata) — fonte COMPLEMENTARE regionale (ITC4).
#   flussi MENSILI per PROVINCIA × PROVENIENZA (regioni IT + paesi esteri) × categoria.
#   Unico nel motore: MERCATO ESTERO MENSILE a livello sub-nazionale. [[tdh-engine-project]]
# ════════════════════════════════════════════════════════════════════════════
_LOMB_PATH = ".cache/lombardia_flussi_provincia_mese.csv"
_LOMB_MESI = {"Gennaio": 1, "Febbraio": 2, "Marzo": 3, "Aprile": 4, "Maggio": 5, "Giugno": 6,
              "Luglio": 7, "Agosto": 8, "Settembre": 9, "Ottobre": 10, "Novembre": 11, "Dicembre": 12}
# provenienze italiane (regioni) e non specificate → NON sono "mercato estero"
_LOMB_ITA = {"ABRUZZO", "BASILICATA", "BOLZANO - BOZEN", "CALABRIA", "CAMPANIA", "EMILIA-ROMAGNA",
             "FRIULI-VENEZIA GIULIA", "LAZIO", "LIGURIA", "LOMBARDIA", "MARCHE", "MOLISE", "PIEMONTE",
             "PUGLIA", "SARDEGNA", "SICILIA", "TOSCANA", "TRENTO", "UMBRIA", "VALLE D'AOSTA", "VENETO"}
_LOMB_UNSPEC = {"Non specificato", "Regione non specificata", "Paese estero non specificato"}


def _lomb_is_foreign(p: str) -> bool:
    """True se la provenienza è un PAESE ESTERO (non regione italiana né 'non specificato')."""
    return p not in _LOMB_ITA and p not in _LOMB_UNSPEC


@st.cache_data(show_spinner=False)
def lomb_flussi() -> pd.DataFrame:
    """Flussi mensili Lombardia: provincia·mese·provenienza·arrivi·presenze + colonna date.
    Vuoto se cache assente."""
    if not os.path.exists(_LOMB_PATH):
        return pd.DataFrame()
    df = pd.read_csv(_LOMB_PATH)
    df = df[df["mese"].isin(_LOMB_MESI)].copy()
    df["m"] = df["mese"].map(_LOMB_MESI)
    df["date"] = pd.to_datetime(dict(year=df["anno"], month=df["m"], day=1))
    return df


def lomb_provinces() -> list[str]:
    df = lomb_flussi()
    return sorted(df["provincia"].dropna().unique()) if not df.empty else []


def _lomb_scope(prov: str | None):
    """Sottoinsieme per provincia (o tutta la Lombardia se prov None/'Tutta la Lombardia')."""
    df = lomb_flussi()
    if df.empty:
        return df
    if prov and prov != "Tutta la Lombardia":
        df = df[df["provincia"] == prov]
    return df


def lomb_monthly_totals(prov: str | None = None) -> pd.DataFrame:
    """Totale mensile (tutte le provenienze) per provincia/regione: date·arrivi·presenze."""
    df = _lomb_scope(prov)
    if df.empty:
        return df
    g = df.groupby("date").agg(arrivi=("arrivi_totale", "sum"),
                               presenze=("presenze_totale", "sum")).reset_index()
    return g.sort_values("date")


def lomb_markets_table(prov: str | None = None, year: int | None = None,
                       top: int | None = 12) -> pd.DataFrame:
    """Mercati ESTERI per provincia/anno: nome·presenze·arrivi·quota% (sul totale estero)."""
    df = _lomb_scope(prov)
    if df.empty:
        return df
    if year is None:
        year = int(df["anno"].max())
    d = df[(df["anno"] == year) & (df["provenienza_turisti"].map(_lomb_is_foreign))]
    if d.empty:
        return d
    g = d.groupby("provenienza_turisti").agg(presenze=("presenze_totale", "sum"),
                                             arrivi=("arrivi_totale", "sum")).reset_index()
    g = g.rename(columns={"provenienza_turisti": "nome"}).sort_values("presenze", ascending=False)
    tot = g["presenze"].sum()
    g["quota"] = (g["presenze"] / tot * 100) if tot else 0.0
    return g.head(top).reset_index(drop=True) if top else g.reset_index(drop=True)


def lomb_kpi(prov: str | None = None) -> dict:
    """KPI per l'header del blocco: presenze/arrivi ultimo anno, quota estero, top mercato estero."""
    df = _lomb_scope(prov)
    out = {"anno": None, "presenze": None, "arrivi": None, "quota_estero": None, "top_estero": None}
    if df.empty:
        return out
    y = int(df["anno"].max()); out["anno"] = y
    dy = df[df["anno"] == y]
    out["presenze"] = float(dy["presenze_totale"].sum())
    out["arrivi"] = float(dy["arrivi_totale"].sum())
    est = dy[dy["provenienza_turisti"].map(_lomb_is_foreign)]["presenze_totale"].sum()
    out["quota_estero"] = (est / out["presenze"] * 100) if out["presenze"] else None
    mk = lomb_markets_table(prov, year=y, top=1)
    if mk is not None and not mk.empty:
        out["top_estero"] = f"{mk['nome'].iloc[0]} ({mk['quota'].iloc[0]:.0f}%)"
    return out


def chart_lomb_flussi_mensili(prov: str | None = None) -> go.Figure | None:
    g = lomb_monthly_totals(prov)
    if g is None or g.empty:
        return None
    fig = go.Figure(go.Scatter(x=g["date"], y=g["presenze"], mode="lines",
                               line=dict(color="#0e6b70", width=2),
                               fill="tozeroy", fillcolor="rgba(14,116,144,0.08)",
                               hovertemplate="%{x|%m/%Y}<br>%{y:,.0f} presenze<extra></extra>"))
    fig.update_yaxes(title="presenze (mese)")
    fig.update_xaxes(title=None)
    return _layout(fig, h=340)


def chart_lomb_markets(prov: str | None = None, year: int | None = None,
                       top: int = 12) -> go.Figure | None:
    d = lomb_markets_table(prov, year=year, top=top)
    if d is None or d.empty:
        return None
    d = d.sort_values("presenze")
    fig = go.Figure(go.Bar(x=d["presenze"], y=d["nome"], orientation="h", marker_color="#f59e0b",
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} presenze<extra></extra>"))
    fig.update_xaxes(title="presenze di turisti esteri")
    return _layout(fig, h=420)


# ════════════════════════════════════════════════════════════════════════════
# TOSCANA (Open Data Regione Toscana, CKAN) — fonte COMPLEMENTARE regionale (ITE1).
#   movimento clienti ANNUALE per COMUNE (~272 comuni) × ambito turistico, split
#   italiani/stranieri, 2018-2025. Unico nel motore: granularità COMUNALE. [[tdh-engine-project]]
# ════════════════════════════════════════════════════════════════════════════
_TOSC_PATH = ".cache/toscana_movimento_comune_anno.csv"
_TOSC_PROV = {"AR": "Arezzo", "FI": "Firenze", "GR": "Grosseto", "LI": "Livorno",
              "LU": "Lucca", "MS": "Massa-Carrara", "PI": "Pisa", "PO": "Prato",
              "PT": "Pistoia", "SI": "Siena"}


@st.cache_data(show_spinner=False)
def tosc_movimento() -> pd.DataFrame:
    """Movimento clienti Toscana per comune×anno + colonne derivate arrivi_tot/presenze_tot.
    Vuoto se cache assente."""
    if not os.path.exists(_TOSC_PATH):
        return pd.DataFrame()
    df = pd.read_csv(_TOSC_PATH)
    df["arrivi_tot"] = df["arrivi_italiani"].fillna(0) + df["arrivi_stranieri"].fillna(0)
    df["presenze_tot"] = df["presenze_italiane"].fillna(0) + df["presenze_straniere"].fillna(0)
    return df


def tosc_provinces() -> list[str]:
    """Nomi estesi delle province presenti (per il selettore)."""
    df = tosc_movimento()
    if df.empty:
        return []
    sig = sorted(df["sigla_provincia"].dropna().unique())
    return [_TOSC_PROV.get(s, s) for s in sig]


def _tosc_scope(prov: str | None) -> pd.DataFrame:
    """Sottoinsieme per provincia (nome esteso) o tutta la Toscana se None/'Tutta la Toscana'."""
    df = tosc_movimento()
    if df.empty:
        return df
    if prov and prov != "Tutta la Toscana":
        sig = {v: k for k, v in _TOSC_PROV.items()}.get(prov, prov)
        df = df[df["sigla_provincia"] == sig]
    return df


def tosc_yearly_totals(prov: str | None = None) -> pd.DataFrame:
    """Serie annuale: anno·presenze_italiane·presenze_straniere·arrivi_tot·presenze_tot."""
    df = _tosc_scope(prov)
    if df.empty:
        return df
    g = df.groupby("anno").agg(presenze_italiane=("presenze_italiane", "sum"),
                               presenze_straniere=("presenze_straniere", "sum"),
                               arrivi_tot=("arrivi_tot", "sum"),
                               presenze_tot=("presenze_tot", "sum")).reset_index()
    return g.sort_values("anno")


def tosc_top_comuni(prov: str | None = None, year: int | None = None,
                    top: int | None = 12) -> pd.DataFrame:
    """Comuni per presenze in un anno: comune·presenze·arrivi·quota_estero%·quota% sul totale."""
    df = _tosc_scope(prov)
    if df.empty:
        return df
    if year is None:
        year = int(df["anno"].max())
    d = df[df["anno"] == year].copy()
    if d.empty:
        return d
    d = d.rename(columns={"comune": "nome", "presenze_tot": "presenze", "arrivi_tot": "arrivi"})
    d["quota_estero"] = (d["presenze_straniere"] / d["presenze"] * 100).where(d["presenze"] > 0, 0.0)
    d = d.sort_values("presenze", ascending=False)
    tot = d["presenze"].sum()
    d["quota"] = (d["presenze"] / tot * 100) if tot else 0.0
    cols = ["nome", "presenze", "arrivi", "quota_estero", "quota"]
    return (d[cols].head(top) if top else d[cols]).reset_index(drop=True)


def tosc_ambiti(prov: str | None = None, year: int | None = None,
                top: int | None = 12) -> pd.DataFrame:
    """Presenze per AMBITO turistico in un anno: ambito·presenze (ordinate)."""
    df = _tosc_scope(prov)
    if df.empty:
        return df
    if year is None:
        year = int(df["anno"].max())
    d = df[(df["anno"] == year) & (df["ambito"].astype(str).str.strip() != "")]
    if d.empty:
        return d
    g = d.groupby(d["ambito"].astype(str).str.strip()).agg(
        presenze=("presenze_tot", "sum")).reset_index()
    g = g.rename(columns={"ambito": "nome"}).sort_values("presenze", ascending=False)
    return (g.head(top) if top else g).reset_index(drop=True)


def tosc_kpi(prov: str | None = None) -> dict:
    """KPI header: anno·presenze·arrivi·quota estero·top comune·n. comuni."""
    df = _tosc_scope(prov)
    out = {"anno": None, "presenze": None, "arrivi": None, "quota_estero": None,
           "top_comune": None, "n_comuni": None}
    if df.empty:
        return out
    y = int(df["anno"].max()); out["anno"] = y
    dy = df[df["anno"] == y]
    out["presenze"] = float(dy["presenze_tot"].sum())
    out["arrivi"] = float(dy["arrivi_tot"].sum())
    est = float(dy["presenze_straniere"].sum())
    out["quota_estero"] = (est / out["presenze"] * 100) if out["presenze"] else None
    out["n_comuni"] = int(dy["comune"].nunique())
    tc = tosc_top_comuni(prov, year=y, top=1)
    if tc is not None and not tc.empty:
        out["top_comune"] = f"{tc['nome'].iloc[0]} ({tc['quota'].iloc[0]:.0f}%)"
    return out


def chart_tosc_yearly(prov: str | None = None) -> go.Figure | None:
    """Presenze annuali italiani vs stranieri (barre impilate)."""
    g = tosc_yearly_totals(prov)
    if g is None or g.empty:
        return None
    fig = go.Figure()
    fig.add_bar(x=g["anno"], y=g["presenze_italiane"], name="italiani", marker_color="#0e6b70",
                hovertemplate="<b>%{x}</b><br>%{y:,.0f} presenze italiani<extra></extra>")
    fig.add_bar(x=g["anno"], y=g["presenze_straniere"], name="stranieri", marker_color="#f59e0b",
                hovertemplate="<b>%{x}</b><br>%{y:,.0f} presenze stranieri<extra></extra>")
    fig.update_layout(barmode="stack", legend=dict(orientation="h", y=1.12, x=0))
    fig.update_yaxes(title="presenze")
    fig.update_xaxes(dtick=1, title=None)
    return _layout(fig, h=340)


def chart_tosc_top_comuni(prov: str | None = None, year: int | None = None,
                          top: int = 12) -> go.Figure | None:
    d = tosc_top_comuni(prov, year=year, top=top)
    if d is None or d.empty:
        return None
    d = d.sort_values("presenze")
    fig = go.Figure(go.Bar(x=d["presenze"], y=d["nome"], orientation="h", marker_color="#0e6b70",
                           customdata=d["quota_estero"],
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} presenze<br>"
                                         "estero %{customdata:.0f}%<extra></extra>"))
    fig.update_xaxes(title="presenze totali")
    return _layout(fig, h=420)


def chart_tosc_ambiti(prov: str | None = None, year: int | None = None,
                      top: int = 12) -> go.Figure | None:
    d = tosc_ambiti(prov, year=year, top=top)
    if d is None or d.empty:
        return None
    d = d.sort_values("presenze")
    fig = go.Figure(go.Bar(x=d["presenze"], y=d["nome"], orientation="h", marker_color="#16a34a",
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} presenze<extra></extra>"))
    fig.update_xaxes(title="presenze per ambito turistico")
    return _layout(fig, h=420)


# ════════════════════════════════════════════════════════════════════════════
# SARDEGNA (Osservatorio del Turismo, export manuale CC-BY) — fonte COMPLEMENTARE (ITG2).
#   movimento MENSILE × comune × mercato di provenienza (paese) × macro-tipologia, 2021-2025.
#   La più ricca del motore: mensile (come Lombardia) + comunale (come Toscana) + mercato estero.
#   Cache proiettate: comune×mese (totale+estero) e mercato×mese. [[sardegna-fonti-turismo-accesso]]
# ════════════════════════════════════════════════════════════════════════════
_SARD_CM_PATH = ".cache/sardegna_flussi_comune_mese.csv"
_SARD_MM_PATH = ".cache/sardegna_flussi_mercato_mese.csv"


@st.cache_data(show_spinner=False)
def sard_comune_mese() -> pd.DataFrame:
    """Cache comune×mese Sardegna + colonna date (mese 0 = 'non disponibile' -> NaT)."""
    if not os.path.exists(_SARD_CM_PATH):
        return pd.DataFrame()
    df = pd.read_csv(_SARD_CM_PATH)
    m = pd.to_numeric(df["mese"], errors="coerce")
    df["date"] = pd.to_datetime(dict(year=df["anno"], month=m.where((m >= 1) & (m <= 12), 1), day=1),
                                errors="coerce")
    df.loc[~m.between(1, 12), "date"] = pd.NaT
    return df


@st.cache_data(show_spinner=False)
def sard_mercato_mese() -> pd.DataFrame:
    if not os.path.exists(_SARD_MM_PATH):
        return pd.DataFrame()
    return pd.read_csv(_SARD_MM_PATH)


def sard_provinces() -> list[str]:
    df = sard_comune_mese()
    return sorted(df["provincia"].dropna().unique()) if not df.empty else []


def _sard_scope(df: pd.DataFrame, prov: str | None) -> pd.DataFrame:
    if df.empty:
        return df
    if prov and prov != "Tutta la Sardegna":
        df = df[df["provincia"] == prov]
    return df


def sard_monthly_totals(prov: str | None = None) -> pd.DataFrame:
    """Totale mensile (date·arrivi·presenze·presenze_est) per zona/regione."""
    df = _sard_scope(sard_comune_mese(), prov)
    if df.empty:
        return df
    d = df.dropna(subset=["date"])
    g = d.groupby("date").agg(arrivi=("arrivi", "sum"), presenze=("presenze", "sum"),
                              presenze_est=("presenze_est", "sum")).reset_index()
    return g.sort_values("date")


def sard_top_comuni(prov: str | None = None, year: int | None = None,
                    top: int | None = 12) -> pd.DataFrame:
    df = _sard_scope(sard_comune_mese(), prov)
    if df.empty:
        return df
    if year is None:
        year = int(df["anno"].max())
    d = df[df["anno"] == year].groupby("comune").agg(
        presenze=("presenze", "sum"), arrivi=("arrivi", "sum"),
        presenze_est=("presenze_est", "sum")).reset_index()
    if d.empty:
        return d
    d["quota_estero"] = (d["presenze_est"] / d["presenze"] * 100).where(d["presenze"] > 0, 0.0)
    d = d.sort_values("presenze", ascending=False).rename(columns={"comune": "nome"})
    tot = d["presenze"].sum()
    d["quota"] = (d["presenze"] / tot * 100) if tot else 0.0
    cols = ["nome", "presenze", "arrivi", "quota_estero", "quota"]
    return (d[cols].head(top) if top else d[cols]).reset_index(drop=True)


def sard_markets_table(prov: str | None = None, year: int | None = None,
                       top: int | None = 12) -> pd.DataFrame:
    """Mercati ESTERI (provenienza) per presenze, con quota% sul totale estero."""
    df = _sard_scope(sard_mercato_mese(), prov)
    if df.empty:
        return df
    if year is None:
        year = int(df["anno"].max())
    d = df[(df["anno"] == year) & (df["estero"] == 1)]
    if d.empty:
        return d
    g = d.groupby("provenienza").agg(presenze=("presenze", "sum"),
                                     arrivi=("arrivi", "sum")).reset_index()
    g = g.rename(columns={"provenienza": "nome"}).sort_values("presenze", ascending=False)
    tot = g["presenze"].sum()
    g["quota"] = (g["presenze"] / tot * 100) if tot else 0.0
    return (g.head(top) if top else g).reset_index(drop=True)


def sard_kpi(prov: str | None = None) -> dict:
    df = _sard_scope(sard_comune_mese(), prov)
    out = {"anno": None, "presenze": None, "arrivi": None, "quota_estero": None,
           "top_comune": None, "n_comuni": None}
    if df.empty:
        return out
    y = int(df["anno"].max()); out["anno"] = y
    dy = df[df["anno"] == y]
    out["presenze"] = float(dy["presenze"].sum())
    out["arrivi"] = float(dy["arrivi"].sum())
    est = float(dy["presenze_est"].sum())
    out["quota_estero"] = (est / out["presenze"] * 100) if out["presenze"] else None
    out["n_comuni"] = int(dy["comune"].nunique())
    tc = sard_top_comuni(prov, year=y, top=1)
    if tc is not None and not tc.empty:
        out["top_comune"] = f"{tc['nome'].iloc[0]} ({tc['quota'].iloc[0]:.0f}%)"
    return out


def chart_sard_flussi_mensili(prov: str | None = None) -> go.Figure | None:
    g = sard_monthly_totals(prov)
    if g is None or g.empty:
        return None
    fig = go.Figure(go.Scatter(x=g["date"], y=g["presenze"], mode="lines",
                               line=dict(color="#0e6b70", width=2),
                               fill="tozeroy", fillcolor="rgba(14,116,144,0.08)",
                               hovertemplate="%{x|%m/%Y}<br>%{y:,.0f} presenze<extra></extra>"))
    fig.update_yaxes(title="presenze (mese)")
    fig.update_xaxes(title=None)
    return _layout(fig, h=340)


def chart_sard_markets(prov: str | None = None, year: int | None = None,
                       top: int = 12) -> go.Figure | None:
    d = sard_markets_table(prov, year=year, top=top)
    if d is None or d.empty:
        return None
    d = d.sort_values("presenze")
    fig = go.Figure(go.Bar(x=d["presenze"], y=d["nome"], orientation="h", marker_color="#f59e0b",
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} presenze<extra></extra>"))
    fig.update_xaxes(title="presenze di turisti esteri")
    return _layout(fig, h=440)


def chart_sard_top_comuni(prov: str | None = None, year: int | None = None,
                          top: int = 12) -> go.Figure | None:
    d = sard_top_comuni(prov, year=year, top=top)
    if d is None or d.empty:
        return None
    d = d.sort_values("presenze")
    fig = go.Figure(go.Bar(x=d["presenze"], y=d["nome"], orientation="h", marker_color="#16a34a",
                           customdata=d["quota_estero"],
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} presenze<br>"
                                         "estero %{customdata:.0f}%<extra></extra>"))
    fig.update_xaxes(title="presenze totali")
    return _layout(fig, h=440)


# ════════════════════════════════════════════════════════════════════════════
# DOMANDA RESIDENTI — scopo del viaggio (ISTAT Viaggi&Vacanze, TURNOT_CAPI_1).
#   notti dei RESIDENTI per regione di destinazione × scopo (lavoro / vacanza lunga /
#   breve), annuale, TUTTE le regioni. Universo diverso dal movimento negli esercizi:
#   è la DOMANDA (perché e quanto viaggiano gli italiani). Region-aware via istat_area.
# ════════════════════════════════════════════════════════════════════════════
_TURNOT_PATH = ".cache/istat_turnot_scopo_regione.csv"
_TURNOT_LABEL = {"BUS": "Lavoro", "HOLL": "Vacanza lunga (4+ notti)", "HOLS": "Vacanza breve (1-3 notti)"}


@st.cache_data(show_spinner=False)
def turnot_purpose() -> pd.DataFrame:
    if not os.path.exists(_TURNOT_PATH):
        return pd.DataFrame()
    return pd.read_csv(_TURNOT_PATH)


def turnot_pivot(code: str) -> pd.DataFrame:
    """Pivot anno×scopo (notti) per la regione (o IT). Ricava BUS mancante = TRIP − HOL."""
    df = turnot_purpose()
    if df.empty:
        return df
    d = df[df["area"] == RG.istat_area(code)]
    if d.empty:
        return d
    p = d.pivot_table(index="anno", columns="scopo", values="notti", aggfunc="sum")
    if "TRIP" in p and "HOL" in p:
        if "BUS" not in p:
            p["BUS"] = pd.NA
        p["BUS"] = p["BUS"].fillna(p["TRIP"] - p["HOL"])
    return p


def turnot_kpi(code: str) -> dict:
    """KPI ultimo anno: notti totali residenti, quota vacanza lunga, quota lavoro."""
    out = {"anno": None, "notti": None, "quota_lavoro": None, "quota_vac_lunga": None}
    p = turnot_pivot(code)
    if p is None or p.empty or "TRIP" not in p.columns:
        return out
    y = int(p.index.max()); r = p.loc[y]
    tot = r.get("TRIP")
    if pd.isna(tot) or not tot:
        return out
    out["anno"] = y
    out["notti"] = float(tot)
    if "BUS" in p.columns and pd.notna(r.get("BUS")):
        out["quota_lavoro"] = float(r["BUS"]) / tot * 100
    if "HOLL" in p.columns and pd.notna(r.get("HOLL")):
        out["quota_vac_lunga"] = float(r["HOLL"]) / tot * 100
    return out


def chart_turnot_purpose(code: str) -> go.Figure | None:
    """Barre impilate per anno: notti dei residenti per scopo (Lavoro/Vacanza breve/lunga)."""
    p = turnot_pivot(code)
    if p is None or p.empty:
        return None
    idx = p.index.astype(int)
    fig = go.Figure()
    for scopo, color in [("BUS", "#0e6b70"), ("HOLS", "#f59e0b"), ("HOLL", "#16a34a")]:
        if scopo in p.columns:
            fig.add_bar(x=idx, y=p[scopo], name=_TURNOT_LABEL[scopo], marker_color=color,
                        hovertemplate="<b>%{x}</b><br>" + _TURNOT_LABEL[scopo]
                                      + ": %{y:,.0f} notti<extra></extra>")
    if not fig.data:
        return None
    fig.update_layout(barmode="stack", legend=dict(orientation="h", y=1.12, x=0))
    fig.update_yaxes(title="notti dei residenti")
    fig.update_xaxes(dtick=1, title=None)
    return _layout(fig, h=340)


# ════════════════════════════════════════════════════════════════════════════
# AFFITTI BREVI (STR) — Inside Airbnb (CC BY 4.0). Mercato degli affitti brevi per
#   territorio: 7 città + 3 regioni intere (niente Abruzzo). Cache aggregate:
#   territorio (KPI), zona (quartiere), room_type. Occupazione = PROXY. [[tdh-engine-project]]
# ════════════════════════════════════════════════════════════════════════════
_STR_TERR_PATH = ".cache/str_airbnb_territorio.csv"
_STR_ZONA_PATH = ".cache/str_airbnb_zona.csv"
_STR_ROOM_PATH = ".cache/str_airbnb_roomtype.csv"


@st.cache_data(show_spinner=False)
def str_territori() -> pd.DataFrame:
    if not os.path.exists(_STR_TERR_PATH):
        return pd.DataFrame()
    return pd.read_csv(_STR_TERR_PATH)


@st.cache_data(show_spinner=False)
def _str_zone_all() -> pd.DataFrame:
    if not os.path.exists(_STR_ZONA_PATH):
        return pd.DataFrame()
    return pd.read_csv(_STR_ZONA_PATH)


@st.cache_data(show_spinner=False)
def _str_room_all() -> pd.DataFrame:
    if not os.path.exists(_STR_ROOM_PATH):
        return pd.DataFrame()
    return pd.read_csv(_STR_ROOM_PATH)


def str_kpi(slug: str) -> dict:
    """Riga KPI del territorio (dict) o {} se assente."""
    df = str_territori()
    if df.empty:
        return {}
    r = df[df["slug"] == slug]
    return r.iloc[0].to_dict() if not r.empty else {}


def str_zone(slug: str, top: int | None = 15) -> pd.DataFrame:
    df = _str_zone_all()
    if df.empty:
        return df
    d = df[df["slug"] == slug].sort_values("n_annunci", ascending=False)
    return d.head(top) if top else d


def str_roomtypes(slug: str) -> pd.DataFrame:
    df = _str_room_all()
    return df[df["slug"] == slug] if not df.empty else df


def chart_str_adr_territori(sel_slug: str | None = None) -> go.Figure | None:
    """Confronto ADR mediano tra tutti i territori (barre orizzontali), evidenzia il selezionato."""
    df = str_territori()
    if df.empty:
        return None
    d = df.dropna(subset=["adr_mediano"]).sort_values("adr_mediano")
    colors = ["#f59e0b" if s == sel_slug else "#0e6b70" for s in d["slug"]]
    fig = go.Figure(go.Bar(x=d["adr_mediano"], y=d["territorio"], orientation="h", marker_color=colors,
                           customdata=d["n_annunci"],
                           hovertemplate="<b>%{y}</b><br>ADR mediano € %{x:,.0f}<br>"
                                         "%{customdata:,.0f} annunci<extra></extra>"))
    fig.update_xaxes(title="ADR mediano (€/notte)")
    return _layout(fig, h=420)


def chart_str_zone(slug: str, top: int = 12) -> go.Figure | None:
    d = str_zone(slug, top=top)
    if d is None or d.empty:
        return None
    d = d.sort_values("n_annunci")
    fig = go.Figure(go.Bar(x=d["n_annunci"], y=d["zona"], orientation="h", marker_color="#0e6b70",
                           customdata=d["adr_mediano"],
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} annunci<br>"
                                         "ADR mediano € %{customdata:,.0f}<extra></extra>"))
    fig.update_xaxes(title="numero di annunci")
    return _layout(fig, h=440)


def chart_str_roomtype(slug: str) -> go.Figure | None:
    d = str_roomtypes(slug)
    if d is None or d.empty:
        return None
    d = d.sort_values("n_annunci")
    fig = go.Figure(go.Bar(x=d["n_annunci"], y=d["room_type"], orientation="h", marker_color="#16a34a",
                           customdata=d["adr_mediano"],
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} annunci<br>"
                                         "ADR mediano € %{customdata:,.0f}<extra></extra>"))
    fig.update_xaxes(title="numero di annunci")
    return _layout(fig, h=300)


_STR_REV_PATH = ".cache/str_airbnb_recensioni_mese.csv"


@st.cache_data(show_spinner=False)
def str_reviews_monthly() -> pd.DataFrame:
    if not os.path.exists(_STR_REV_PATH):
        return pd.DataFrame()
    df = pd.read_csv(_STR_REV_PATH)
    df["date"] = pd.to_datetime(df["mese"].astype(str) + "-01", errors="coerce")
    return df


def chart_str_struttura() -> go.Figure | None:
    """Confronto tra territori su struttura del mercato: % intero alloggio, % operatori
    professionali (multi-annuncio), % con licenza/CIR (barre orizzontali raggruppate)."""
    df = str_territori()
    if df.empty:
        return None
    d = df.sort_values("pct_licenza")
    fig = go.Figure()
    for col, color, lbl in [("pct_intero", "#0e6b70", "Intero alloggio"),
                            ("pct_multihost", "#f59e0b", "Operatori professionali"),
                            ("pct_licenza", "#16a34a", "Con licenza/CIR")]:
        if col in d.columns:
            fig.add_bar(y=d["territorio"], x=d[col], orientation="h", name=lbl, marker_color=color,
                        hovertemplate="<b>%{y}</b><br>" + lbl + ": %{x:.0f}%<extra></extra>")
    fig.update_layout(barmode="group", legend=dict(orientation="h", y=1.08, x=0))
    fig.update_xaxes(title="% degli annunci", range=[0, 100])
    return _layout(fig, h=520)


def chart_str_reviews(slug: str) -> go.Figure | None:
    """Andamento nel tempo dell'attività: recensioni per mese del territorio (proxy della domanda)."""
    df = str_reviews_monthly()
    if df.empty:
        return None
    d = df[df["slug"] == slug].dropna(subset=["date"]).sort_values("date")
    if d.empty:
        return None
    fig = go.Figure(go.Scatter(x=d["date"], y=d["n_recensioni"], mode="lines",
                               line=dict(color="#0e6b70", width=1.8),
                               fill="tozeroy", fillcolor="rgba(14,116,144,0.08)",
                               hovertemplate="%{x|%m/%Y}<br>%{y:,.0f} recensioni<extra></extra>"))
    fig.update_yaxes(title="recensioni / mese")
    fig.update_xaxes(title=None)
    return _layout(fig, h=340)


# ════════════════════════════════════════════════════════════════════════════
# STR NAZIONALE (ANTEPRIMA) — front-end su schema AirROI/AirDNA con DATI DI ESEMPIO.
#   Metriche di mercato reali (occupazione, ADR, RevPAR, ricavi, annunci, permanenza),
#   copertura nazionale (tutte le regioni). I valori qui sono SINTETICI e realistici:
#   nel prodotto PRIVATO si sostituiscono con le chiamate API (AirROI /markets/metrics/*).
#   NB: non committare dati AirROI/AirDNA reali (licenza); qui è tutto mock. [[tdh-engine-project]]
# ════════════════════════════════════════════════════════════════════════════
# (market, regione, occupancy 0-1, ADR €, active_listings, avg_length_of_stay notti)
_STR_NAT_SAMPLE = [
    ("Roma", "Lazio", 0.45, 162, 37084, 3.4), ("Milano", "Lombardia", 0.48, 136, 25750, 3.0),
    ("Firenze", "Toscana", 0.43, 166, 13472, 3.3), ("Venezia", "Veneto", 0.44, 202, 8766, 3.1),
    ("Napoli", "Campania", 0.36, 107, 11575, 3.5), ("Bologna", "Emilia-Romagna", 0.53, 128, 4948, 2.9),
    ("Torino", "Piemonte", 0.44, 95, 6000, 3.1), ("Palermo", "Sicilia", 0.40, 92, 5200, 3.8),
    ("Catania", "Sicilia", 0.42, 88, 3400, 3.7), ("Bari", "Puglia", 0.43, 105, 2600, 3.6),
    ("Lecce", "Puglia", 0.38, 120, 3100, 4.2), ("Matera", "Basilicata", 0.40, 110, 900, 2.8),
    ("Verona", "Veneto", 0.47, 130, 3000, 2.9), ("Rimini", "Emilia-Romagna", 0.41, 115, 4200, 4.5),
    ("Como", "Lombardia", 0.46, 175, 2800, 3.2), ("Costiera Amalfitana", "Campania", 0.50, 260, 2200, 4.0),
    ("Taormina", "Sicilia", 0.48, 190, 1400, 4.1), ("Cortina d'Ampezzo", "Veneto", 0.38, 240, 900, 4.3),
    ("Bolzano-Dolomiti", "Trentino-Alto Adige", 0.52, 165, 3500, 4.6), ("Cagliari", "Sardegna", 0.42, 130, 2400, 4.0),
    ("Alghero", "Sardegna", 0.40, 140, 1600, 4.4), ("L'Aquila", "Abruzzo", 0.34, 95, 700, 3.5),
    ("Perugia", "Umbria", 0.39, 100, 1500, 3.0),
]
# profilo di stagionalità (mensile, picco estivo) — usato per la vista di dettaglio mock
_STR_NAT_SEASON = [0.60, 0.62, 0.78, 0.95, 1.05, 1.20, 1.35, 1.40, 1.15, 0.90, 0.62, 0.68]


@st.cache_data(show_spinner=False)
def str_nat_sample() -> pd.DataFrame:
    """Dataset di ESEMPIO (sintetico) del mercato STR nazionale, schema AirROI/AirDNA."""
    df = pd.DataFrame(_STR_NAT_SAMPLE, columns=["market", "regione", "occ", "adr", "listings", "alos"])
    df["revpar"] = (df["occ"] * df["adr"]).round(1)
    df["revenue"] = (df["revpar"] * 365).round(0).astype("int64")  # ricavo annuo per annuncio
    df["occ_pct"] = (df["occ"] * 100).round(1)
    return df.sort_values("revenue", ascending=False).reset_index(drop=True)


_STR_NAT_METRICS = {"revenue": ("Ricavo annuo/annuncio", "€ %{x:,.0f}"),
                    "revpar": ("RevPAR", "€ %{x:,.1f}"),
                    "adr": ("ADR", "€ %{x:,.0f}"),
                    "occ_pct": ("Occupazione", "%{x:.0f}%"),
                    "listings": ("Annunci attivi", "%{x:,.0f}")}


def chart_str_nat_ranking(metric: str = "revenue", top: int = 15) -> go.Figure | None:
    df = str_nat_sample()
    if df.empty or metric not in df.columns:
        return None
    d = df.sort_values(metric, ascending=False).head(top).sort_values(metric)
    lbl, tmpl = _STR_NAT_METRICS.get(metric, (metric, "%{x}"))
    fig = go.Figure(go.Bar(x=d[metric], y=d["market"], orientation="h", marker_color="#0e6b70",
                           hovertemplate="<b>%{y}</b><br>" + lbl + ": " + tmpl + "<extra></extra>"))
    fig.update_xaxes(title=lbl)
    return _layout(fig, h=460)


def str_nat_seasonality(market: str) -> pd.DataFrame:
    """Curva mensile di ESEMPIO (occupazione/ADR) per un mercato, dal profilo stagionale."""
    df = str_nat_sample()
    r = df[df["market"] == market]
    if r.empty:
        return pd.DataFrame()
    occ, adr = float(r["occ"].iloc[0]), float(r["adr"].iloc[0])
    mean = sum(_STR_NAT_SEASON) / 12
    rows = []
    for m, f in enumerate(_STR_NAT_SEASON, start=1):
        rows.append({"mese": m, "occ_pct": round(min(occ * f / mean, 0.98) * 100, 1),
                     "adr": round(adr * (0.85 + 0.30 * (f / max(_STR_NAT_SEASON))))})
    return pd.DataFrame(rows)


def chart_str_nat_seasonality(market: str) -> go.Figure | None:
    d = str_nat_seasonality(market)
    if d.empty:
        return None
    mesi = ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"]
    fig = go.Figure()
    fig.add_bar(x=mesi, y=d["occ_pct"], name="Occupazione %", marker_color="#0e6b70",
                hovertemplate="%{x}<br>occ. %{y:.0f}%<extra></extra>")
    fig.add_scatter(x=mesi, y=d["adr"], name="ADR €", yaxis="y2", mode="lines+markers",
                    line=dict(color="#f59e0b", width=2),
                    hovertemplate="%{x}<br>ADR € %{y:,.0f}<extra></extra>")
    fig.update_layout(yaxis=dict(title="occupazione %"),
                      yaxis2=dict(title="ADR €", overlaying="y", side="right", showgrid=False),
                      legend=dict(orientation="h", y=1.12, x=0))
    return _layout(fig, h=340)


# ════════════════════════════════════════════════════════════════════════════
# MERCATI D'ORIGINE — i 10 paesi da cui arrivano i turisti stranieri.
#   spesa in Italia + n. turisti (Banca d'Italia, TS1 per paese, completi al 2025)
#   spesa per turismo all'estero = taglia del mercato (World Bank ST.INT.XPND.CD)
#   quota Italia = spesa in Italia ÷ spesa outbound (stesso anno, cambio EUR/USD stimato)
# ════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def outbound_long() -> pd.DataFrame:
    """Spesa per turismo all'estero UNIFICATA in M€ (Eurostat per i reporter europei +
    World Bank per gli extra-UE): code, paese, anno, spesa_out_eur_m, fonte."""
    try:
        return ECS.fetch_outbound_expenditure()
    except Exception:  # noqa: BLE001
        return pd.DataFrame()


@st.cache_data(show_spinner="Assemblo i mercati d'origine…")
def origin_markets_table() -> list[dict]:
    """Riga per ciascuno dei 10 paesi d'origine: spesa in Italia, turisti, spesa/turista (BdI,
    ultimo anno completo), spesa outbound (World Bank, ultimo anno disp.) e quota Italia.
    Ordinata per spesa in Italia decrescente."""
    long = bdi_country_long()
    if long is None or long.empty:
        return []
    long = long.copy()
    long["anno"] = long["date"].dt.year
    ann = (long.groupby(["code", "anno"])
           .agg(spesa=("spesa", "sum"), notti=("notti", "sum"),
                viaggiatori=("viaggiatori", "sum"), q=("date", "count")).reset_index())
    ann = ann[ann["q"] >= 4]                       # solo anni completi (4 trimestri)
    out = outbound_long()
    rows = []
    for code, (iso3, nome) in ECS.ORIGIN_COUNTRIES.items():
        a = ann[ann["code"] == code].sort_values("anno")
        if a.empty:
            continue
        last = a.iloc[-1]
        spesa_it, turisti, notti, yr = (float(last["spesa"]), float(last["viaggiatori"]),
                                        float(last["notti"]), int(last["anno"]))
        spt = spesa_it * 1e6 / (turisti * 1e3) if turisti else None
        out_eur_mld = out_yr = out_fonte = quota = None
        w = out[out["code"] == code].sort_values("anno") if out is not None and not out.empty else None
        if w is not None and not w.empty:
            last_o = w.iloc[-1]
            out_eur_m, out_yr, out_fonte = (float(last_o["spesa_out_eur_m"]),
                                            int(last_o["anno"]), str(last_o["fonte"]))
            out_eur_mld = out_eur_m / 1000
            same = a[a["anno"] == out_yr]                       # quota su anno comune
            sp_same = float(same["spesa"].iloc[0]) if not same.empty else spesa_it
            quota = sp_same / out_eur_m * 100 if out_eur_m else None
        rows.append({"code": code, "paese": nome, "anno_bdi": yr, "spesa_it_M": spesa_it,
                     "turisti_k": turisti, "notti_k": notti, "spesa_per_turista": spt,
                     "out_eur_mld": out_eur_mld, "out_anno": out_yr, "out_fonte": out_fonte,
                     "quota_pct": quota})
    rows.sort(key=lambda r: -r["spesa_it_M"])
    return rows


def origin_markets_coverage() -> dict:
    """Quanto i 10 paesi coprono del totale nazionale (riconciliazione con la «riga gialla»).
    Ritorna {sum10_M, naz_M, quota_pct, anno}."""
    rows = origin_markets_table()
    if not rows:
        return {}
    sum10 = sum(r["spesa_it_M"] for r in rows)
    naz = bdi_national_annual()
    naz_M = anno = None
    if naz is not None and not naz.empty:
        full = naz[naz["trimestri"] >= 4]
        last = (full if not full.empty else naz).iloc[-1]
        naz_M, anno = float(last["spesa"]), int(last["anno"])
    return {"sum10_M": sum10, "naz_M": naz_M, "anno": anno,
            "quota_pct": (sum10 / naz_M * 100 if naz_M else None)}


# Registro variabili per il TREND storico di un mercato d'origine (chiave · etichetta · unità · decimali)
MARKET_VARS = [
    ("spesa_it_M", "Spesa in Italia", "M€", 0),
    ("turisti_k", "Turisti", "migliaia", 0),
    ("spesa_per_turista", "Spesa per turista", "€", 0),
    ("out_eur_mld", "Spesa all'estero (totale)", "mld€", 1),
    ("quota_pct", "Quota Italia", "%", 1),
]
MARKET_VAR_LABEL = {k: lab for k, lab, _u, _d in MARKET_VARS}
MARKET_VAR_UNIT = {k: u for k, _l, u, _d in MARKET_VARS}
MARKET_VAR_DEC = {k: d for k, _l, _u, d in MARKET_VARS}


@st.cache_data(show_spinner="Carico la serie storica del mercato…")
def origin_market_panel(code: str) -> pd.DataFrame:
    """Serie storica ANNUALE di un mercato d'origine (anno × variabili): spesa in Italia,
    turisti, spesa/turista (BdI, anni completi), spesa all'estero (Eurostat/WB) e quota Italia."""
    long = bdi_country_long()
    if long is None or long.empty:
        return pd.DataFrame()
    long = long.copy()
    long["anno"] = long["date"].dt.year
    a = (long[long["code"] == code].groupby("anno")
         .agg(spesa=("spesa", "sum"), viaggiatori=("viaggiatori", "sum"), q=("date", "count")).reset_index())
    a = a[a["q"] >= 4]
    if a.empty:
        return pd.DataFrame()
    df = pd.DataFrame(index=a["anno"].astype(int))
    df["spesa_it_M"] = a.set_index("anno")["spesa"]
    df["turisti_k"] = a.set_index("anno")["viaggiatori"]
    df["spesa_per_turista"] = df["spesa_it_M"] * 1e6 / (df["turisti_k"] * 1e3)
    out = outbound_long()
    if out is not None and not out.empty:
        o = out[out["code"] == code].set_index("anno")["spesa_out_eur_m"]
        if not o.empty:
            df["out_eur_mld"] = o / 1000
            df["quota_pct"] = df["spesa_it_M"] / o * 100
    df.index.name = "anno"
    return df.sort_index()


def chart_origin_spesa_italia(rows: list[dict]) -> go.Figure:
    """Grafico 1 (area Mercati d'origine): spesa in Italia per paese (M€), barre orizzontali."""
    s = sorted(rows, key=lambda r: r["spesa_it_M"])
    fig = go.Figure(go.Bar(x=[r["spesa_it_M"] for r in s], y=[r["paese"] for r in s],
                           orientation="h", marker_color="#0e6b70",
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} M€<extra></extra>"))
    fig.update_xaxes(title="spesa dei turisti in Italia (M€)")
    return _layout(fig, h=380)


def chart_origin_share(rows: list[dict]) -> go.Figure:
    """Grafico 2 (area Mercati d'origine): taglia del mercato (outbound, mld €) vs quota Italia (%).
    Bolla = spesa in Italia. Solo i paesi con outbound disponibile."""
    pts = [r for r in rows if r["out_eur_mld"] and r["quota_pct"] is not None]
    if not pts:
        return _layout(go.Figure(), h=380)
    fig = go.Figure(go.Scatter(
        x=[r["out_eur_mld"] for r in pts], y=[r["quota_pct"] for r in pts], mode="markers+text",
        text=[r["paese"] for r in pts], textposition="top center",
        marker=dict(size=[max(12, (r["spesa_it_M"] / 200)) for r in pts], color="#f59e0b",
                    line=dict(color="#b45309", width=1)),
        hovertemplate="<b>%{text}</b><br>outbound %{x:,.0f} mld€<br>quota Italia %{y:.1f}%<extra></extra>"))
    fig.update_xaxes(title="taglia del mercato — spesa turismo all'estero (mld €)")
    fig.update_yaxes(title="quota catturata dall'Italia (%)", ticksuffix="%")
    return _layout(fig, h=400)


def chart_bdi_country(metric: str = "notti") -> go.Figure | None:
    g = bdi_country_annual()
    if g is None or g.empty:
        return None
    names = {mk.code: mk.name for mk in DEFAULT_MARKETS}
    fig = go.Figure()
    for code in BDI_MARKETS:
        d = g[g["code"] == code].sort_values("anno")
        if not d.empty:
            fig.add_trace(go.Scatter(x=d["anno"], y=d[metric], name=names.get(code, code),
                                     mode="lines+markers"))
    fig.update_yaxes(title=metric)
    return _layout(fig, h=380)


def chart_value_bar(summary: list[dict]) -> go.Figure:
    s = sorted([x for x in summary if x.get("valore")], key=lambda x: x["valore"])
    fig = go.Figure(go.Bar(x=[x["valore"] for x in s], y=[x["market"] for x in s], orientation="h",
                           marker_color="#0e6b70",
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} €/viaggiatore<extra></extra>"))
    fig.update_layout(xaxis_title="spesa €/viaggiatore")
    return _layout(fig, h=300)


def chart_region_weight_bar(summary: list[dict]) -> go.Figure | None:
    """Peso economico REALE del mercato nella regione (stima M€) — il peso usato nello score."""
    s = sorted([x for x in summary if x.get("peso")], key=lambda x: x["peso"])
    if not s:
        return None
    fig = go.Figure(go.Bar(x=[x["peso"] for x in s], y=[x["market"] for x in s], orientation="h",
                           marker_color="#0e6b70",
                           hovertemplate="<b>%{y}</b><br>%{x:,.1f} M€ (stima regionale)<extra></extra>"))
    fig.update_layout(xaxis_title="valore economico stimato nella regione (M€/anno)")
    return _layout(fig, h=300)


# --- Eurostat avia_par: connettività aerea diretta su Pescara (fattibilità reale) ---
@st.cache_data(show_spinner=False)
def pescara_connectivity() -> dict:
    p = ".cache/eurostat_pescara_flights.csv"
    if not os.path.exists(p):
        return {}
    df = pd.read_csv(p)
    return {str(r["code"]): int(r["pax"]) for _, r in df.iterrows()}


def feas_weights() -> dict:
    """Peso di fattibilità 0.4..1.0 dalla connettività voli (0 voli → 0.4; più connesso → 1.0)."""
    pax = pescara_connectivity()
    if not pax:
        return {}
    mx = max(pax.values()) or 1
    return {k: round(0.4 + 0.6 * min(v / mx, 1.0), 3) for k, v in pax.items()}


@st.cache_data(show_spinner=False)
def region_connectivity(code: str | None = None) -> dict:
    """Connettività aerea per mercato verso gli aeroporti della regione (Eurostat avia)."""
    code = code or RG.DEFAULT_REGION
    airports = RG.region(code)["airports"]
    if not airports:
        return {}
    try:
        return ECS.fetch_region_connectivity(airports, cache_name=f"conn_{code}.csv")
    except Exception:  # noqa: BLE001
        return {}


def chart_connectivity(code: str | None = None):
    code = code or RG.DEFAULT_REGION
    pax = region_connectivity(code)
    if not pax:
        return None
    names = {mk.code: mk.name for mk in DEFAULT_MARKETS}
    items = sorted(((mk.code, pax.get(mk.code, 0)) for mk in DEFAULT_MARKETS), key=lambda x: x[1])
    fig = go.Figure(go.Bar(x=[v for _, v in items], y=[names.get(k, k) for k, _ in items], orientation="h",
                           marker_color=["#16a34a" if v > 0 else "#cbd5e1" for _, v in items],
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} passeggeri 2024<extra></extra>"))
    fig.update_xaxes(title=f"passeggeri diretti verso {RG.region(code)['nome']} (2024)")
    return _layout(fig, h=260)


# --- Salute mercato (fiducia consumatori) e accessibilità nel tempo (voli Pescara) ---
# Criteri DECISIONALI di contesto (non predittori del forecast: vedi bake-off).
HEALTH_GEO = {"DE": "DE", "AT": "AT", "NL": "NL"}   # fiducia disponibile (Eurostat)


@st.cache_data(show_spinner=False)
def market_health() -> dict:
    """Trend della fiducia dei consumatori per mercato: {code: {conf, delta, label, series}}.
    Disponibile per DE/AT/NL; per gli altri mercati non c'è dato Eurostat."""
    out = {}
    for code, geo in HEALTH_GEO.items():
        p = f".cache/econ_confidence_{geo}.csv"
        if not os.path.exists(p):
            continue
        df = pd.read_csv(p, parse_dates=["date"]).sort_values("date")
        if df.empty:
            continue
        s = df["confidence"].astype(float)
        latest = float(s.iloc[-1])
        prev = float(s.iloc[-7]) if len(s) >= 7 else float(s.iloc[0])   # ~6 mesi prima
        delta = latest - prev
        label = "📈 in ripresa" if delta >= 3 else "📉 in calo" if delta <= -3 else "➡️ stabile"
        out[code] = {"conf": round(latest, 1), "delta": round(delta, 1), "label": label,
                     "series": df[["date", "confidence"]]}
    return out


@st.cache_data(show_spinner=False)
def flights_monthly_df():
    p = ".cache/econ_flights_pescara_monthly.csv"
    return pd.read_csv(p, parse_dates=["date"]).sort_values("date") if os.path.exists(p) else None


def chart_flights_monthly():
    df = flights_monthly_df()
    if df is None or df.empty:
        return None
    df = df.copy(); df["roll12"] = df["pax"].rolling(12).mean()
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df["date"], y=df["pax"], name="passeggeri/mese",
                             line=dict(color="#cbd5e1", width=1)))
    fig.add_trace(go.Scatter(x=df["date"], y=df["roll12"], name="media 12 mesi",
                             line=dict(color="#0e6b70", width=3)))
    fig.update_yaxes(title="passeggeri Pescara")
    return _layout(fig, h=280)


def chart_confidence(health: dict):
    """Mini-serie della fiducia per i mercati con dato."""
    if not health:
        return None
    fig = go.Figure()
    for code, h in health.items():
        s = h["series"]
        fig.add_trace(go.Scatter(x=s["date"], y=s["confidence"], name=code, mode="lines"))
    fig.update_yaxes(title="fiducia (saldo)")
    return _layout(fig, h=260)


# --- Wikipedia pageviews: secondo segnale anticipatore (per LINGUA, corroborazione) ---
WIKI_LANG = {"DE": "de", "AT": "de", "CH": "de", "GB": "en", "US": "en", "NL": "nl"}
WIKI_LABEL = {"de": "Tedesco (DE/AT/CH)", "en": "Inglese (GB/US)", "nl": "Olandese (NL)",
              "it": "Italiano (domestico)"}


@st.cache_data(show_spinner=False)
def wiki_series(lang: str, code: str | None = None):
    code = code or RG.DEFAULT_REGION
    art = RG.region(code)["wiki"].get(lang)
    if not art:
        return None
    cache_name = f"wiki_{lang}.csv" if code == RG.DEFAULT_REGION else f"wiki_{code}_{lang}.csv"
    p = f".cache/{cache_name}"
    if not os.path.exists(p):  # regione nuova: scarico l'articolo (Wikipedia è affidabile)
        try:
            RS.fetch_wikipedia_monthly(lang, article=art, cache_name=cache_name)
        except Exception:  # noqa: BLE001
            return None
    if not os.path.exists(p):
        return None
    df = pd.read_csv(p, parse_dates=["date"])
    today = pd.Timestamp.today()
    cutoff = pd.Timestamp(today.year, today.month, 1)  # scarta il mese corrente (incompleto)
    return df[df["date"] < cutoff].reset_index(drop=True)


def wiki_momentum(lang: str, code: str | None = None):
    df = wiki_series(lang, code)
    if df is None or df.empty:
        return None
    s = df.sort_values("date")["views"]
    recent = s.tail(3).mean()
    yoy = s.tail(15).head(3).mean() if len(s) >= 15 else None
    return round((recent - yoy) / yoy * 100, 1) if (yoy and yoy != 0) else None


def online_interest(summary: list[dict], code: str | None = None) -> list[dict]:
    out = []
    for m in summary:
        lang = WIKI_LANG.get(m["code"])
        wmom = wiki_momentum(lang, code)
        tmom = m.get("momentum")
        conc = "—"
        if tmom is not None and wmom is not None:
            conc = "✓ concorde" if (tmom >= 0) == (wmom >= 0) else "✗ discorde"
        out.append({"Mercato": m["market"], "Wikipedia": WIKI_LABEL.get(lang, lang or "—"),
                    "Trends mom %": (round(tmom) if tmom is not None else None),
                    "Wikipedia mom %": wmom, "Concordanza": conc})
    return out


def chart_wiki(yr_range=None, code: str | None = None) -> go.Figure:
    colors = {"de": "#0e6b70", "en": "#f59e0b", "nl": "#16a34a"}
    fig = go.Figure()
    for lang in ["de", "en", "nl"]:
        df = wiki_series(lang, code)
        if df is None:
            continue
        if yr_range:
            df = df[(df["date"].dt.year >= yr_range[0]) & (df["date"].dt.year <= yr_range[1])]
        fig.add_trace(go.Scatter(x=df["date"], y=df["views"], name=WIKI_LABEL[lang],
                                 line=dict(color=colors[lang], width=2)))
    fig.update_yaxes(title="visualizzazioni pagina (per lingua)")
    return _layout(fig, h=360)


def get_context(region_code: str | None = None) -> dict:
    """Ritorna il contesto della modalità corrente per la REGIONE attiva (dati + ctx per i tool)."""
    region_code = region_code or st.session_state.get("region_code", RG.DEFAULT_REGION)
    # In modalità «Italia» (vista d'insieme) il motore per-regione non ha un target nazionale
    # diretto (arriverà nello Step B): usiamo l'Abruzzo come stand-in per non far crashare il
    # caricamento globale. Le pagine motore restano comunque "guardate" lato app.
    eng_code = RG.DEFAULT_REGION if RG.is_national(region_code) else region_code
    mode = st.session_state.get("mode", MODE_REAL)
    if MODE_REAL in mode or "Reale" in mode:
        R = compute_real(eng_code)
        return {"mode": "real", "R": R, "region": region_code}
    ranked, rows = compute_synthetic()
    return {"mode": "synthetic", "ranked": ranked, "rows": rows, "region": region_code}


def markets_summary(ctx: dict) -> list[dict]:
    """Lista normalizzata per grafici/mappa: code, market, reco, score, forza, momentum, valore, capacity_ok, rank."""
    if ctx["mode"] == "real":
        return [{"code": c["code"], "market": c["market"], "reco": c["raccomandazione"],
                 "score": c["score"], "forza": c["forza_anticipatrice"],
                 "momentum": c["momentum_search_pct"], "valore": c["valore_eur_per_visitatore"],
                 "peso": c.get("peso_score"), "capacity_ok": c["capacita_voli_ok"],
                 "rank": c["rank"]} for c in ctx["R"]["ranked"]]
    rows = ctx["rows"]; name2code = {r["name"]: c for c, r in rows.items()}
    out = []
    for r in ctx["ranked"]:
        code = name2code.get(r["market"]); rr = rows.get(code, {})
        out.append({"code": code, "market": r["market"], "reco": r["raccomandazione"],
                    "score": r["opportunity_score"], "forza": None, "momentum": None,
                    "valore": rr.get("spend_per_visitor"), "capacity_ok": rr.get("capacity_ok"),
                    "rank": r["rank"]})
    return out


# ════════════════════════════════════════════════════════════════════════════
# GRAFICI PLOTLY
# ════════════════════════════════════════════════════════════════════════════
def _layout(fig: go.Figure, h: int = 420, title: str | None = None) -> go.Figure:
    fig.update_layout(
        height=h,
        margin=dict(l=10, r=10, t=44 if title else 16, b=10),
        paper_bgcolor="rgba(255,255,255,0.0)",   # trasparente: si integra col contenitore
        plot_bgcolor="#ffffff",                   # area grafico bianca
        title=(title or ""),
        font=dict(family=FONT_STACK, size=13, color="#5c7176"),
        colorway=["#0e6b70", "#e08a1e", "#16a34a", "#7c3aed", "#b91c1c", "#5c7176"],
        legend=dict(orientation="h", y=-0.18, font=dict(size=12),
                    bgcolor="rgba(255,255,255,0)", bordercolor="rgba(0,0,0,0)"),
        hoverlabel=dict(bgcolor="white", bordercolor="#e4ebec",
                        font_size=12, font_family=FONT_STACK))
    fig.update_xaxes(showgrid=False, color="#5c7176", linecolor="#e4ebec")
    fig.update_yaxes(gridcolor="#eef3f3", color="#5c7176", zeroline=False)
    # Barre ORIZZONTALI: le linee guida utili sono VERTICALI (lette sull'asse x
    # degli importi), non orizzontali. Inverto la griglia così che ogni barra
    # corrisponda visivamente al suo valore sull'ascissa.
    is_hbar = any(getattr(tr, "type", None) == "bar" and getattr(tr, "orientation", None) == "h"
                  for tr in fig.data)
    if is_hbar:
        fig.update_xaxes(showgrid=True, gridcolor="#eef3f3")
        fig.update_yaxes(showgrid=False)
    return fig


def chart_map(summary: list[dict]) -> go.Figure:
    gj = json.load(open(".cache/world_countries.geojson", encoding="utf-8"))
    all_ids = [f.get("id") for f in gj["features"] if f.get("id")]
    max_s = max((s["score"] for s in summary), default=1) or 1
    fig = go.Figure()
    fig.add_trace(go.Choropleth(geojson=gj, featureidkey="id", locations=all_ids,
                                z=[0] * len(all_ids), colorscale=[[0, "#e6e8eb"], [1, "#e6e8eb"]],
                                showscale=False, marker_line_color="#fff", marker_line_width=0.4,
                                hoverinfo="skip"))
    lon_l, lat_l = [], []
    for s in summary:
        if s["code"] not in CENTROID:
            continue
        la, lo = CENTROID[s["code"]]; lon_l += [lo, ABRUZZO[1], None]; lat_l += [la, ABRUZZO[0], None]
    fig.add_trace(go.Scattergeo(lon=lon_l, lat=lat_l, mode="lines",
                                line=dict(width=0.8, color="rgba(120,120,120,0.4)"),
                                hoverinfo="skip", showlegend=False))
    by_cat: dict = {}
    for s in summary:
        by_cat.setdefault(reco_cat(s["reco"]), []).append(s)
    for cat, grp in by_cat.items():
        grp = [g for g in grp if g["code"] in CENTROID]
        if not grp:
            continue
        fig.add_trace(go.Scattergeo(
            lon=[CENTROID[g["code"]][1] for g in grp], lat=[CENTROID[g["code"]][0] for g in grp],
            mode="markers+text", text=[g["market"] for g in grp], textposition="top center",
            textfont=dict(size=11, color="#111827"),
            marker=dict(size=[14 + 44 * math.sqrt(max(g["score"], 1) / max_s) for g in grp],
                        color=reco_color(grp[0]["reco"]), opacity=0.85, line=dict(width=1.2, color="white")),
            name=cat,
            customdata=[[g["rank"], g["reco"], round(g["score"])] for g in grp],
            hovertemplate="<b>%{text}</b><br>#%{customdata[0]} · %{customdata[1]}"
                          "<br>Score: %{customdata[2]:,}<extra></extra>"))
    fig.add_trace(go.Scattergeo(lon=[ABRUZZO[1]], lat=[ABRUZZO[0]], mode="markers+text",
                                text=["ABRUZZO"], textposition="bottom center",
                                textfont=dict(size=12, color="#7c2d12"),
                                marker=dict(size=16, color="#7c2d12", symbol="star"),
                                name="Destinazione", hoverinfo="text"))
    fig.update_layout(height=560, margin=dict(l=0, r=0, t=10, b=0), paper_bgcolor="white", title_text="",
                      legend=dict(title="Raccomandazione", orientation="h", y=-0.02, x=0.02),
                      geo=dict(projection_type="natural earth", showland=False, showcountries=False,
                               showcoastlines=False, showframe=False, showocean=True, oceancolor="#eaf2fb",
                               bgcolor="rgba(0,0,0,0)", lataxis_range=[22, 66], lonaxis_range=[-120, 32]))
    return fig


def chart_ranking_bar(summary: list[dict]) -> go.Figure:
    s = sorted(summary, key=lambda x: x["score"])  # crescente -> #1 in alto
    fig = go.Figure(go.Bar(
        x=[x["score"] for x in s], y=[x["market"] for x in s], orientation="h",
        marker_color=[reco_color(x["reco"]) for x in s],
        customdata=[[x["reco"], x["rank"]] for x in s],
        hovertemplate="<b>%{y}</b><br>#%{customdata[1]} · %{customdata[0]}<br>Score: %{x:,.0f}<extra></extra>"))
    fig.update_layout(xaxis_title="Score (opportunità)")
    return _layout(fig, h=360)


def chart_forecast(history: pd.DataFrame, fc: dict, ylabel: str = "presenze") -> go.Figure:
    hist = history.tail(24)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=fc["dates"], y=fc["hi"], line=dict(width=0), showlegend=False, hoverinfo="skip"))
    fig.add_trace(go.Scatter(x=fc["dates"], y=fc["lo"], fill="tonexty", fillcolor="rgba(245,158,11,0.18)",
                             line=dict(width=0), name="intervallo 80%", hoverinfo="skip"))
    fig.add_trace(go.Scatter(x=hist["date"], y=hist[hist.columns[1]], name="storico",
                             line=dict(color="#2563eb", width=2)))
    fig.add_trace(go.Scatter(x=fc["dates"], y=fc["mean"], name="forecast",
                             line=dict(color="#f59e0b", width=2), mode="lines+markers"))
    ref_x = [d for d, v in zip(fc["dates"], fc["lastyear"]) if v is not None]
    ref_y = [v for v in fc["lastyear"] if v is not None]
    if ref_y:
        fig.add_trace(go.Scatter(x=ref_x, y=ref_y, name="anno prima", mode="markers",
                                 marker=dict(color="#16a34a", size=9)))
    fig.update_yaxes(title=ylabel)
    return _layout(fig, h=380)


def chart_search(panel: pd.DataFrame, code: str) -> go.Figure:
    col = f"search_{code}"
    d = panel[["date", col]].dropna()
    fig = go.Figure(go.Scatter(x=d["date"], y=d[col], line=dict(color="#7c3aed", width=2),
                               name="interesse di ricerca"))
    fig.update_yaxes(title="interesse di ricerca (0–100)")
    return _layout(fig, h=300)


def chart_seasonal_heatmap(panel: pd.DataFrame, summary: list[dict]) -> go.Figure:
    """Heatmap mese×mercato del profilo stagionale di ricerca (normalizzato per riga):
    mostra QUANDO l'interesse di ogni mercato è al massimo (timing campagne)."""
    order = [s["code"] for s in summary if f"search_{s['code']}" in panel.columns]
    names = {s["code"]: s["market"] for s in summary}
    d = panel.copy(); d["m"] = d["date"].dt.month
    z, ytxt = [], []
    for code in order:
        prof = d.groupby("m")[f"search_{code}"].mean().reindex(range(1, 13))
        mn, mx = prof.min(), prof.max()
        norm = (prof - mn) / (mx - mn) if mx > mn else prof * 0
        z.append(norm.tolist()); ytxt.append(names[code])
    fig = go.Figure(go.Heatmap(z=z, x=MONTHS_IT, y=ytxt, colorscale="YlOrRd",
                               colorbar=dict(title="interesse<br>(per mercato)"),
                               hovertemplate="%{y} · %{x}<br>intensità %{z:.0%}<extra></extra>"))
    return _layout(fig, h=320)


# ════════════════════════════════════════════════════════════════════════════
# ALLOCATORE BUDGET
# ════════════════════════════════════════════════════════════════════════════
def allocate_budget(summary: list[dict], total: float, categorie: set[str]) -> list[dict]:
    """Ripartisce il budget proporzionalmente allo score, tra i mercati la cui categoria
    di raccomandazione è in `categorie`. I mercati esclusi/score<=0 ricevono 0."""
    elig = [s for s in summary if reco_cat(s["reco"]) in categorie and s["score"] > 0]
    tot_w = sum(s["score"] for s in elig) or 1
    out = []
    for s in summary:
        w = s["score"] if s in elig else 0
        quota = total * w / tot_w if w > 0 else 0.0
        out.append({**s, "quota_eur": quota, "quota_pct": (w / tot_w * 100 if w > 0 else 0.0)})
    return sorted(out, key=lambda x: x["quota_eur"], reverse=True)


def chart_allocation(alloc: list[dict]) -> go.Figure:
    a = [x for x in alloc if x["quota_eur"] > 0]
    fig = go.Figure(go.Bar(x=[x["quota_eur"] for x in a][::-1], y=[x["market"] for x in a][::-1],
                           orientation="h", marker_color=[reco_color(x["reco"]) for x in a][::-1],
                           customdata=[[x["quota_pct"]] for x in a][::-1],
                           hovertemplate="<b>%{y}</b><br>€ %{x:,.0f} (%{customdata[0]:.0f}%)<extra></extra>"))
    fig.update_layout(xaxis_title="budget allocato (€)")
    return _layout(fig, h=340)


# ════════════════════════════════════════════════════════════════════════════
# KPI (modalità reale)
# ════════════════════════════════════════════════════════════════════════════
def kpi_real(R: dict) -> dict:
    h = R["history"].sort_values("date")
    last12 = h.tail(12)["presences"].sum()
    prev12 = h.tail(24).head(12)["presences"].sum()
    yoy = (last12 - prev12) / prev12 * 100 if prev12 else float("nan")
    top = R["ranked"][0]["market"] if R["ranked"] else "—"
    anno = h["date"].max().year if not h.empty else "—"
    return {"presenze_anno": last12, "anno": anno, "yoy": yoy, "top": top,
            "n_aumentare": sum(1 for c in R["ranked"] if c["raccomandazione"].startswith("Aumentare"))}


# ════════════════════════════════════════════════════════════════════════════
# ASSISTENTE (chat con tool, mode-aware)
# ════════════════════════════════════════════════════════════════════════════
def _name_match(items, key, getname, getcode):
    key = (key or "").strip().lower()
    return next((x for x in items if key in (str(getcode(x)).lower(), getname(x).lower())), None)


def run_tool(name, args, ctx) -> str:
    if ctx["mode"] == "real":
        R = ctx["R"]
        if name == "run_engine":
            agg = R["agg"]
            return json.dumps({"ranking": [
                {"rank": c["rank"], "mercato": c["market"], "raccomandazione": c["raccomandazione"],
                 "forza_anticipatrice": c["forza_anticipatrice"], "momentum_search_pct": c["momentum_search_pct"],
                 "score": c["score"]} for c in R["ranked"]],
                "forecast_aggregato": {"lag_mesi": agg["lag"], "batte_naive": agg["beats_naive"],
                "mape_pct": round(agg["mape_model"], 1),
                "nota": "A livello aggregato la stagionalità domina: il modello NON batte la naive; "
                        "per il numero ci si appoggia al riferimento stagionale. Il valore è nel ranking."}},
                ensure_ascii=False)
        if name == "get_market":
            c = _name_match(R["ranked"], args.get("market", ""), lambda x: x["market"], lambda x: x["code"])
            if not c:
                return json.dumps({"errore": "mercato non trovato",
                                   "validi": [x["market"] for x in R["ranked"]]}, ensure_ascii=False)
            return json.dumps({"mercato": c["market"], "rank": c["rank"], "raccomandazione": c["raccomandazione"],
                "forza_anticipatrice": c["forza_anticipatrice"], "lag_mesi": c["lag_mesi"],
                "momentum_search_pct": c["momentum_search_pct"], "search_recente": c["search_recente"],
                "valore_eur_per_visitatore": c["valore_eur_per_visitatore"],
                "capacita_voli_ok": c["capacita_voli_ok"], "score": c["score"]}, ensure_ascii=False)
        if name == "explain_coefficients":
            return json.dumps({"modello": "presenze straniere totali Abruzzo ~ basket search (lag) + "
                "stagionalità + trend + dummy COVID", "lag_mesi": R["agg"]["lag"],
                "batte_naive": R["agg"]["beats_naive"], "mape_pct": round(R["agg"]["mape_model"], 1),
                "lettura_coefficienti": R["agg"]["coeff_reading"]}, ensure_ascii=False)
    else:
        rows, ranked = ctx["rows"], ctx["ranked"]
        if name == "run_engine":
            return json.dumps({"ranking": [
                {"rank": r["rank"], "mercato": r["market"], "raccomandazione": r["raccomandazione"],
                 "opportunity_score": round(r["opportunity_score"]), "confidenza": r["confidenza"]}
                for r in ranked]}, ensure_ascii=False)
        if name in ("get_market", "explain_coefficients"):
            code = _name_match(list(rows.values()), args.get("market", ""),
                               lambda r: r["name"], lambda r: r["code"])
            if not code:
                return json.dumps({"errore": "mercato non trovato",
                                   "validi": [r["name"] for r in rows.values()]}, ensure_ascii=False)
            if name == "get_market":
                return json.dumps({"mercato": code["name"], "scheda": code["card"]}, ensure_ascii=False)
            return json.dumps({"mercato": code["name"], "lag_mesi": code["lag"],
                "batte_naive": code["beats_naive"], "lettura_coefficienti": code["coeff_reading"]},
                ensure_ascii=False)
    return json.dumps({"errore": f"tool sconosciuto: {name}"}, ensure_ascii=False)


TOOLS_SCHEMA = [
    {"name": "run_engine", "description": "Ranking dei mercati + diagnostica del forecast. "
     "Per confronti tra mercati o ordinamento.",
     "input_schema": {"type": "object", "properties": {}, "required": []}},
    {"name": "get_market", "description": "Scheda di un singolo mercato (raccomandazione e numeri chiave).",
     "input_schema": {"type": "object", "properties": {"market": {"type": "string",
                      "description": "Nome o codice, es. 'Germania' o 'DE'."}}, "required": ["market"]}},
    {"name": "explain_coefficients", "description": "Lettura dei coefficienti del modello + backtest. "
     "Per il 'perché' o spiegare il modello.",
     "input_schema": {"type": "object", "properties": {"market": {"type": "string",
                      "description": "Opzionale: nome o codice del mercato."}}, "required": []}},
]

SYSTEM_BASE = ("Sei l'assistente del cruscotto del TDH Engine, che ordina i mercati esteri per allocare "
    "il budget promo turistico regionale. Confine SEMPRE: il motore NON stima l'effetto "
    "causale della spesa ('metti X€ → +Y presenze'); ordina 'dove conviene agire'. Usa SEMPRE i tool "
    "per i numeri (non a memoria); rispondi in italiano, conciso e orientato alla decisione; risposta "
    "finale diretta senza ragionamento visibile.")
SYSTEM_REAL = (" MODALITÀ REALE: un solo modello sulle presenze straniere totali (ISTAT, non per paese); a "
    "livello aggregato la stagionalità domina e NON batte la naive (per il numero si usa il riferimento "
    "stagionale). Il valore è nel RANKING per-mercato = forza anticipatrice (Google Trends) × momentum × "
    "valore economico (spesa €/viaggiatore Banca d'Italia) × fattibilità (voli diretti su Pescara, Eurostat). "
    "Wikipedia (per lingua) corrobora il segnale ma non entra nello score. Sii onesto su queste distinzioni.")
SYSTEM_SYN = (" MODALITÀ SINTETICA (dati di collaudo): modello per-mercato con barriera di onestà (forecast "
    "solo se batte la naive). Puoi dire che i dati sono sintetici.")


def system_prompt_for(ctx: dict) -> str:
    return SYSTEM_BASE + (SYSTEM_REAL if ctx["mode"] == "real" else SYSTEM_SYN)


def render_assistant(ctx: dict, height: int = 460):
    """Componente chat con Claude (streaming + tool sul motore)."""
    if "messages" not in st.session_state:
        st.session_state.messages = []
    env_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if "api_key" not in st.session_state:
        st.session_state.api_key = env_key
    if not st.session_state.api_key:
        st.session_state.api_key = st.text_input("ANTHROPIC_API_KEY", type="password",
            help="A consumo, separata da Claude Code. Oppure imposta la variabile d'ambiente.")

    chat_box = st.container(height=height)
    with chat_box:
        for m in st.session_state.messages:
            if m["role"] == "user" and isinstance(m["content"], str):
                with st.chat_message("user"):
                    st.markdown(m["content"])
            elif m["role"] == "assistant":
                with st.chat_message("assistant"):
                    for b in m["content"]:
                        bt = getattr(b, "type", None) or (isinstance(b, dict) and b.get("type"))
                        if bt == "text":
                            st.markdown(b.text if hasattr(b, "text") else b["text"])
                        elif bt == "tool_use":
                            nm = b.name if hasattr(b, "name") else b["name"]
                            inp = b.input if hasattr(b, "input") else b["input"]
                            st.caption(f"🔧 {nm}({', '.join(f'{k}={v}' for k, v in inp.items())})")

    prompt = st.chat_input("Chiedi all'assistente…")
    if not prompt:
        return
    if not st.session_state.api_key:
        st.warning("Inserisci una API key Anthropic per usare l'assistente."); st.stop()

    import anthropic
    client = anthropic.Anthropic(api_key=st.session_state.api_key)
    st.session_state.messages.append({"role": "user", "content": prompt})
    with chat_box:
        with st.chat_message("user"):
            st.markdown(prompt)
        with st.chat_message("assistant"):
            try:
                for _ in range(6):
                    ph = st.empty(); acc = ""
                    with client.messages.stream(model=MODEL, max_tokens=2000,
                            system=system_prompt_for(ctx), thinking={"type": "disabled"},
                            tools=TOOLS_SCHEMA, messages=st.session_state.messages) as stream:
                        for delta in stream.text_stream:
                            acc += delta; ph.markdown(acc + " ▌")
                        resp = stream.get_final_message()
                    ph.markdown(acc) if acc else ph.empty()
                    st.session_state.messages.append({"role": "assistant", "content": resp.content})
                    calls = [b for b in resp.content if b.type == "tool_use"]
                    for b in calls:
                        st.caption(f"🔧 {b.name}({', '.join(f'{k}={v}' for k, v in b.input.items())})")
                    if resp.stop_reason != "tool_use":
                        break
                    res = [{"type": "tool_result", "tool_use_id": b.id,
                            "content": run_tool(b.name, b.input, ctx)} for b in calls]
                    st.session_state.messages.append({"role": "user", "content": res})
            except anthropic.AuthenticationError:
                st.error("API key non valida.")
            except anthropic.APIError as e:
                st.error(f"Errore API: {e}")


# ════════════════════════════════════════════════════════════════════════════
# GESTIONE DATI — fonti in uso, upload di file con descrizione/fonte, registro
# ════════════════════════════════════════════════════════════════════════════
DATA_DIR = "data"
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
REGISTRY_PATH = os.path.join(DATA_DIR, "registry.json")


def _fmt_mtime(path: str) -> str:
    try:
        return datetime.datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M")
    except OSError:
        return "—"


def _csv_rows(path: str):
    try:
        with open(path, encoding="utf-8", errors="ignore") as f:
            return max(sum(1 for _ in f) - 1, 0)
    except OSError:
        return None


# Schema colonne UNIFICATO per le due tabelle (Presenti e in Valutazione)
SRC_COLS = ["Dataset", "Descrizione", "Fonte", "URL", "Frequenza", "Stato", "Righe", "Aggiornato"]


def _row(dataset, descr, fonte, url, freq, stato, righe, agg):
    return {"Dataset": dataset, "Descrizione": descr, "Fonte": fonte, "URL": url,
            "Frequenza": freq, "Stato": stato, "Righe": righe, "Aggiornato": agg}


def fmt_count_col(df: pd.DataFrame, col: str = "Righe") -> pd.DataFrame:
    """Uniforma a stringa una colonna-conteggio che mescola interi e placeholder.

    Le tabelle delle fonti hanno conteggi righe interi per i CSV in cache e un
    «—» per le fonti live (ECB, ecc.): il mix int/str rompe la serializzazione
    Arrow di ``st.dataframe`` (la colonna diventa ``object`` e va in errore).
    Qui la rendiamo coerente: «12.345» per i numeri, «—» per i mancanti.
    """
    if col not in df.columns:
        return df
    out = df.copy()

    def _f(v):
        if isinstance(v, bool):
            return str(v)
        if isinstance(v, int):
            return f"{v:,}".replace(",", ".")
        if isinstance(v, float):
            return "—" if pd.isna(v) else f"{int(v):,}".replace(",", ".")
        s = str(v).strip()
        if s.isdigit():
            return f"{int(s):,}".replace(",", ".")
        return s if s else "—"

    out[col] = out[col].map(_f)
    return out


def _safe_kw(kw: str) -> str:
    """Slug del nome regione per i file di cache Trends (allineato a _trends_precache.py)."""
    return "".join(c if c.isalnum() else "_" for c in kw.lower())


def builtin_sources() -> list[dict]:
    """Fonti dati GIÀ in uso dal motore (cache + live), colonne unificate con URL."""
    A = "🟢 Attiva"
    rows = []
    istat = ".cache/istat_presenze_straniere_abruzzo.csv"
    if os.path.exists(istat):
        rows.append(_row("Presenze straniere per regione",
                         "Presenze di stranieri al mese (movimento clienti) — tutte le 21 regioni",
                         "ISTAT", "https://esploradati.istat.it/", "mensile", A, _csv_rows(istat), _fmt_mtime(istat)))
    tur9 = ".cache/istat_estero_regione_prov_annuale.csv"
    if os.path.exists(tur9):
        rows.append(_row("Turisti esteri per paese × regione/provincia",
                         "Arrivi/presenze per singolo paese estero, per ogni regione e provincia (annuale, dal 2008)",
                         "ISTAT · Movimento clienti (cube DCSC_TUR_9)", "https://esploradati.istat.it/",
                         "annuale", A, _csv_rows(tur9), _fmt_mtime(tur9)))
    astat_f = ".cache/astat_bolzano_flussi_mensili.csv"
    astat_c = ".cache/astat_bolzano_capacita_categoria.csv"
    if os.path.exists(astat_f):
        rows.append(_row("Alto Adige · flussi e capacità (ASTAT)",
                         "Arrivi/presenze mensili + esercizi/posti letto per categoria — SOLO provincia di "
                         "Bolzano (fonte COMPLEMENTARE, non multi-regione, senza paese di origine)",
                         "ASTAT · Provincia autonoma di Bolzano", "https://astat.provincia.bz.it/",
                         "mensile", A, (_csv_rows(astat_f) or 0) + (_csv_rows(astat_c) or 0), _fmt_mtime(astat_f)))
    lomb = ".cache/lombardia_flussi_provincia_mese.csv"
    if os.path.exists(lomb):
        rows.append(_row("Lombardia · flussi provincia × mercato (Open Data)",
                         "Arrivi/presenze mensili per provincia lombarda e paese estero di origine "
                         "(fonte COMPLEMENTARE, 12 province, 2019-2024)",
                         "Regione Lombardia · Open Data (Socrata)", "https://www.dati.lombardia.it/",
                         "mensile", A, _csv_rows(lomb), _fmt_mtime(lomb)))
    tosc = ".cache/toscana_movimento_comune_anno.csv"
    if os.path.exists(tosc):
        rows.append(_row("Toscana · movimento per comune (Open Data)",
                         "Arrivi/presenze annuali per singolo comune toscano, split italiani/stranieri "
                         "(fonte COMPLEMENTARE, ~272 comuni, 10 province, 2018-2025)",
                         "Regione Toscana · Open Data (CKAN)", "https://dati.toscana.it/",
                         "annuale", A, _csv_rows(tosc), _fmt_mtime(tosc)))
    sard = ".cache/sardegna_flussi_comune_mese.csv"
    if os.path.exists(sard):
        rows.append(_row("Sardegna · movimento mensile per comune × mercato (Osservatorio)",
                         "Arrivi/presenze MENSILI per comune sardo, mercato estero di origine e "
                         "macro-tipologia (fonte COMPLEMENTARE, ~330 comuni, 2021-2025; export manuale CC-BY)",
                         "Osservatorio Turismo Sardegna (SIRED)", "https://osservatorio.sardegnaturismo.it/it/open-data",
                         "mensile", A, (_csv_rows(sard) or 0) + (_csv_rows(".cache/sardegna_flussi_mercato_mese.csv") or 0),
                         _fmt_mtime(sard)))
    turnot = ".cache/istat_turnot_scopo_regione.csv"
    if os.path.exists(turnot):
        rows.append(_row("Domanda residenti per scopo (ISTAT Viaggi e Vacanze)",
                         "Notti dei residenti per regione di DESTINAZIONE e scopo del viaggio "
                         "(lavoro / vacanza lunga / breve), annuale dal 2014 — tutte le regioni",
                         "ISTAT · Viaggi e Vacanze (DCCV_TURNOT)", "https://esploradati.istat.it/",
                         "annuale", A, _csv_rows(turnot), _fmt_mtime(turnot)))
    strp = ".cache/str_airbnb_territorio.csv"
    if os.path.exists(strp):
        rows.append(_row("Affitti brevi / STR (Inside Airbnb)",
                         "Mercato degli affitti brevi per territorio: prezzo (ADR), n. annunci, occupazione "
                         "proxy, qualità, operatori, licenze — 7 città + 3 regioni (Puglia/Sicilia/Trentino)",
                         "Inside Airbnb", "https://insideairbnb.com/get-the-data/",
                         "snapshot", A, (_csv_rows(strp) or 0) + (_csv_rows(".cache/str_airbnb_zona.csv") or 0),
                         _fmt_mtime(strp)))
    trends = sorted(glob.glob(".cache/trends_*.csv"))
    if trends:
        _geos = [mk.code for mk in DEFAULT_MARKETS]
        _full = sum(1 for info in RG.REGIONS.values()
                    if all(os.path.exists(f".cache/trends_{_safe_kw(info['trends_kw'])}_{g}.csv") for g in _geos))
        rows.append(_row(f"Google Trends — {_full}/{len(RG.REGIONS)} regioni × {len(_geos)} mercati",
                         "Interesse di ricerca per regione e per paese di origine (segnale leading)",
                         "Google Trends", "https://trends.google.com/", "mensile", A,
                         sum((_csv_rows(t) or 0) for t in trends), _fmt_mtime(trends[0])))
    rows.append(_row("Cambio valute", "Indice del cambio EUR vs USD/GBP/CHF (driver economico)",
                     "ECB Statistical Data Warehouse", "https://data.ecb.europa.eu/", "mensile", A, "—", "live"))
    bdi = ".cache/bdi_spend_per_market.csv"
    if os.path.exists(bdi):
        rows.append(_row("Spesa turisti stranieri", "Spesa €/viaggiatore per paese + per regione (valore economico)",
                         "Banca d'Italia · turismo internazionale",
                         "https://www.bancaditalia.it/statistiche/tematiche/rapporti-estero/turismo-internazionale/",
                         "trimestrale", A, _csv_rows(bdi), _fmt_mtime(bdi)))
    cap = ".cache/istat_capacity_letti_ITF1.csv"
    if os.path.exists(cap):
        rows.append(_row("Capacità ricettiva (posti letto)", "Posti letto per regione per anno (per l'occupazione)",
                         "ISTAT · Capacità esercizi ricettivi", "https://esploradati.istat.it/", "annuale", A,
                         _csv_rows(cap), _fmt_mtime(cap)))
    eu = ".cache/eurostat_pescara_flights.csv"
    if os.path.exists(eu):
        rows.append(_row("Connettività aerea per regione",
                         "Passeggeri voli diretti per paese verso ogni regione (fattibilità) — serie mensile per Pescara",
                         "Eurostat · avia_par", "https://ec.europa.eu/eurostat/", "annuale", A,
                         _csv_rows(eu), _fmt_mtime(eu)))
    wk = ".cache/wiki_de.csv"
    if os.path.exists(wk):
        rows.append(_row("Wikipedia pageviews", "Visualizzazioni pagina della regione per lingua (2° segnale leading)",
                         "Wikimedia · Pageviews API", "https://wikimedia.org/api/rest_v1/", "mensile", A,
                         _csv_rows(wk), _fmt_mtime(wk)))
    etd = ".cache/eurostat_travel_debit.csv"
    if os.path.exists(etd):
        rows.append(_row("Spesa turismo all'estero — UE/EFTA",
                         "Spesa dei residenti per turismo all'estero (taglia del mercato), paesi UE/EFTA, in M€",
                         "Eurostat · bop_its6_det (travel, debit)",
                         "https://ec.europa.eu/eurostat/databrowser/view/bop_its6_det",
                         "annuale", A, _csv_rows(etd), _fmt_mtime(etd)))
    wbx = ".cache/wb_tourism_expenditure.csv"
    if os.path.exists(wbx):
        rows.append(_row("Spesa turismo all'estero — extra-UE",
                         "Spesa per turismo all'estero dei mercati extra-UE (USA, Canada, Giappone, Russia; US$)",
                         "World Bank · ST.INT.XPND.CD",
                         "https://data.worldbank.org/indicator/ST.INT.XPND.CD",
                         "annuale", A, _csv_rows(wbx), _fmt_mtime(wbx)))
    geo = ".cache/world_countries.geojson"
    if os.path.exists(geo):
        rows.append(_row("Confini paesi (mappa)", "Geometrie dei confini per la mappa GIS (offline)",
                         "openpolis / johan world.geo.json", "https://github.com/openpolis/geojson-italy",
                         "statico", A, "—", _fmt_mtime(geo)))
    return rows


def load_registry() -> list[dict]:
    if os.path.exists(REGISTRY_PATH):
        try:
            return json.load(open(REGISTRY_PATH, encoding="utf-8"))
        except (OSError, ValueError):
            return []
    return []


def _save_registry(reg: list[dict]):
    os.makedirs(DATA_DIR, exist_ok=True)
    json.dump(reg, open(REGISTRY_PATH, "w", encoding="utf-8"), ensure_ascii=False, indent=2)


def save_upload(uploaded_file, nome: str, descrizione: str, fonte: str, url: str = "") -> dict:
    """Salva un file caricato in data/uploads/ e ne registra i metadati (nome, descrizione, fonte, URL)."""
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in uploaded_file.name)
    fid = f"{int(time.time())}_{safe}"
    path = os.path.join(UPLOAD_DIR, fid)
    data = uploaded_file.getvalue()
    with open(path, "wb") as f:
        f.write(data)
    righe = None
    if safe.lower().endswith(".csv"):
        try:
            righe = max(sum(1 for _ in io.StringIO(data.decode("utf-8", "ignore"))) - 1, 0)
        except ValueError:
            righe = None
    rec = {"id": fid, "nome": (nome or "").strip() or uploaded_file.name, "file": uploaded_file.name,
           "descrizione": (descrizione or "").strip(), "fonte": (fonte or "").strip(),
           "url": (url or "").strip(), "path": path,
           "caricato": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
           "dimensione_kb": round(len(data) / 1024, 1), "righe": righe}
    reg = load_registry(); reg.append(rec); _save_registry(reg)
    return rec


def delete_upload(rec_id: str):
    keep = []
    for r in load_registry():
        if r["id"] == rec_id:
            try:
                os.remove(r["path"])
            except OSError:
                pass
        else:
            keep.append(r)
    _save_registry(keep)


# ════════════════════════════════════════════════════════════════════════════
# FONTI CANDIDATE — scoperta con NULLA OSTA umano (Livello 1)
# ════════════════════════════════════════════════════════════════════════════
CAND_STATE_PATH = os.path.join(DATA_DIR, "candidates_state.json")

CANDIDATES = [
    {"id": "openmeteo", "nome": "Meteo Pescara (storico)",
     "descrizione": "Temperatura media e precipitazioni mensili a Pescara — driver della domanda balneare/montana.",
     "fonte": "Open-Meteo", "url": "https://open-meteo.com/", "frequenza": "mensile",
     "probe": CS.probe_open_meteo, "load": CS.load_open_meteo},
    {"id": "holidays", "nome": "Festività per paese",
     "descrizione": "Festività nazionali dei mercati (DE/AT/GB/US/NL/CH) — per il timing delle campagne.",
     "fonte": "Nager.Date", "url": "https://date.nager.at/", "frequenza": "annuale",
     "probe": CS.probe_holidays, "load": CS.load_holidays},
    {"id": "istat_viaggi", "nome": "Viaggi dei residenti abruzzesi",
     "descrizione": "Viaggi/notti dei residenti dell'Abruzzo per destinazione (le «partenze»).",
     "fonte": "ISTAT · DCCV_TURNOT", "url": "https://esploradati.istat.it/", "frequenza": "annuale",
     "probe": CS.probe_istat_viaggi, "load": CS.load_istat_viaggi},
    {"id": "musei", "nome": "Musei statali — visitatori e introiti",
     "descrizione": "Visitatori e introiti dei musei statali per istituto (1996-2024). NAZIONALE, "
                    "filtrabile per regione → segnale di domanda culturale del territorio.",
     "fonte": "MiC · dati.gov.it", "frequenza": "annuale",
     "url": "https://www.dati.gov.it/opendata/dataset/musei-statali-visitatori-ed-introiti-1996-2024",
     "probe": CS.probe_musei, "load": CS.load_musei},
    {"id": "aeroporti", "nome": "Mappa aeroporti commerciali italiani",
     "descrizione": "Anagrafica degli aeroporti italiani aperti al traffico (coord., IATA, località). "
                    "NAZIONALE → base per la mappa regione→aeroporto (accessibilità multi-regione).",
     "fonte": "Autorità Trasporti · dati.gov.it", "frequenza": "annuale",
     "url": "https://www.dati.gov.it/opendata/dataset/mappa-degli-aeroporti-aperti-al-traffico-commerciale",
     "probe": CS.probe_aeroporti, "load": CS.load_aeroporti},
    {"id": "ferrovia", "nome": "Traffico passeggeri ferroviario (nazionale)",
     "descrizione": "Passeggeri ferroviari per anno (totale nazionale). Indicatore di contesto "
                    "sull'accessibilità ferroviaria; non disaggregato per regione.",
     "fonte": "Autorità Trasporti · dati.gov.it", "frequenza": "annuale",
     "url": "https://www.dati.gov.it/opendata/dataset/evoluzione-del-traffico-passeggeri-per-ferrovia",
     "probe": CS.probe_ferrovia, "load": CS.load_ferrovia},
]


def candidates_state() -> dict:
    if os.path.exists(CAND_STATE_PATH):
        try:
            return json.load(open(CAND_STATE_PATH, encoding="utf-8"))
        except (OSError, ValueError):
            return {}
    return {}


def _save_candidates_state(s: dict):
    os.makedirs(DATA_DIR, exist_ok=True)
    json.dump(s, open(CAND_STATE_PATH, "w", encoding="utf-8"), ensure_ascii=False, indent=2)


def candidate_by_id(cid: str):
    return next((c for c in CANDIDATES if c["id"] == cid), None)


def approve_candidate(cid: str, descrizione: str) -> dict:
    """Carica (load) la fonte candidata DOPO il nulla osta e la marca approvata."""
    c = candidate_by_id(cid)
    if not c:
        return {"ok": False, "msg": "candidato sconosciuto"}
    try:
        res = c["load"]()
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "msg": f"caricamento fallito: {type(e).__name__}: {str(e)[:120]}"}
    st_ = candidates_state()
    st_[cid] = {"approved": True, "descrizione": (descrizione or c["descrizione"]).strip(),
                "caricato": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                "righe": res.get("rows"), "path": res.get("path")}
    _save_candidates_state(st_)
    return res


def revoke_candidate(cid: str):
    st_ = candidates_state()
    st_.pop(cid, None)
    _save_candidates_state(st_)


def present_sources() -> list[dict]:
    """Tabella DATI PRESENTI: fonti in uso + file caricati + candidati approvati (colonne unificate)."""
    rows = list(builtin_sources())
    for u in load_registry():
        url = u.get("url") or (u.get("fonte", "") if str(u.get("fonte", "")).startswith("http") else "")
        rows.append(_row(u.get("nome"), u.get("descrizione", ""), u.get("fonte", "caricato manualmente"),
                         url, "—", "📤 Caricato", u.get("righe"), u.get("caricato")))
    state = candidates_state()
    for c in CANDIDATES:
        s = state.get(c["id"])
        if s and s.get("approved"):
            rows.append(_row(c["nome"], s.get("descrizione", c["descrizione"]), c["fonte"], c["url"],
                             c["frequenza"], "🟢 Attiva (approvata)", s.get("righe"), s.get("caricato")))
    return rows


def pending_candidates() -> list[dict]:
    """Tabella DATI IN VALUTAZIONE: candidati non ancora approvati (colonne unificate)."""
    state = candidates_state()
    rows = []
    for c in CANDIDATES:
        if state.get(c["id"], {}).get("approved"):
            continue
        rows.append(_row(c["nome"], state.get(c["id"], {}).get("descrizione", c["descrizione"]),
                         c["fonte"], c["url"], c["frequenza"], "⏳ In valutazione", "—", "—"))
    return rows


COVERAGE_COLS = ["Informazione", "Nazionale", "Regionale", "Provinciale", "Fonte", "Note / buco da coprire"]


def coverage_matrix() -> list[dict]:
    """Mappa cosa abbiamo per LIVELLO geografico (nazionale/regionale/provinciale).
    ✅ disponibile · 🟡 parziale · ❌ assente. Serve a vedere i buchi per il multi-regione."""
    M = [
        ("Presenze turistiche (totali)", "✅", "✅", "✅", "ISTAT DCSC_TUR", "completo per tutte le regioni"),
        ("Presenze straniere (aggregato)", "✅", "✅", "✅", "ISTAT DCSC_TUR", "stranieri totali, non per singolo paese"),
        ("Presenze per PAESE di origine", "✅", "✅", "✅", "ISTAT DCSC_TUR_9", "✅ risolto: annuale dal 2008 per regione E provincia (cube _9); il mensile per-paese resta solo nazionale"),
        ("Capacità ricettiva (posti letto)", "✅", "✅", "🟡", "ISTAT DCSC_TUR_1", "regionale ok; provinciale solo alcune"),
        ("Flussi MENSILI per categoria ricettiva", "❌", "❌", "🟡", "ASTAT (Bolzano)", "🟡 solo Alto Adige: arrivi/presenze mensili + esercizi/posti letto per stelle/categoria (fonte provinciale complementare)"),
        ("Mercato ESTERO mensile sub-nazionale", "❌", "🟡", "🟡", "Lombardia + Sardegna", "🟡 Lombardia (12 province) e Sardegna (per comune!): presenze/arrivi mensili per paese estero di origine; ISTAT lo dà solo annuale (_9)"),
        ("Movimento a livello COMUNALE", "❌", "🟡", "🟡", "Toscana + Sardegna", "🟡 Toscana (~272 comuni, annuale, ita/str) e Sardegna (~330 comuni, MENSILE × mercato estero): sotto il livello provinciale ISTAT"),
        ("Anagrafica strutture ricettive", "🟡", "❌", "❌", "registri regionali", "🔴 BUCO: registri non federati come open data"),
        ("Affitti brevi / STR (prezzo, annunci)", "❌", "🟡", "🟡", "Inside Airbnb", "🟡 mercato affitti brevi: 7 città + 3 regioni intere (Puglia/Sicilia/Trentino), NIENTE Abruzzo; ADR/n.annunci/occ.proxy/licenze per annuncio (CC BY 4.0)"),
        ("Spesa turistica per paese", "✅", "🟡", "🟡", "BdI + stima da _9", "regionale/provinciale STIMATA (presenze _9 × spesa/notte BdI); dato diretto solo TOTALE regionale (BdI TS2)"),
        ("Accessibilità aerea (voli/pax)", "✅", "✅", "—", "Eurostat avia_par", "regionale via mappa regione→aeroporto"),
        ("Accessibilità ferroviaria", "✅", "❌", "❌", "Aut. Trasporti", "🟡 solo totale nazionale; manca il dettaglio regionale"),
        ("Interesse online — Google Trends", "✅", "✅", "—", "Google Trends", "per regione: keyword = nome regione"),
        ("Interesse online — Wikipedia", "✅", "✅", "—", "Wikimedia", "per regione: articolo della regione"),
        ("Cambio valute (FX mercati)", "✅", "—", "—", "ECB", "nazionale per i mercati esteri (non è dato regionale)"),
        ("Fiducia consumatori (mercati)", "✅", "—", "—", "Eurostat", "per paese estero (DE/AT/NL); GB/CH/US via OCSE"),
        ("Domanda culturale (musei)", "✅", "✅", "🟡", "MiC (candidato)", "per istituto → filtrabile per regione"),
        ("Meteo destinazione", "✅", "✅", "✅", "Open-Meteo (candidato)", "per coordinate: qualunque capoluogo/comune"),
        ("Festività mercati esteri", "✅", "—", "—", "Nager.Date (candidato)", "per paese estero (timing campagne)"),
        ("Eventi / POI turistici", "🟡", "🟡", "🟡", "open data regionali", "🟡 alcune regioni pubblicano, molte no (Abruzzo scarso)"),
        ("Viaggi residenti per regione d'origine", "✅", "❌", "❌", "ISTAT Viaggi&Vacanze", "🔴 open data solo nazionale/macro-area; il dettaglio regionale è solo MICRODATI (richiesta istituzionale ISTAT)"),
        ("Domanda residenti per SCOPO (destinazione)", "✅", "✅", "❌", "ISTAT Viaggi&Vacanze", "✅ a livello REGIONALE: notti dei residenti per regione di DESTINAZIONE × scopo (lavoro/vacanza lunga/breve), annuale dal 2014 (TURNOT_CAPI_1)"),
        ("Confini amministrativi (geojson)", "✅", "✅", "✅", "openpolis / RNDT", "completo: usato per la mappa d'Italia"),
    ]
    return [dict(zip(COVERAGE_COLS, r)) for r in M]


# ════════════════════════════════════════════════════════════════════════════
# AZIONI RACCOMANDATE + BOZZA CAMPAGNA  (ispirate al "Suggerimento AI" del POC)
# ════════════════════════════════════════════════════════════════════════════
def peak_search_month(panel, code: str):
    col = f"search_{code}"
    if panel is None or col not in panel.columns:
        return None
    d = panel.copy(); d["m"] = d["date"].dt.month
    prof = d.groupby("m")[col].mean()
    return int(prof.idxmax()) if not prof.empty else None


def recommended_actions(ctx: dict) -> list[dict]:
    """Trasforma il ranking in azioni operative con priorità (Alta/Media/Bassa)."""
    s = markets_summary(ctx)
    panel = ctx["R"]["panel"] if ctx["mode"] == "real" else None
    acts = []
    for m in s:
        cat = reco_cat(m["reco"])
        forza, mom = m.get("forza") or 0, m.get("momentum") or 0
        if cat == "Aumentare" and forza >= 0.4 and mom > 20:
            prio = "Alta"
        elif cat == "Aumentare":
            prio = "Media"
        elif cat == "Ridurre":
            prio = "Media"
        else:
            prio = "Bassa"
        acts.append({"market": m["market"], "code": m["code"], "reco": m["reco"], "cat": cat,
                     "priorita": prio, "forza": m.get("forza"), "momentum": m.get("momentum"),
                     "valore": m.get("valore"), "score": m["score"], "rank": m["rank"],
                     "picco_mese": peak_search_month(panel, m["code"])})
    order = {"Alta": 0, "Media": 1, "Bassa": 2}
    return sorted(acts, key=lambda a: (order[a["priorita"]], -a["score"]))


def campaign_brief(ctx: dict, code: str) -> dict | None:
    a = next((x for x in recommended_actions(ctx) if x["code"] == code), None)
    if not a:
        return None
    pk = MONTHS_IT[a["picco_mese"] - 1] if a["picco_mese"] else "—"
    lag = ctx["R"]["agg"]["lag"] if ctx["mode"] == "real" else 2
    leva = []
    if a["valore"]:
        leva.append(f"alto valore economico (~€{a['valore']:.0f}/visitatore)")
    if a["momentum"] and a["momentum"] > 0:
        leva.append(f"interesse di ricerca in crescita ({a['momentum']:+.0f}% su anno prima)")
    if a["forza"] and a["forza"] > 0.3:
        leva.append(f"segnale anticipatore solido (forza {a['forza']:+.2f})")
    return {
        "mercato": a["market"], "priorita": a["priorita"], "azione": a["reco"],
        "finestra": f"presidiare il picco di ricerca (~{pk}); il search anticipa gli arrivi di ~{lag} mesi, "
                    f"quindi avviare la campagna prima del picco di presenze",
        "leva": "; ".join(leva) or "monitorare il segnale prima di investire",
        "kpi": "presenze straniere dal mercato · costo per arrivo · quota di ricerca 'Abruzzo' nel paese",
        "nota": "Stima d'opportunità (dove e quando conviene agire), NON una garanzia di ritorno: "
                "il motore non modella l'effetto causale della spesa promozionale.",
    }


# ════════════════════════════════════════════════════════════════════════════
# ARCHITETTURA & SORGENTI (dal documento tecnico TDH — vision + stato attuale)
# ════════════════════════════════════════════════════════════════════════════
def tdh_architecture() -> dict:
    return {
        "livelli": [
            ("1 · Sorgenti dati", "Portali regionali, MaaS, registri ricettivi, Hub nazionale PDND"),
            ("2 · Ingestion", "Batch ETL · Streaming · API connectors · connettore PDND"),
            ("3 · Data Lake (Medallion)", "Bronze (grezzo) → Silver (normalizzato/anonimizzato GDPR) → Gold (KPI)"),
            ("4 · Governance & processing", "Trasformazioni · GDPR engine · identity resolution · Auth/ACL (SPID/CIE)"),
            ("5 · Servizi di output", "Cruscotto Regione · API Operatori · Analytics ML · integrazione MaaS"),
        ],
        "sorgenti": [
            {"Sorgente": "ISTAT — movimento clienti", "Tipo": "Presenze per provenienza",
             "Frequenza": "Mensile", "Stato": "🟢 Attiva"},
            {"Sorgente": "ECB — cambio valute", "Tipo": "Driver economico (FX)",
             "Frequenza": "Mensile", "Stato": "🟢 Attiva"},
            {"Sorgente": "Google Trends", "Tipo": "Segnale leading di ricerca",
             "Frequenza": "Mensile", "Stato": "🟢 Attiva"},
            {"Sorgente": "Banca d'Italia — turismo internazionale", "Tipo": "Spesa per paese/regione (valore economico)",
             "Frequenza": "Trimestrale", "Stato": "🟢 Attiva"},
            {"Sorgente": "Eurostat — trasporto aereo (avia_par)", "Tipo": "Connettività voli per paese (fattibilità)",
             "Frequenza": "Annuale", "Stato": "🟢 Attiva"},
            {"Sorgente": "Wikipedia (Wikimedia Pageviews)", "Tipo": "Interesse per lingua (2° segnale leading)",
             "Frequenza": "Mensile", "Stato": "🟢 Attiva"},
            {"Sorgente": "PDND — Ministero Turismo SIT/CIR", "Tipo": "Statistiche turismo nazionali",
             "Frequenza": "Periodica", "Stato": "🟡 Pianificata"},
            {"Sorgente": "PDND — ANPR", "Tipo": "Presenze cittadini",
             "Frequenza": "Periodica", "Stato": "🟡 Pianificata"},
            {"Sorgente": "Portale Regionale (GA4)", "Tipo": "Web analytics, funnel visitatori",
             "Frequenza": "Giornaliera", "Stato": "🟡 Pianificata"},
            {"Sorgente": "MaaS4Abruzzo (GTFS-RT)", "Tipo": "Mobilità, percorsi, viaggi",
             "Frequenza": "Real-time", "Stato": "🟡 Pianificata"},
            {"Sorgente": "Registro Ricettivo (~2.800 strutture)", "Tipo": "Offerta ricettiva",
             "Frequenza": "Batch", "Stato": "🟡 Pianificata"},
            {"Sorgente": "App TUA", "Tipo": "Ticketing / transiti",
             "Frequenza": "Real-time", "Stato": "⚪ Da valutare"},
            {"Sorgente": "Meteo ARPAM", "Tipo": "Dati territoriali",
             "Frequenza": "Oraria", "Stato": "⚪ Da valutare"},
        ],
        "roadmap": [
            "Fase 1 — Fondamenta: sorgenti nazionali (PDND/ISTAT) + primo motore decisionale. "
            "*Questo MVP è il primo mattone reale.*",
            "Fase 2 — Integrazione regionale: portale GA4, MaaS4Abruzzo, registro ricettivo.",
            "Fase 3 — Analytics ML: forecasting flussi, segmentazione turisti, alerting anomalie.",
            "Fase 4 — Servizi: cruscotto Regione, API operatori, integrazione MaaS in tempo reale.",
        ],
        "principi": [
            "**GDPR by design** — anonimizzazione PII nelle zone Silver/Gold, nessun dato nominale.",
            "**Cloud PA-first** — infrastruttura su provider qualificato ACN (PSNC), dati nell'UE.",
            "**Non sostitutivo** — il TDH si appoggia sopra i sistemi esistenti come layer di integrazione.",
            "**PDND** — sorgenti nazionali senza convenzioni bilaterali (art. 50-ter CAD).",
        ],
        "motore_statistico": [
            "**Spina dorsale** — modello lineare trasparente (OLS): stagionalità esplicita (effetti di "
            "mese) + trend + dummy COVID. Ogni coefficiente è leggibile in italiano corrente.",
            "**Barriera di onestà** — si prevede un numero solo se il modello batte la *naive stagionale* "
            "(ripetere lo stesso periodo dell'anno prima) in un backtest; altrimenti «segnale insufficiente».",
            "**Sfidanti** (si tiene il migliore, serie per serie): ETS/Holt-Winters, SARIMAX, stato latente "
            "(UCM), Theta, STL+ARIMA — scelti con backtest **rolling-origin** (più finestre mobili, scelta "
            "affidabile); + opzione di trend **robusto** (Huber) per gli outlier.",
            "**Tra regioni** — *partial pooling* (empirical-Bayes): stabilizza le serie corte tirando il "
            "trend di ogni regione verso la media nazionale; + riconciliazione Italia↔regioni (i totali tornano).",
            "**Confine** — è un **ranking decisionale** (forza anticipatrice × momentum × valore × "
            "fattibilità), NON una stima causale della spesa promozionale.",
            "*Evidenza:* sui dati reali i modelli complessi non battono di norma la naive stagionale → la "
            "complessità si attiva solo se supera il backtest (vedi script `bakeoff.py` e il documento di studio).",
        ],
    }


# ════════════════════════════════════════════════════════════════════════════
# VISTE TERRITORIALI — per Provincia e per tipologia di Struttura (dati ISTAT reali)
# ════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner="Carico le presenze per provincia (ISTAT)…")
def compute_provinces(code: str | None = None) -> dict:
    """Presenze per provincia (NUTS3) della regione richiesta. Multi-regione: le
    province vengono dal registro; le serie che ISTAT non espone vengono saltate."""
    from tourism_wedge.real_sources import fetch_istat_presences
    code = code or RG.DEFAULT_REGION
    rows, wide, wide_str = [], None, None
    for area, nome in RG.provinces(code).items():
        try:  # il TOTALE (WORLD) è indispensabile per mostrare la provincia
            tot = fetch_istat_presences(area=area, country="WORLD").rename(columns={"presences": nome})
        except Exception:  # noqa: BLE001 — provincia non esposta/abolita: salto
            continue
        est_wide = None
        try:  # gli stranieri possono mancare per qualche provincia: in quel caso restano vuoti
            est = fetch_istat_presences(area=area, country="WRL_X_ITA").sort_values("date")
            e12 = float(est.tail(12)["presences"].sum())
            est_wide = est.rename(columns={"presences": nome})[["date", nome]]  # serie mensile stranieri
        except Exception:  # noqa: BLE001
            e12 = float("nan")
        wide = tot if wide is None else wide.merge(tot, on="date", how="outer")
        if est_wide is not None:
            wide_str = est_wide if wide_str is None else wide_str.merge(est_wide, on="date", how="outer")
        t12 = float(tot.sort_values("date").tail(12)[nome].sum())
        rows.append({"provincia": nome, "presenze": t12, "stranieri": e12,
                     "quota_stranieri": (e12 / t12 * 100 if (t12 and pd.notna(e12)) else float("nan"))})
    empty = pd.DataFrame({"date": []})
    if wide is None:
        return {"rows": [], "panel": empty, "panel_str": empty}
    return {"rows": sorted(rows, key=lambda r: -r["presenze"]),
            "panel": wide.sort_values("date"),
            "panel_str": (wide_str.sort_values("date") if wide_str is not None else empty)}


@st.cache_data(show_spinner="Carico le presenze per struttura (ISTAT)…")
def compute_structure(code: str | None = None) -> dict:
    from tourism_wedge.real_sources import fetch_istat_presences
    area = RG.istat_area(code)
    hot = fetch_istat_presences(area=area, accom="HOTELLIKE", country="WORLD").rename(columns={"presences": "alberghiero"})
    oth = fetch_istat_presences(area=area, accom="OTHER", country="WORLD").rename(columns={"presences": "extra"})
    panel = hot.merge(oth, on="date", how="outer").sort_values("date")
    return {"alberghiero": float(hot.tail(12)["alberghiero"].sum()),
            "extra": float(oth.tail(12)["extra"].sum()), "panel": panel}


@st.cache_data(show_spinner=False)
def _italy_provinces_geojson():
    p = "assets/italy_provinces.geojson"
    return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else None


def chart_province_map(rows: list[dict]) -> go.Figure | None:
    gj = _italy_provinces_geojson()
    if gj is None or not rows:
        return None
    names = [r["provincia"] for r in rows]
    fig = go.Figure(go.Choropleth(
        geojson=gj, featureidkey="properties.prov_name", locations=names, name="presenze",
        z=[r["presenze"] for r in rows], colorscale="Teal", marker_line_color="white", marker_line_width=1,
        colorbar=dict(title=dict(text="presenze (12 mesi)", side="right")),
        customdata=[[r["stranieri"], r["quota_stranieri"]] for r in rows],
        hovertemplate="<b>%{location}</b><br>presenze %{z:,.0f}"
                      "<br>di cui stranieri %{customdata[0]:,.0f} (%{customdata[1]:.0f}%)<extra></extra>"))
    fig.update_geos(fitbounds="locations", visible=False, projection_type="mercator", bgcolor="rgba(0,0,0,0)")
    fig.update_layout(height=460, margin=dict(l=0, r=0, t=10, b=0), paper_bgcolor="white",
                      title_text="", font=dict(family=FONT_STACK))
    return fig


def chart_province_trend(panel, rangeslider: bool = False) -> go.Figure:
    fig = go.Figure()
    for col in [c for c in panel.columns if c != "date"]:
        d = panel[["date", col]].dropna()
        fig.add_trace(go.Scatter(x=d["date"], y=d[col], name=col, mode="lines+markers"))
    fig.update_yaxes(title="presenze")
    if rangeslider:
        fig.update_xaxes(rangeslider_visible=True, rangeslider_thickness=0.08)
    return _layout(fig, h=400 if rangeslider else 360)


def province_yoy(panel) -> list[dict]:
    """Variazione % anno-su-anno per provincia: somma ultimi 12 mesi vs 12 precedenti
    (sul totale presenze). Ordinata dal calo più forte alla crescita più forte."""
    if panel is None or panel.empty:
        return []
    d = panel.sort_values("date")
    out = []
    for c in [c for c in panel.columns if c != "date"]:
        s = d[["date", c]].dropna()
        if len(s) < 24:  # servono due finestre da 12 mesi
            continue
        cur = float(s[c].tail(12).sum())
        prev = float(s[c].iloc[-24:-12].sum())
        var = (cur / prev - 1) * 100 if prev else float("nan")
        out.append({"provincia": c, "cur": cur, "prev": prev, "var_pct": var})
    out.sort(key=lambda r: (r["var_pct"] if r["var_pct"] == r["var_pct"] else -9e99))
    return out


def chart_province_yoy(yoy: list[dict]) -> go.Figure:
    data = [r for r in yoy if r.get("var_pct") == r.get("var_pct")]
    if not data:
        return _layout(go.Figure(), h=260)
    fig = go.Figure(go.Bar(
        x=[r["var_pct"] for r in data], y=[r["provincia"] for r in data], orientation="h",
        marker_color=["#16a34a" if r["var_pct"] >= 0 else "#dc2626" for r in data],
        hovertemplate="<b>%{y}</b><br>%{x:+.1f}% a/a<extra></extra>"))
    fig.update_xaxes(title="variazione presenze · ultimi 12 mesi vs 12 precedenti (%)",
                     ticksuffix="%", zeroline=True, zerolinecolor="#94a3b8", zerolinewidth=1)
    return _layout(fig, h=max(260, 90 + 30 * len(data)))


def chart_province_seasonality(panel) -> go.Figure:
    """Heatmap mese × provincia normalizzata: per ogni provincia il mese di picco = 100%.
    Mostra QUANDO lavora ciascuna provincia (costa estiva vs montagna invernale), a parità
    di scala (non conta la dimensione)."""
    if panel is None or panel.empty:
        return _layout(go.Figure(), h=260)
    d = panel.copy()
    d["m"] = d["date"].dt.month
    cols = [c for c in panel.columns if c != "date"]
    prof = d.groupby("m")[cols].mean().reindex(range(1, 13))
    z, prov = [], []
    for c in cols:
        col = prof[c]
        mx = col.max()
        if not mx or mx != mx:  # provincia senza dati
            continue
        z.append([(None if v != v else v / mx * 100) for v in col])
        prov.append(c)
    if not z:
        return _layout(go.Figure(), h=260)
    fig = go.Figure(go.Heatmap(
        z=z, x=MONTHS_IT, y=prov, colorscale="Teal", zmin=0, zmax=100,
        colorbar=dict(title=dict(text="% del picco", side="right")),
        hovertemplate="<b>%{y}</b><br>%{x}: %{z:.0f}% del picco<extra></extra>"))
    return _layout(fig, h=max(260, 80 + 30 * len(prov)))


def chart_structure_donut(s: dict) -> go.Figure:
    fig = go.Figure(go.Pie(labels=["Alberghiero", "Extra-alberghiero"],
                           values=[s["alberghiero"], s["extra"]], hole=0.55,
                           marker=dict(colors=["#0e6b70", "#f59e0b"])))
    return _layout(fig, h=320)


def chart_structure_trend(panel) -> go.Figure:
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=panel["date"], y=panel["alberghiero"], name="Alberghiero",
                             line=dict(color="#0e6b70", width=2)))
    fig.add_trace(go.Scatter(x=panel["date"], y=panel["extra"], name="Extra-alberghiero",
                             line=dict(color="#f59e0b", width=2)))
    fig.update_yaxes(title="presenze")
    return _layout(fig, h=340)


# ════════════════════════════════════════════════════════════════════════════
# OCCUPAZIONE REALE — indice di utilizzazione lorda dei posti letto (ISTAT)
#   occ = presenze ÷ (posti letto × giorni del mese)
# presenze totali Abruzzo = somma delle 4 province (WORLD); posti letto: dataset Capacità.
# ════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner="Calcolo l'occupazione (ISTAT)…")
def compute_occupancy(code: str | None = None) -> dict:
    from tourism_wedge.real_sources import fetch_istat_presences, fetch_istat_capacity
    area = RG.istat_area(code)
    try:
        beds = fetch_istat_capacity(area=area)
        tot = fetch_istat_presences(area=area, country="WORLD")
    except Exception:  # noqa: BLE001 — ISTAT instabile/serie assente
        return {"available": False}
    beds_map = dict(zip(beds["anno"].astype(int), beds["letti"].astype(float)))
    if not beds_map or tot.empty:
        return {"available": False}
    panel = tot.rename(columns={"presences": "presenze"})[["date", "presenze"]].dropna().copy()
    last_year = max(beds_map)
    panel["letti"] = panel["date"].dt.year.map(lambda y: beds_map.get(int(y), beds_map[last_year]))
    panel["giorni"] = panel["date"].dt.daysinmonth
    panel["occ"] = panel["presenze"] / (panel["letti"] * panel["giorni"]) * 100
    occ12 = float(panel.sort_values("date").tail(12)["occ"].mean())
    return {"available": True, "panel": panel.sort_values("date"), "occ_media12": occ12,
            "letti_ultimo": int(beds_map[last_year]), "anno_letti": int(last_year)}


def chart_occupancy(panel) -> go.Figure:
    fig = go.Figure(go.Scatter(x=panel["date"], y=panel["occ"], line=dict(color="#0e6b70", width=2),
                               name="occupazione", fill="tozeroy", fillcolor="rgba(14,116,144,.12)"))
    fig.update_yaxes(title="occupazione lorda", ticksuffix="%")
    return _layout(fig, h=360)


# ── MULTI-REGIONE: quadro ISTAT per qualunque regione (NUTS2) ─────────────────
@st.cache_data(show_spinner="Carico i dati ISTAT della regione…")
def region_overview(code: str) -> dict:
    """Presenze mensili (totale + stranieri) e capacità della regione richiesta.
    Funziona per tutte le 20 regioni (ISTAT per NUTS2). Cache per-regione."""
    info = RG.region(code)
    area = RG.istat_area(code)
    stran = RS.fetch_istat_presences(area=area, country="WRL_X_ITA", start="2019-01")
    tot = RS.fetch_istat_presences(area=area, country="WORLD", start="2019-01")
    df = (stran.rename(columns={"presences": "stranieri"})
          .merge(tot.rename(columns={"presences": "totale"}), on="date", how="outer")
          .sort_values("date").reset_index(drop=True))
    letti = anno = None
    try:
        cap = RS.fetch_istat_capacity(area=area)
        letti, anno = int(cap["letti"].iloc[-1]), int(cap["anno"].iloc[-1])
    except Exception:  # noqa: BLE001
        pass
    return {"info": info, "presenze": df, "letti": letti, "anno_letti": anno}


def chart_region_presences(df) -> go.Figure:
    fig = go.Figure()
    if "totale" in df:
        fig.add_trace(go.Scatter(x=df["date"], y=df["totale"], name="totale",
                                 mode="lines", line=dict(color="#0e6b70", width=2)))
    fig.add_trace(go.Scatter(x=df["date"], y=df["stranieri"], name="stranieri",
                             mode="lines", line=dict(color="#f59e0b", width=2)))
    fig.update_yaxes(title="presenze (mese)")
    return _layout(fig, h=360)


# ── Pannello ANNUALE multi-variabile per la scheda Regione (base per le proiezioni) ──
# Registro variabili: chiave · etichetta · unità · decimali. Le derivate (quota, spesa/turista,
# occupazione) si calcolano da quelle grezze. Spesa/notti/viaggiatori = SOLO stranieri (BdI).
REGION_VARS = [
    ("presenze_tot", "Presenze totali", "n", 0),
    ("presenze_str", "Presenze straniere", "n", 0),
    ("quota_str", "Quota stranieri", "%", 1),
    ("spesa", "Spesa straniera", "M€", 0),
    ("spesa_per_viagg", "Spesa per viaggiatore", "€", 0),
    ("spesa_per_notte", "Spesa per notte", "€", 0),
    ("viaggiatori", "Viaggiatori stranieri", "migliaia", 0),
    ("notti", "Pernottamenti stranieri", "migliaia", 0),
    ("letti", "Posti letto", "n", 0),
    ("occ", "Occupazione", "%", 1),
]
REGION_VAR_LABEL = {k: lab for k, lab, _u, _d in REGION_VARS}
REGION_VAR_UNIT = {k: u for k, _l, u, _d in REGION_VARS}
REGION_VAR_DEC = {k: d for k, _l, _u, d in REGION_VARS}


@st.cache_data(show_spinner="Costruisco il quadro pluriennale…")
def region_annual_panel(code: str) -> pd.DataFrame:
    """Pannello ANNUALE (anno × variabili) per regione o Italia. Unisce ISTAT (presenze
    mensili→somma annua su anni completi, capacità) e Banca d'Italia (spesa/notti/viaggiatori,
    anni completi). Aggiunge le derivate: quota stranieri, spesa/viaggiatore, spesa/notte,
    occupazione. Le celle mancanti restano NaN (le serie partono da anni diversi)."""
    out: dict[int, dict] = {}
    # ISTAT presenze: mensile → somma annua, solo anni con 12 mesi
    try:
        p = region_overview(code)["presenze"].copy()
        p["anno"] = p["date"].dt.year
        for col, key in (("totale", "presenze_tot"), ("stranieri", "presenze_str")):
            if col in p:
                g = p.dropna(subset=[col]).groupby("anno")[col].agg(["sum", "count"])
                for y, r in g.iterrows():
                    if r["count"] >= 12:
                        out.setdefault(int(y), {})[key] = float(r["sum"])
    except Exception:  # noqa: BLE001
        pass
    # Banca d'Italia: spesa/notti/viaggiatori (stranieri), solo anni completi (4 trimestri)
    g = bdi_region_annual(code)
    if g is not None and not g.empty:
        for _, r in g.iterrows():
            if r["trimestri"] >= 4:
                d = out.setdefault(int(r["anno"]), {})
                d["spesa"], d["notti"], d["viaggiatori"] = float(r["spesa"]), float(r["notti"]), float(r["viaggiatori"])
    # ISTAT capacità: posti letto annuali
    try:
        cap = RS.fetch_istat_capacity(area=RG.istat_area(code))
        for _, r in cap.iterrows():
            out.setdefault(int(r["anno"]), {})["letti"] = float(r["letti"])
    except Exception:  # noqa: BLE001
        pass
    if not out:
        return pd.DataFrame()
    df = pd.DataFrame(out).T.sort_index()
    df.index.name = "anno"
    # derivate
    if {"presenze_str", "presenze_tot"} <= set(df.columns):
        df["quota_str"] = df["presenze_str"] / df["presenze_tot"] * 100
    if {"spesa", "viaggiatori"} <= set(df.columns):
        df["spesa_per_viagg"] = df["spesa"] * 1e6 / (df["viaggiatori"] * 1e3)
    if {"spesa", "notti"} <= set(df.columns):
        df["spesa_per_notte"] = df["spesa"] * 1e6 / (df["notti"] * 1e3)
    if {"presenze_tot", "letti"} <= set(df.columns):
        df["occ"] = df["presenze_tot"] / (df["letti"] * 365) * 100
    # ordina le colonne come nel registro
    cols = [k for k, *_ in REGION_VARS if k in df.columns]
    return df[cols]


def chart_region_indexed(df: pd.DataFrame, keys: list[str], labels: dict | None = None) -> go.Figure:
    """Linee a INDICE (base 100 = primo anno della finestra) per le variabili scelte:
    confronta i TREND relativi a prescindere dall'unità (es. turisti vs spesa/turista).
    `labels` mappa chiave→etichetta (default: variabili Regione)."""
    labels = labels or REGION_VAR_LABEL
    palette = ["#0e6b70", "#f59e0b", "#7c3aed", "#059669", "#dc2626",
               "#2563eb", "#db2777", "#65a30d", "#0e6b70", "#ca8a04"]
    fig = go.Figure()
    for i, k in enumerate(keys):
        if k not in df:
            continue
        s = df[k].dropna()
        if s.empty or not s.iloc[0]:
            continue
        fig.add_trace(go.Scatter(x=s.index.astype(int), y=s / s.iloc[0] * 100,
                                 name=labels.get(k, k), mode="lines+markers",
                                 line=dict(color=palette[i % len(palette)], width=2)))
    fig.add_hline(y=100, line=dict(color="#94a3b8", dash="dot"))
    fig.update_yaxes(title="indice (base 100 = primo anno)")
    fig.update_xaxes(title="anno", dtick=1)
    return _layout(fig, h=380)


def project_var(panel: pd.DataFrame, key: str, horizon: int = 2, drop_covid: bool = True,
                n_years: int | None = 5, robust: bool = False) -> dict | None:
    """Proiezione a trend LINEARE della variabile annuale `key` su `horizon` anni, con
    intervallo di previsione all'80%. `n_years` limita la BASE del trend agli ultimi N anni
    (None = tutto lo storico): utile quando c'è un cambio di regime (es. salto post-COVID
    della spesa/turista). `drop_covid` esclude il 2020 prima di contare gli anni.
    `robust=True` (Tier 1): stima con regressione robusta di Huber (statsmodels RLM), che
    pesa di meno gli anni anomali (outlier tipo COVID) invece di lasciarli inclinare la retta;
    la banda 80% deriva dalla scala robusta dei residui."""
    import numpy as np
    import statsmodels.api as sm
    if panel is None or panel.empty or key not in panel:
        return None
    s = panel[key].dropna()
    s.index = s.index.astype(int)
    if drop_covid and 2020 in s.index:
        s = s.drop(index=2020)
    if n_years:
        s = s.tail(int(n_years))
    if len(s) < 3:
        return None
    x, y = s.index.values.astype(float), s.values.astype(float)
    last = int(s.index.max())
    fut = list(range(last + 1, last + 1 + int(horizon)))
    X, Xf = sm.add_constant(x), sm.add_constant(np.array(fut, dtype=float), has_constant="add")
    if robust:
        m = sm.RLM(y, X, M=sm.robust.norms.HuberT()).fit()
        fit_vals, fc_mean = m.predict(X), m.predict(Xf)
        # banda 80% ≈ ±z(0.9)·scala robusta, allargata in estrapolazione (leverage)
        xbar = x.mean(); Sxx = float(((x - xbar) ** 2).sum()) or 1.0
        widen = np.sqrt(1.0 + (np.array(fut, float) - xbar) ** 2 / Sxx)
        half = 1.2816 * float(m.scale) * widen
        fc_lo, fc_hi = fc_mean - half, fc_mean + half
        r2 = _pseudo_r2(y, fit_vals)
    else:
        m = sm.OLS(y, X).fit()
        fit_vals, fc_mean = m.predict(X), m.predict(Xf)
        pr = m.get_prediction(Xf).summary_frame(alpha=0.2)  # 80% → obs_ci = intervallo di previsione
        fc_mean = pr["mean"].to_numpy()
        fc_lo, fc_hi = pr["obs_ci_lower"].to_numpy(), pr["obs_ci_upper"].to_numpy()
        r2 = float(m.rsquared)
    return {"hist_years": [int(v) for v in s.index], "hist_vals": [float(v) for v in y],
            "fit_vals": [float(v) for v in fit_vals],
            "fut_years": fut, "fc_mean": [float(v) for v in fc_mean],
            "fc_lo": [float(v) for v in fc_lo], "fc_hi": [float(v) for v in fc_hi],
            "slope": float(m.params[1]), "r2": float(r2), "robust": bool(robust)}


# ── TIER 2 · partial pooling tra regioni + riconciliazione gerarchica ─────────
#   Idea (dallo studio): le serie annuali per singola regione sono corte/rumorose; stimare
#   il trend in isolamento è instabile. Il PARTIAL POOLING (modello a effetti casuali, stima
#   empirical-Bayes) restringe la pendenza di ogni regione verso la media nazionale, di più
#   dove la stima della singola regione è incerta. La RICONCILIAZIONE rende coerenti i livelli
#   della gerarchia Italia↔regioni (il totale = somma delle regioni, per costruzione bottom-up).

def annual_slope_se(years, vals) -> tuple[float, float] | None:
    """Pendenza annua (OLS) della serie e suo errore standard. None se <3 punti validi."""
    import numpy as np
    import statsmodels.api as sm
    x = np.asarray(years, float)
    y = np.asarray(vals, float)
    ok = np.isfinite(x) & np.isfinite(y)
    if ok.sum() < 3:
        return None
    m = sm.OLS(y[ok], sm.add_constant(x[ok])).fit()
    return float(m.params[1]), float(m.bse[1])


def partial_pool_slopes(slopes: dict) -> dict:
    """Empirical-Bayes / James-Stein su pendenze regionali.
    `slopes`: {code: (pendenza, errore_standard)}. Restituisce, per regione,
    pendenza grezza, pendenza ristretta (shrunk), peso w (1=fido la regione, 0=tutta media),
    più la media nazionale `mu` e la varianza tra regioni `tau2` (metodo dei momenti)."""
    import numpy as np
    codes = [c for c, v in slopes.items() if v is not None and np.isfinite(v[0]) and np.isfinite(v[1])]
    if len(codes) < 3:
        return {"by_code": {}, "mu": float("nan"), "tau2": float("nan"), "n": len(codes)}
    b = np.array([slopes[c][0] for c in codes], float)
    s2 = np.array([slopes[c][1] for c in codes], float) ** 2
    mu = float(np.mean(b))
    tau2 = max(0.0, float(np.var(b, ddof=1) - np.mean(s2)))  # varianza "vera" tra regioni
    out = {}
    for c, bi, s2i in zip(codes, b, s2):
        w = tau2 / (tau2 + s2i) if (tau2 + s2i) > 0 else 0.0
        out[c] = {"raw": float(bi), "shrunk": float(mu + w * (bi - mu)),
                  "weight": float(w), "se": float(np.sqrt(s2i))}
    return {"by_code": out, "mu": mu, "tau2": tau2, "n": len(codes)}


def reconcile_bottom_up(region_fc: dict, national_fc: float | None = None) -> dict:
    """Riconciliazione gerarchica BOTTOM-UP: il totale coerente è la somma delle regioni
    (coerente per costruzione). Se è data una previsione nazionale indipendente
    (`national_fc`, top-down), riporta lo scostamento % tra le due viste (incoerenza)."""
    import numpy as np
    vals = {c: float(v) for c, v in region_fc.items() if v is not None and np.isfinite(v)}
    bottom_up = float(sum(vals.values()))
    gap = ((bottom_up / national_fc - 1) * 100
           if national_fc and np.isfinite(national_fc) and national_fc != 0 else float("nan"))
    return {"bottom_up_total": bottom_up, "national_top_down": national_fc,
            "gap_pct": gap, "n_regioni": len(vals), "per_regione": vals}


@st.cache_data(show_spinner=False)
def load_tier2_artifact() -> dict | None:
    """Legge l'artefatto PRECALCOLATO del Tier 2 (`.cache/tier2_pooled_trends.json`):
    trend pooled per regione + riconciliazione. Lettura istantanea (lo genera lo script
    offline `tier2_structural.py`). None se assente: l'app funziona comunque."""
    import json
    import os
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".cache", "tier2_pooled_trends.json")
    if not os.path.exists(p):
        return None
    try:
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    except Exception:  # noqa: BLE001
        return None


def pooled_trend_for(code: str, key: str) -> dict | None:
    """Info di partial pooling per (regione, variabile) dall'artefatto: crescita grezza vs
    stabilizzata (%/anno), peso della regione, media nazionale; + riconciliazione se `presenze_str`.
    None se la variabile non è nell'artefatto o la regione manca."""
    art = load_tier2_artifact()
    if not art:
        return None
    pt = (art.get("pooled_trends") or {}).get(key)
    bc = (pt.get("by_code") or {}).get(code) if pt else None
    if not bc:
        return None
    return {"raw": bc.get("raw"), "shrunk": bc.get("shrunk"), "weight": bc.get("weight"),
            "mu": pt.get("mu"), "window": art.get("window"),
            "recon": art.get("reconciliation_presenze_str") if key == "presenze_str" else None}


def chart_projection(proj: dict, label: str, unit: str = "") -> go.Figure:
    """Storico + trend fittato + proiezione con banda di previsione 80% (unità assolute)."""
    hx, hy = proj["hist_years"], proj["hist_vals"]
    fx, fm = proj["fut_years"], proj["fc_mean"]
    lo, hi = proj["fc_lo"], proj["fc_hi"]
    bx = [hx[-1]] + fx                       # aggancia la banda all'ultimo punto storico
    bl = [hy[-1]] + [max(0.0, v) for v in lo]
    bh = [hy[-1]] + hi
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=bx + bx[::-1], y=bh + bl[::-1], fill="toself",
                             fillcolor="rgba(245,158,11,.15)", line=dict(width=0),
                             name="intervallo 80%", hoverinfo="skip"))
    fig.add_trace(go.Scatter(x=hx, y=proj["fit_vals"], name="trend", mode="lines",
                             line=dict(color="#94a3b8", width=1, dash="dot")))
    fig.add_trace(go.Scatter(x=hx, y=hy, name="storico", mode="lines+markers",
                             line=dict(color="#0e6b70", width=2)))
    fig.add_trace(go.Scatter(x=[hx[-1]] + fx, y=[hy[-1]] + fm, name="proiezione",
                             mode="lines+markers", line=dict(color="#f59e0b", width=2, dash="dash")))
    fig.update_yaxes(title=label + (f" ({unit})" if unit and unit != "n" else ""))
    fig.update_xaxes(title="anno", dtick=1)
    return _layout(fig, h=380)


# ── Proiezione STAGIONALE (mensile) — riusa la metodologia del motore aggregato:
#    value ~ C(mese) + trend + dummy COVID. Disponibile per le serie mensili (presenze ISTAT).
MONTHLY_VARS = {"presenze_tot": "totale", "presenze_str": "stranieri"}


def region_monthly_series(code: str, key: str) -> pd.DataFrame | None:
    """Serie MENSILE (date, value) per le variabili con dato mensile (presenze ISTAT)."""
    col = MONTHLY_VARS.get(key)
    if not col:
        return None
    try:
        df = region_overview(code)["presenze"]
    except Exception:  # noqa: BLE001
        return None
    if col not in df:
        return None
    out = df[["date", col]].rename(columns={col: "value"}).dropna().sort_values("date")
    return out if not out.empty else None


def project_seasonal(series: pd.DataFrame, horizon_years: int = 2) -> dict | None:
    """Proiezione mensile stagionale: value ~ C(mese) + trend + dummy COVID (2020-03→2021-12),
    intervallo di previsione 80%. Ritorna serie mensile (storico+forecast) e roll-up annuale
    (somma 12 mesi) per gli anni futuri completi. Stessa metodologia del motore aggregato."""
    import numpy as np
    import statsmodels.formula.api as smf
    if series is None or len(series) < 24:  # servono ~2 anni di mesi
        return None
    d = series.dropna().sort_values("date").reset_index(drop=True).copy()
    d["month"] = d["date"].dt.month
    d["t"] = np.arange(len(d))
    d["covid"] = ((d["date"] >= pd.Timestamp("2020-03-01")) &
                  (d["date"] <= pd.Timestamp("2021-12-01"))).astype(int)
    model = smf.ols("value ~ C(month) + t + covid", data=d).fit()
    n = int(horizon_years) * 12
    last_t = int(d["t"].iloc[-1])
    fut_dates = pd.date_range(d["date"].max() + pd.DateOffset(months=1), periods=n, freq="MS")
    fut = pd.DataFrame({"date": fut_dates, "month": fut_dates.month,
                        "t": last_t + np.arange(1, n + 1), "covid": 0})
    pr = model.get_prediction(fut).summary_frame(alpha=0.2)
    fut["mean"] = pr["mean"].values
    fut["lo"] = pr["obs_ci_lower"].clip(lower=0).values
    fut["hi"] = pr["obs_ci_upper"].values
    ha = d.assign(anno=d["date"].dt.year).groupby("anno")["value"].agg(["sum", "count"])
    hist_ann = {int(y): float(r["sum"]) for y, r in ha.iterrows() if r["count"] >= 12}
    fut["anno"] = fut["date"].dt.year
    fa = fut.groupby("anno").agg(mean=("mean", "sum"), lo=("lo", "sum"),
                                 hi=("hi", "sum"), n=("mean", "count"))
    fc_ann = {int(y): (float(r["mean"]), float(r["lo"]), float(r["hi"]))
              for y, r in fa.iterrows() if r["n"] >= 12}
    return {"hist": d[["date", "value"]], "fut": fut[["date", "mean", "lo", "hi"]],
            "hist_ann": hist_ann, "fc_ann": fc_ann, "r2": float(model.rsquared)}


# ── Sfidanti statistici dentro la barriera di onestà ──────────────────────────
#   Sei motori competono con il modello stagionale OLS: ETS (Holt-Winters), SARIMAX,
#   stato latente (UCM), Theta, STL+ARIMA. Si tiene quello con l'errore più basso in un
#   backtest ROLLING-ORIGIN (più finestre mobili da 12 mesi → scelta affidabile), e si
#   riporta lo skill vs la naive stagionale. Tutto in statsmodels — nessuna nuova dipendenza.
#   Ogni candidato è in try/except: se uno non converge, esce dalla gara; l'OLS stagionale
#   resta sempre come fallback garantito (il metodo già validato).

def _seasonal_design(series: pd.DataFrame) -> pd.DataFrame:
    """Pannello mensile con month/t/covid (stessa specifica di project_seasonal)."""
    import numpy as np
    d = series.dropna().sort_values("date").reset_index(drop=True).copy()
    d["month"] = d["date"].dt.month
    d["t"] = np.arange(len(d))
    d["covid"] = ((d["date"] >= pd.Timestamp("2020-03-01")) &
                  (d["date"] <= pd.Timestamp("2021-12-01"))).astype(int)
    return d


def _pseudo_r2(y, fitted) -> float:
    import numpy as np
    y = np.asarray(y, float); fitted = np.asarray(fitted, float)
    ss_tot = float(np.sum((y - y.mean()) ** 2))
    ss_res = float(np.sum((y - fitted) ** 2))
    return float(1 - ss_res / ss_tot) if ss_tot else float("nan")


def _fc_seas_ols(train: pd.DataFrame, fut: pd.DataFrame):
    """Modello attuale: value ~ C(mese) + trend + dummy COVID, intervallo 80%."""
    import statsmodels.formula.api as smf
    m = smf.ols("value ~ C(month) + t + covid", data=train).fit()
    pr = m.get_prediction(fut).summary_frame(alpha=0.2)
    return (pr["mean"].to_numpy(), pr["obs_ci_lower"].to_numpy(),
            pr["obs_ci_upper"].to_numpy(), float(m.rsquared))


def _fc_seas_ets(train: pd.DataFrame, fut: pd.DataFrame):
    """ETS / Holt-Winters additivo con stagionalità 12 e trend smorzato."""
    import numpy as np
    from statsmodels.tsa.exponential_smoothing.ets import ETSModel
    ser = pd.Series(train["value"].to_numpy(float),
                    index=pd.PeriodIndex(train["date"], freq="M"))
    m = ETSModel(ser, error="add", trend="add", seasonal="add",
                 seasonal_periods=12, damped_trend=True).fit(disp=False)
    nf = len(fut)
    sf = m.get_prediction(start=len(ser), end=len(ser) + nf - 1).summary_frame(alpha=0.2)
    return (sf["mean"].to_numpy(), sf["pi_lower"].to_numpy(),
            sf["pi_upper"].to_numpy(), _pseudo_r2(ser.to_numpy(), m.fittedvalues))


def _fc_seas_sarimax(train: pd.DataFrame, fut: pd.DataFrame):
    """SARIMAX (1,0,0)(0,1,1)[12] con dummy COVID come regressore esogeno."""
    import numpy as np
    from statsmodels.tsa.statespace.sarimax import SARIMAX
    y = train["value"].to_numpy(float)
    ex = train[["covid"]].to_numpy(float)
    exf = fut[["covid"]].to_numpy(float)
    m = SARIMAX(y, exog=ex, order=(1, 0, 0), seasonal_order=(0, 1, 1, 12),
                enforce_stationarity=False, enforce_invertibility=False).fit(disp=False)
    pr = m.get_prediction(start=len(y), end=len(y) + len(fut) - 1, exog=exf)
    ci = np.asarray(pr.conf_int(alpha=0.2))
    return (np.asarray(pr.predicted_mean), ci[:, 0], ci[:, 1],
            _pseudo_r2(y, m.fittedvalues))


def _fc_seas_ucm(train: pd.DataFrame, fut: pd.DataFrame):
    """TIER 2 · modello a STATO LATENTE (UnobservedComponents): livello e trend stocastici
    + stagionalità stocastica 12, con dummy COVID esogena. È un modello strutturale che
    scompone la serie in componenti interpretabili (livello/trend/stagione) e dà incertezza
    nativa; gestisce bene serie con regime che cambia lentamente."""
    import numpy as np
    import statsmodels.api as sm
    y = train["value"].to_numpy(float)
    ex = train[["covid"]].to_numpy(float)
    exf = fut[["covid"]].to_numpy(float)
    m = sm.tsa.UnobservedComponents(y, level="local linear trend", seasonal=12,
                                    exog=ex).fit(disp=False)
    pr = m.get_prediction(start=len(y), end=len(y) + len(fut) - 1, exog=exf)
    ci = np.asarray(pr.conf_int(alpha=0.2))
    return (np.asarray(pr.predicted_mean), ci[:, 0], ci[:, 1],
            _pseudo_r2(y, m.fittedvalues))


def _fc_seas_theta(train: pd.DataFrame, fut: pd.DataFrame):
    """Metodo THETA (Assimakopoulos-Nikolopoulos): semplice e robusto, spesso vincente su
    serie stagionali quando ETS/ARIMA faticano. Intervalli nativi. Niente esogene."""
    import numpy as np
    from statsmodels.tsa.forecasting.theta import ThetaModel
    ser = pd.Series(train["value"].to_numpy(float),
                    index=pd.PeriodIndex(train["date"], freq="M"))
    m = ThetaModel(ser, period=12).fit()
    nf = len(fut)
    fc = np.asarray(m.forecast(nf), float)
    pi = m.prediction_intervals(nf, alpha=0.2)
    return fc, pi["lower"].to_numpy(), pi["upper"].to_numpy(), float("nan")


def _fc_seas_stl(train: pd.DataFrame, fut: pd.DataFrame):
    """STL + ARIMA: scompone stagionalità/trend con STL e modella il resto con un ARIMA.
    Utile quando la stagionalità è marcata ma il resto ha una dinamica propria."""
    import numpy as np
    from statsmodels.tsa.arima.model import ARIMA
    from statsmodels.tsa.forecasting.stl import STLForecast
    ser = pd.Series(train["value"].to_numpy(float),
                    index=pd.PeriodIndex(train["date"], freq="M"))
    m = STLForecast(ser, ARIMA, model_kwargs=dict(order=(1, 1, 1)), period=12).fit()
    nf = len(fut)
    sf = m.get_prediction(len(ser), len(ser) + nf - 1).summary_frame(alpha=0.2)
    return (sf["mean"].to_numpy(), sf["mean_ci_lower"].to_numpy(),
            sf["mean_ci_upper"].to_numpy(), float("nan"))


_SEASONAL_METHODS = {"OLS stagionale": _fc_seas_ols,
                     "ETS (Holt-Winters)": _fc_seas_ets,
                     "SARIMAX": _fc_seas_sarimax,
                     "Stato latente (UCM)": _fc_seas_ucm,
                     "Theta": _fc_seas_theta,
                     "STL + ARIMA": _fc_seas_stl}


def _backtest_seasonal(d: pd.DataFrame, horizon: int = 12, folds: int = 3,
                       min_train: int = 24) -> dict:
    """Backtest ROLLING-ORIGIN: per fino a `folds` origini mobili, addestra fino all'origine
    e prevede i `horizon` mesi successivi; media l'errore (MAE) di ogni motore sulle finestre
    e lo confronta con la naive stagionale. Riporta anche quante finestre ha vinto ciascuno.
    Più finestre = scelta del motore più affidabile (non frutto di una sola finestra fortunata)."""
    import numpy as np
    n = len(d)
    origins = [n - k * horizon for k in range(folds, 0, -1) if n - k * horizon >= min_train]
    if not origins:  # storia corta: ripiega su una singola finestra
        origins = [n - horizon] if n >= min_train + horizon else []
    if not origins:
        return {}
    full = d.set_index("date")["value"]
    agg = {name: [] for name in _SEASONAL_METHODS}
    wins = {name: 0 for name in _SEASONAL_METHODS}
    naive_maes = []
    for cut in origins:
        train, test = d.iloc[:cut], d.iloc[cut:cut + horizon]
        if test.empty:
            continue
        actual = test["value"].to_numpy(float)
        naive = np.array([full.get(dt - pd.DateOffset(years=1), np.nan) for dt in test["date"]])
        mask = ~np.isnan(naive)
        naive_maes.append(float(np.mean(np.abs(actual[mask] - naive[mask]))) if mask.any() else np.nan)
        fold_mae = {}
        for name, fn in _SEASONAL_METHODS.items():
            try:
                mean = np.asarray(fn(train, test)[0], float)[:len(actual)]
                mae = float(np.mean(np.abs(actual - mean)))
            except Exception:  # noqa: BLE001 — motore fuori gara in questa finestra
                mae = float("nan")
            fold_mae[name] = mae
            if np.isfinite(mae):
                agg[name].append(mae)
        finite = {k: v for k, v in fold_mae.items() if np.isfinite(v)}
        if finite:
            wins[min(finite, key=finite.get)] += 1
    naive_mae = float(np.nanmean(naive_maes)) if naive_maes else float("nan")
    out = {"_naive_mae": naive_mae, "_folds": len(origins)}
    for name in _SEASONAL_METHODS:
        if agg[name]:
            mae = float(np.mean(agg[name]))
            skill = (1 - mae / naive_mae) * 100 if (naive_mae and np.isfinite(naive_mae)) else float("nan")
            out[name] = {"mae": mae, "skill": skill,
                         "beats_naive": bool(np.isfinite(skill) and skill > 0),
                         "wins": wins[name], "n_folds": len(agg[name])}
        else:
            out[name] = {"mae": float("nan"), "skill": float("nan"), "beats_naive": False,
                         "wins": 0, "n_folds": 0}
    return out


def project_seasonal_best(series: pd.DataFrame, horizon_years: int = 2,
                          folds: int = 3) -> dict | None:
    """Proiezione mensile che SCEGLIE il modello tra 6 motori (OLS · ETS · SARIMAX · stato
    latente · Theta · STL+ARIMA), selezionato per minor errore in un backtest ROLLING-ORIGIN
    (più finestre mobili) e confrontato con la naive stagionale. Stesso contratto di output di
    project_seasonal (hist/fut/hist_ann/fc_ann/r2) + i metadati della gara (method/beats_naive/
    backtest, con finestre vinte), così grafici e UI non cambiano."""
    import numpy as np
    if series is None or len(series.dropna()) < 24:
        return None
    d = _seasonal_design(series)
    n = int(horizon_years) * 12
    last_t = int(d["t"].iloc[-1])
    fut_dates = pd.date_range(d["date"].max() + pd.DateOffset(months=1), periods=n, freq="MS")
    fut = pd.DataFrame({"date": fut_dates, "month": fut_dates.month,
                        "t": last_t + np.arange(1, n + 1), "covid": 0})

    bt = _backtest_seasonal(d, horizon=12, folds=folds)
    ranked = sorted(((k, bt[k]["mae"]) for k in _SEASONAL_METHODS
                     if k in bt and np.isfinite(bt[k].get("mae", np.nan))),
                    key=lambda kv: kv[1])
    winner = ranked[0][0] if ranked else "OLS stagionale"
    try:
        mean, lo, hi, r2 = _SEASONAL_METHODS[winner](d, fut)
    except Exception:  # noqa: BLE001 — fallback inossidabile all'OLS già validato
        winner = "OLS stagionale"
        mean, lo, hi, r2 = _fc_seas_ols(d, fut)

    fut["mean"] = np.asarray(mean, float)
    fut["lo"] = np.clip(np.asarray(lo, float), 0, None)
    fut["hi"] = np.asarray(hi, float)

    ha = d.assign(anno=d["date"].dt.year).groupby("anno")["value"].agg(["sum", "count"])
    hist_ann = {int(y): float(r["sum"]) for y, r in ha.iterrows() if r["count"] >= 12}
    fa = fut.assign(anno=fut["date"].dt.year).groupby("anno").agg(
        mean=("mean", "sum"), lo=("lo", "sum"), hi=("hi", "sum"), n=("mean", "count"))
    fc_ann = {int(y): (float(r["mean"]), float(r["lo"]), float(r["hi"]))
              for y, r in fa.iterrows() if r["n"] >= 12}
    return {"hist": d[["date", "value"]], "fut": fut[["date", "mean", "lo", "hi"]],
            "hist_ann": hist_ann, "fc_ann": fc_ann, "r2": float(r2),
            "method": winner, "beats_naive": bool(bt.get(winner, {}).get("beats_naive", False)),
            "backtest": bt}


def chart_projection_monthly(proj: dict, label: str, unit: str = "") -> go.Figure:
    """Storico mensile + proiezione stagionale con banda di previsione 80%."""
    h, f = proj["hist"], proj["fut"]
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=list(f["date"]) + list(f["date"])[::-1],
                             y=list(f["hi"]) + list(f["lo"])[::-1], fill="toself",
                             fillcolor="rgba(245,158,11,.15)", line=dict(width=0),
                             name="intervallo 80%", hoverinfo="skip"))
    fig.add_trace(go.Scatter(x=h["date"], y=h["value"], name="storico", mode="lines",
                             line=dict(color="#0e6b70", width=2)))
    mx = [h["date"].iloc[-1]] + list(f["date"])   # aggancia all'ultimo punto storico
    my = [h["value"].iloc[-1]] + list(f["mean"])
    fig.add_trace(go.Scatter(x=mx, y=my, name="proiezione stagionale", mode="lines",
                             line=dict(color="#f59e0b", width=2, dash="dash")))
    fig.update_yaxes(title=label + (f" ({unit})" if unit and unit != "n" else "") + " · mese")
    return _layout(fig, h=380)


def regions_spend_ranking() -> list[dict]:
    """Classifica delle regioni per spesa turistica straniera 2024 (Banca d'Italia).
    code (NUTS2) · regione · spesa_M · rank. Multi-regione by construction."""
    ext = bdi_extended()
    spese = (ext or {}).get("regioni_2024", {})
    if not spese:
        return []
    rows = []
    for code, info in RG.REGIONS.items():
        sp = spese.get(info["bdi"])
        if sp is not None:
            rows.append({"code": code, "regione": info["nome"], "spesa_M": float(sp)})
    rows.sort(key=lambda r: r["spesa_M"], reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return rows


@st.cache_data(show_spinner=False)
def bdi_region_years() -> list[int]:
    """Anni con i 4 trimestri completi nella serie regionale BdI (per i selettori)."""
    df = bdi_region_long()
    if df is None or df.empty:
        return []
    g = df.assign(anno=df["date"].dt.year).groupby("anno")["date"].nunique()
    return sorted(int(y) for y, n in g.items() if n >= 4)


@st.cache_data(show_spinner=False)
def regions_spend_ranking_year(year: int) -> list[dict]:
    """Come regions_spend_ranking() ma per l'ANNO indicato (spesa straniera M€, BdI).
    Ricavata da bdi_region_long(); per il 2024 coincide col dato di bdi_extended."""
    df = bdi_region_long()
    if df is None or df.empty:
        return []
    d = df[df["date"].dt.year == year]
    if d.empty:
        return []
    # Spesa annuale per regione visitata. Trentino: ITD1/ITD2 hanno lo stesso valore,
    # quindi mappando per nome BdI non si duplica (allineato a regions_spend_ranking).
    by_name: dict[str, float] = {}
    for code, val in d.groupby("code")["spesa"].sum().items():
        info = RG.REGIONS.get(code)
        if info:
            by_name[info["bdi"]] = float(val)
    rows = []
    for code, info in RG.REGIONS.items():
        sp = by_name.get(info["bdi"])
        if sp is not None:
            rows.append({"code": code, "regione": info["nome"], "spesa_M": sp})
    rows.sort(key=lambda r: r["spesa_M"], reverse=True)
    for i, r in enumerate(rows):
        r["rank"] = i + 1
    return rows


def region_spend(code: str):
    """(spesa_M, rank, totale_regioni) per la regione richiesta, o None.
    Per la vista nazionale: (totale_Italia, None, n_regioni) — rank=None = «Italia»."""
    rk = regions_spend_ranking()
    if RG.is_national(code):
        return (sum(r["spesa_M"] for r in rk), None, len(rk)) if rk else None
    bdi = RG.region(code)["bdi"]
    hit = next((r for r in rk if RG.REGIONS[r["code"]]["bdi"] == bdi), None)
    return (hit["spesa_M"], hit["rank"], len(rk)) if hit else None


@st.cache_data(show_spinner=False)
def _italy_regions_geojson():
    return json.load(open("assets/italy_regions.geojson", encoding="utf-8"))


def chart_italy_map(highlight: str | None = None, year: int | None = None) -> go.Figure:
    """Mappa d'Italia delle regioni, colorata per spesa straniera (BdI) dell'anno indicato
    (default: dato 2024 da bdi_extended). La regione `highlight` ha il bordo rosso.
    Cliccabile come selettore (gestito in app)."""
    gj = _italy_regions_geojson()
    rk = regions_spend_ranking_year(year) if year is not None else regions_spend_ranking()
    spend = {RG.REGIONS[r["code"]]["bdi"]: r["spesa_M"] for r in rk}
    names, z, codes = [], [], []
    for f in gj["features"]:
        nm = f["properties"]["reg_name"]
        code = RG.code_for_geo(nm)
        names.append(nm)
        codes.append(code or "")
        z.append(spend.get(RG.region(code)["bdi"]) if code else None)
    fig = go.Figure(go.Choropleth(
        geojson=gj, featureidkey="properties.reg_name", locations=names, z=z, name="spesa",
        colorscale="Teal", marker_line_color="white", marker_line_width=0.6, customdata=codes,
        selected=dict(marker=dict(opacity=1.0)), unselected=dict(marker=dict(opacity=1.0)),
        colorbar=dict(title=dict(text=f"spesa straniera<br>{year or 2024} (M€)", side="right")),
        hovertemplate="<b>%{location}</b><br>spesa %{z:,.0f} M€<extra></extra>"))
    hn = RG.GEO_NAME.get(highlight) if highlight else None
    if hn:
        fig.add_trace(go.Choropleth(
            geojson=gj, featureidkey="properties.reg_name", locations=[hn], z=[0], showscale=False,
            colorscale=[[0, "rgba(0,0,0,0)"], [1, "rgba(0,0,0,0)"]],
            marker_line_color="#dc2626", marker_line_width=2.5, hoverinfo="skip"))
    fig.update_geos(fitbounds="locations", visible=False, projection_type="mercator",
                    bgcolor="rgba(0,0,0,0)")
    fig.update_layout(height=620, margin=dict(l=0, r=0, t=10, b=0), paper_bgcolor="white",
                      title_text="", font=dict(family=FONT_STACK))
    return fig


def chart_regions_ranking(highlight: str | None = None, year: int | None = None) -> go.Figure:
    rk = regions_spend_ranking_year(year) if year is not None else regions_spend_ranking()
    if not rk:
        return _layout(go.Figure(), h=300)
    rk = sorted(rk, key=lambda r: r["spesa_M"])
    hi_bdi = RG.region(highlight)["bdi"] if highlight else None
    colors = ["#f59e0b" if RG.REGIONS[r["code"]]["bdi"] == hi_bdi else "#0e6b70" for r in rk]
    fig = go.Figure(go.Bar(x=[r["spesa_M"] for r in rk], y=[r["regione"] for r in rk],
                           orientation="h", marker_color=colors,
                           hovertemplate="<b>%{y}</b><br>%{x:,.0f} M€<extra></extra>"))
    fig.update_xaxes(title=f"spesa turisti stranieri {year or 2024} (M€)")
    return _layout(fig, h=560)


def chart_occupancy_season(panel) -> go.Figure:
    d = panel.copy(); d["m"] = d["date"].dt.month
    prof = d.groupby("m")["occ"].mean().reindex(range(1, 13))
    fig = go.Figure(go.Bar(x=MONTHS_IT, y=prof.values, marker_color="#0e6b70"))
    fig.update_yaxes(title="occupazione media", ticksuffix="%")
    return _layout(fig, h=300)


# ════════════════════════════════════════════════════════════════════════════
# VISTA OPERATORI — DEMO (dati SIMULATI, etichettati): parità col POC su feature
# che richiedono sorgenti non ancora collegate (Registro ricettivo, OTA, GA4).
# ════════════════════════════════════════════════════════════════════════════
def demo_banner():
    st.markdown("""<div style="background:#fffbeb;border:1px solid #fcd34d;border-radius:12px;
        padding:12px 16px;margin-bottom:14px;color:#92400e">
        <b>⚠️ DEMO · dati simulati</b> a scopo illustrativo — anticipa funzionalità che richiedono
        sorgenti non ancora collegate (Registro ricettivo, booking/OTA, analytics del portale).
        I numeri NON sono reali.</div>""", unsafe_allow_html=True)


def demo_operators(provincia: str = "Tutte") -> dict:
    """Dati SIMULATI ma deterministici (seed dalla provincia) per la vista operatori demo."""
    import random
    rng = random.Random(sum(ord(c) for c in provincia) + 7)
    base = {"Organico": 32, "OTA (Booking/Expedia)": 27, "Social": 17,
            "Diretto": 13, "Email": 8, "Referral": 3}
    ch = {k: max(1, v + rng.randint(-4, 4)) for k, v in base.items()}
    tot = sum(ch.values()); ch = {k: round(v / tot * 100, 1) for k, v in ch.items()}
    ric = rng.randint(80, 140) * 1000
    sch = int(ric * rng.uniform(0.40, 0.50))
    clk = int(sch * rng.uniform(0.22, 0.30))
    pren = int(clk * rng.uniform(0.24, 0.32))
    base_w = [0.72, 0.70, 0.76, 0.86, 1.02, 1.22, 1.12]
    scale = rng.randint(900, 1600)
    weekly = {g: int(scale * w * rng.uniform(0.92, 1.08))
              for g, w in zip(["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"], base_w)}
    rev = {"5★": rng.randint(120, 320), "4★": rng.randint(80, 200), "3★": rng.randint(20, 70),
           "2★": rng.randint(5, 25), "1★": rng.randint(2, 15)}
    queries = ["Costa dei Trabocchi", "Gran Sasso trekking", "Rocca Calascio", "borghi Abruzzo",
               "Parco della Majella", "spiagge Abruzzo", "agriturismo Abruzzo", "Lago di Scanno"]
    rng.shuffle(queries)
    return {"occupazione": rng.randint(48, 78), "adr": rng.randint(62, 118),
            "qualita": round(rng.uniform(3.8, 4.6), 1), "conversione": round(rng.uniform(1.8, 4.2), 1),
            "canali": ch, "funnel": {"Ricerche": ric, "Schede viste": sch,
            'Click "prenota"': clk, "Prenotazioni": pren},
            "weekly": weekly, "recensioni": rev, "queries": queries[:6]}


def chart_demo_channels(ch: dict) -> go.Figure:
    return _layout(go.Figure(go.Pie(labels=list(ch), values=list(ch.values()), hole=0.55)), h=320)


def chart_demo_funnel(f: dict) -> go.Figure:
    fig = go.Figure(go.Funnel(y=list(f), x=list(f.values()), textinfo="value+percent initial",
                              marker={"color": ["#0e6b70", "#0e6b70", "#38bdf8", "#f59e0b"]}))
    return _layout(fig, h=320)


def chart_demo_weekly(w: dict) -> go.Figure:
    fig = go.Figure(go.Bar(x=list(w), y=list(w.values()), marker_color="#0e6b70"))
    fig.update_yaxes(title="presenze previste")
    return _layout(fig, h=290)


def chart_demo_reviews(r: dict) -> go.Figure:
    keys = list(r)[::-1]  # 1★..5★ dal basso
    fig = go.Figure(go.Bar(x=[r[k] for k in keys], y=keys, orientation="h",
                           marker_color=["#dc2626", "#f59e0b", "#eab308", "#84cc16", "#16a34a"]))
    fig.update_xaxes(title="recensioni")
    return _layout(fig, h=260)


# ════════════════════════════════════════════════════════════════════════════
# SLICING STORICO — filtro periodo + mappa di copertura delle serie
# ════════════════════════════════════════════════════════════════════════════
def filter_years(df, yr_range, date_col: str = "date"):
    """Ritaglia un DataFrame all'intervallo di anni [a, b]."""
    if df is None or date_col not in df.columns:
        return df
    a, b = yr_range
    return df[(df[date_col].dt.year >= a) & (df[date_col].dt.year <= b)]


def series_coverage() -> list[dict]:
    """Inizio-fine di ogni serie storica in cache (coperture diverse)."""
    def rng(path):
        if not os.path.exists(path):
            return None
        d = pd.read_csv(path)
        if "date" in d.columns:
            t = pd.to_datetime(d["date"]); return t.min(), t.max()
        if "anno" in d.columns:
            return pd.Timestamp(f"{int(d['anno'].min())}-01-01"), pd.Timestamp(f"{int(d['anno'].max())}-12-31")
        return None
    items = [
        ("Presenze straniere (ISTAT)", ".cache/istat_presenze_straniere_abruzzo.csv", "mensile"),
        ("Presenze province (ISTAT)", ".cache/istat_ITF13_NI_ALL_WORLD.csv", "mensile"),
        ("Capacità posti letto (ISTAT)", ".cache/istat_capacity_letti_ITF1.csv", "annuale"),
        ("Google Trends (per mercato)", ".cache/trends_abruzzo_DE.csv", "mensile"),
        ("Wikipedia (per lingua)", ".cache/wiki_de.csv", "mensile"),
    ]
    cov = []
    for name, path, freq in items:
        r = rng(path)
        if r:
            cov.append({"serie": name, "start": r[0], "end": r[1], "freq": freq})
    cov.append({"serie": "Cambio ECB (live)", "start": pd.Timestamp("2019-01-01"),
                "end": pd.Timestamp.today().normalize(), "freq": "mensile"})
    if os.path.exists(".cache/bdi_turismo_ts.xlsx"):
        cov.append({"serie": "Spesa BdI (turismo internaz.)", "start": pd.Timestamp("1997-01-01"),
                    "end": pd.Timestamp("2025-12-31"), "freq": "trimestrale"})
    eu = ".cache/eurostat_pescara_flights.csv"
    if os.path.exists(eu):
        try:
            y = int(pd.read_csv(eu)["anno"].iloc[0])
        except Exception:  # noqa: BLE001
            y = 2024
        cov.append({"serie": "Voli Pescara (Eurostat)", "start": pd.Timestamp(f"{y}-01-01"),
                    "end": pd.Timestamp(f"{y}-12-31"), "freq": "annuale"})
    if os.path.exists(".cache/istat_estero_regione_prov_annuale.csv"):
        cov.append({"serie": "Esteri per paese×regione (ISTAT _9)", "start": pd.Timestamp("2008-01-01"),
                    "end": pd.Timestamp("2024-12-31"), "freq": "annuale"})
    r = rng(".cache/astat_bolzano_flussi_mensili.csv")
    if r:
        cov.append({"serie": "Alto Adige flussi (ASTAT)", "start": r[0], "end": r[1], "freq": "mensile"})
    r = rng(".cache/lombardia_flussi_provincia_mese.csv")
    if r:
        cov.append({"serie": "Lombardia flussi×mercato (Open Data)", "start": r[0], "end": r[1], "freq": "mensile"})
    if os.path.exists(".cache/toscana_movimento_comune_anno.csv"):
        try:
            aa = pd.read_csv(".cache/toscana_movimento_comune_anno.csv", usecols=["anno"])["anno"]
            cov.append({"serie": "Toscana movimento×comune (Open Data)",
                        "start": pd.Timestamp(f"{int(aa.min())}-01-01"),
                        "end": pd.Timestamp(f"{int(aa.max())}-12-31"), "freq": "annuale"})
        except Exception:  # noqa: BLE001
            pass
    if os.path.exists(_SARD_CM_PATH):
        try:
            aa = pd.read_csv(_SARD_CM_PATH, usecols=["anno"])["anno"]
            cov.append({"serie": "Sardegna movimento×comune×mercato (Osservatorio)",
                        "start": pd.Timestamp(f"{int(aa.min())}-01-01"),
                        "end": pd.Timestamp(f"{int(aa.max())}-12-31"), "freq": "mensile"})
        except Exception:  # noqa: BLE001
            pass
    if os.path.exists(_TURNOT_PATH):
        try:
            aa = pd.read_csv(_TURNOT_PATH, usecols=["anno"])["anno"]
            cov.append({"serie": "Domanda residenti per scopo (ISTAT V&V)",
                        "start": pd.Timestamp(f"{int(aa.min())}-01-01"),
                        "end": pd.Timestamp(f"{int(aa.max())}-12-31"), "freq": "annuale"})
        except Exception:  # noqa: BLE001
            pass
    return cov


def chart_coverage(cov: list[dict]) -> go.Figure:
    import plotly.express as px
    df = pd.DataFrame(cov)
    fig = px.timeline(df, x_start="start", x_end="end", y="serie", color="freq",
                      color_discrete_sequence=["#0e6b70", "#f59e0b", "#16a34a"])
    fig.update_yaxes(autorange="reversed", title=None)
    # Linee verticali di riferimento: maggiori per ANNO (solide), minori per SEMESTRE (punteggiate)
    fig.update_xaxes(
        showgrid=True, gridcolor="#94a3b8", gridwidth=1, dtick="M12",
        tickformat="%Y", ticklabelmode="period", ticks="outside",
        minor=dict(dtick="M6", showgrid=True, gridcolor="#e2e8f0", griddash="dot", gridwidth=1),
    )
    fig.update_layout(height=300, margin=dict(l=10, r=10, t=10, b=10), paper_bgcolor="white",
                      plot_bgcolor="white", title_text="", legend_title="frequenza", font=dict(family=FONT_STACK))
    return fig


# ════════════════════════════════════════════════════════════════════════════
# BANCA D'ITALIA — vista estesa (spesa Abruzzo, regioni, motivo, struttura, durata)
# ════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def bdi_extended():
    p = ".cache/bdi_extended.json"
    return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else None


def chart_abruzzo_spend(ext: dict, yr_range=None) -> go.Figure:
    a = ext["abruzzo"]; anni, sp = a["anni"], a["spesa"]
    if yr_range:
        keep = [i for i, y in enumerate(anni) if yr_range[0] <= y <= yr_range[1]]
        anni, sp = [anni[i] for i in keep], [sp[i] for i in keep]
    fig = go.Figure(go.Bar(x=anni, y=sp, marker_color="#0e6b70",
                           hovertemplate="%{x}: %{y:.0f} M€<extra></extra>"))
    fig.update_yaxes(title="spesa straniera (M€)")
    return _layout(fig, h=320)


def chart_bdi_motivo(ext: dict) -> go.Figure:
    m = ext["motivo_2024"]
    return _layout(go.Figure(go.Pie(labels=list(m), values=list(m.values()), hole=0.55,
                   marker=dict(colors=["#0e6b70", "#38bdf8", "#f59e0b"]))), h=320)


def chart_bdi_struttura(ext: dict) -> go.Figure:
    items = sorted(ext["struttura_2024"].items(), key=lambda x: x[1])
    fig = go.Figure(go.Bar(x=[v for _, v in items], y=[k for k, _ in items], orientation="h",
                           marker_color="#0e6b70", hovertemplate="<b>%{y}</b><br>%{x:,.0f} M€<extra></extra>"))
    fig.update_xaxes(title="spesa nazionale 2024 (M€)")
    return _layout(fig, h=300)
